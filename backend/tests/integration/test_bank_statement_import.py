"""
Integration tests for bank statement Excel import endpoint.

Endpoint: POST /api/finance/bank-statements/import_excel/
The view reads the header from row 2 (hard-coded). Row 1 is treated as a
title/metadata row and skipped. Data rows start from row 3 onward.
"""

import io

import openpyxl
import pytest
from django.core.files.uploadedfile import SimpleUploadedFile

pytestmark = pytest.mark.django_db

# Known column names the view recognises
HEADERS = [
    '凭证号',
    '对方账号',
    '交易时间',
    '借贷标志',
    '对方单位',
    '对方行号',
    '用途',
    '摘要',
    '附言',
    '回单个性化信息',
    '转入金额',
    '转出金额',
    '支付凭证种类',
    '余额',
]

IMPORT_URL = '/api/finance/bank-statements/import_excel/'


def _make_xlsx(rows_by_row: dict[int, list]) -> bytes:
    """Build an in-memory xlsx workbook.

    rows_by_row maps 1-based row numbers to lists of cell values.
    Returns raw bytes suitable for upload.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    for row_num, values in rows_by_row.items():
        for col_num, value in enumerate(values, start=1):
            ws.cell(row=row_num, column=col_num, value=value)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload(client, xlsx_bytes: bytes, filename: str = 'test.xlsx'):
    """POST the xlsx bytes as a multipart file upload."""
    uploaded = SimpleUploadedFile(
        filename,
        xlsx_bytes,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    return client.post(IMPORT_URL, {'file': uploaded}, format='multipart')


# ---------------------------------------------------------------------------
# Test 1 – header on row 2 (auto-detected; row 1 is a title/metadata row)
# ---------------------------------------------------------------------------

def test_import_header_on_row_2(api_client_admin):
    """Header on row 2 is the expected format; data row starts at row 3."""
    xlsx = _make_xlsx({
        1: ['银行流水明细'],  # title / metadata row, ignored
        2: HEADERS,           # header row (auto-detected within first 10 rows)
        3: [
            'V001',           # 凭证号
            '123456789',      # 对方账号
            '2024-01-15 10:00:00',  # 交易时间
            '贷',             # 借贷标志
            '测试客户有限公司',  # 对方单位
            'BANK001',        # 对方行号
            '货款',           # 用途
            '销售货款',        # 摘要
            '',               # 附言
            '',               # 回单个性化信息
            '50000.00',       # 转入金额
            '0',              # 转出金额
            '',               # 支付凭证种类
            '150000.00',      # 余额
        ],
    })

    resp = _upload(api_client_admin, xlsx)
    assert resp.status_code == 200, f"Expected 200 but got {resp.status_code}: {resp.data}"
    data = resp.data
    assert data['success_count'] >= 1
    assert data['credit_total'] >= 50000.0


# ---------------------------------------------------------------------------
# Test 2 – header on row 1 (no title row above it)
# The view auto-detects the header row by scanning the first 10 rows, so a
# header on row 1 is correctly recognised and the following data row imported.
# ---------------------------------------------------------------------------

def test_import_header_on_row_1(api_client_admin):
    """Header on row 1 is auto-detected; the data row on row 2 is imported."""
    xlsx = _make_xlsx({
        1: HEADERS,  # header on row 1 (auto-detected)
        2: [
            'V001', '123456789', '2024-01-15 10:00:00', '贷',
            '测试客户有限公司', 'BANK001', '货款', '销售货款',
            '', '', '50000.00', '0', '', '150000.00',
        ],
    })

    resp = _upload(api_client_admin, xlsx)
    assert resp.status_code == 200, f"Unexpected status: {resp.status_code}: {resp.data}"
    assert resp.data['success_count'] >= 1


# ---------------------------------------------------------------------------
# Test 3 – garbage headers → header auto-detection fails → 400
# The current view scans the first 10 rows for a header containing at least 3
# recognised column names. Garbage headers match nothing, so it rejects the
# upload with HTTP 400 and an explanatory error (instead of silently importing
# zero rows). This is the corrected behaviour.
# ---------------------------------------------------------------------------

def test_import_no_matching_header(api_client_admin):
    """Garbage headers are rejected with 400 (no recognisable header row)."""
    xlsx = _make_xlsx({
        1: ['Title row'],
        2: ['Col_A', 'Col_B', 'Col_C', 'Col_D'],
        3: ['foo', 'bar', 'baz', 'qux'],
    })

    resp = _upload(api_client_admin, xlsx)
    # Header auto-detection requires >= 3 known columns; none match → 400.
    assert resp.status_code == 400, f"Unexpected status {resp.status_code}: {resp.data}"
    assert 'error' in resp.data


# ---------------------------------------------------------------------------
# Test 4 – empty Excel file → no header row found → 400
# An empty sheet has no recognisable header, so the view rejects it with 400
# rather than reporting a successful import of zero rows.
# ---------------------------------------------------------------------------

def test_import_empty_file(api_client_admin):
    """Empty xlsx is rejected with 400 (no recognisable header row)."""
    wb = openpyxl.Workbook()
    # Leave the sheet completely empty
    buf = io.BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()

    resp = _upload(api_client_admin, xlsx_bytes)
    assert resp.status_code == 400, f"Unexpected status {resp.status_code}: {resp.data}"
    assert 'error' in resp.data
