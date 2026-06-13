"""
Views for bank statement import/export.
"""
import io
import re
from decimal import Decimal, InvalidOperation
from datetime import datetime
from django.db import transaction
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser

from apps.core.mixins import SoftDeleteMixin, UserTrackingMixin
from apps.core.permission_mixin import PermissionMixin
from apps.masterdata.models import Supplier, Customer
from .models import AccountPayable, AccountReceivable, Payment
from .bank_statement_models import BankStatement, BankStatementImportLog
from .bank_statement_serializers import (
    BankStatementSerializer,
    BankStatementImportLogSerializer,
    BankStatementMatchSerializer
)


class BankStatementViewSet(PermissionMixin, SoftDeleteMixin, UserTrackingMixin, viewsets.ModelViewSet):

    permission_module = 'finance'
    permission_resource = 'bank_statement'
    """
    ViewSet for BankStatement management.
    """
    queryset = BankStatement.objects.all()
    serializer_class = BankStatementSerializer
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    filterset_fields = ['status', 'transaction_type', 'match_type', 'supplier', 'customer', 'import_batch', 'bank_name']
    search_fields = ['counterparty_name', 'summary', 'purpose', 'postscript', 'bank_name']
    ordering_fields = ['transaction_time', 'debit_amount', 'credit_amount', 'created_at']
    
    def _parse_amount(self, value):
        """Parse amount string to Decimal."""
        if not value:
            return Decimal('0')
        # Remove commas and spaces
        value = str(value).replace(',', '').replace(' ', '').strip()
        if not value:
            return Decimal('0')
        try:
            return Decimal(value)
        except InvalidOperation:
            return Decimal('0')
    
    def _parse_datetime(self, value):
        """Parse datetime string. Handles ICBC/华夏银行/建行 etc. export formats."""
        if not value:
            return None
        if isinstance(value, datetime):
            return value

        from datetime import date as date_type
        if isinstance(value, date_type):
            return datetime(value.year, value.month, value.day)

        value = str(value).strip()

        formats = [
            '%Y-%m-%d %H:%M:%S',
            '%Y/%m/%d %H:%M:%S',
            '%Y-%m-%d %H:%M',
            '%Y/%m/%d %H:%M',
            '%Y-%m-%d',
            '%Y/%m/%d',
            '%Y.%m.%d %H:%M:%S',
            '%Y.%m.%d',
            '%Y%m%d%H%M%S',
            '%Y%m%d',
        ]

        for fmt in formats:
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                continue
        return None

    def _load_xls_sheet(self, file_bytes):
        """Load .xls (BIFF) file via xlrd and wrap as openpyxl-compatible sheet."""
        import xlrd

        class _XlrdCellAdapter:
            def __init__(self, value):
                self.value = value

        class _XlrdSheetAdapter:
            def __init__(self, xlrd_sheet):
                self._s = xlrd_sheet
                self.max_row = xlrd_sheet.nrows
                self.max_column = xlrd_sheet.ncols

            def cell(self, row, column):
                cell_obj = self._s.cell(row - 1, column - 1)
                val = cell_obj.value
                # xlrd type 0 = empty, 6 = blank → treat as None
                if cell_obj.ctype in (0, 6) or val == '':
                    val = None
                return _XlrdCellAdapter(val)

        xls_book = xlrd.open_workbook(file_contents=file_bytes)
        return _XlrdSheetAdapter(xls_book.sheet_by_index(0))

    @action(detail=False, methods=['post'])
    def import_excel(self, request):
        """
        Import bank statement from Excel file.
        Supports the bank format with headers on row 2.
        """
        file = request.FILES.get('file')
        if not file:
            return Response(
                {'error': '请上传文件'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        file_name = file.name
        if not file_name.endswith(('.xlsx', '.xls')):
            return Response(
                {'error': '只支持Excel文件格式(.xlsx, .xls)'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        bank_name = request.data.get('bank_name', '')
        bank_account = request.data.get('bank_account', '')

        try:
            import openpyxl

            # Generate batch number (microseconds for uniqueness)
            batch_no = f"BS{timezone.now().strftime('%Y%m%d%H%M%S%f')}"

            # Read Excel file — support both .xlsx (openpyxl) and .xls (xlrd)
            file_bytes = file.read()
            if file_name.endswith('.xls') and not file_name.endswith('.xlsx'):
                sheet = self._load_xls_sheet(file_bytes)
            else:
                wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
                sheet = wb.active

            # Column aliases — different banks export slightly different headers.
            # Key = canonical name used by the rest of this function; value = list
            # of accepted header strings (after strip()). The canonical name itself
            # is always accepted; aliases are extras.
            # Covered banks: 工商银行(ICBC), 华夏银行, 建设银行, 招商银行 etc.
            COLUMN_ALIASES = {
                '凭证号': [
                    '凭证号', '凭证号码', '业务流水号', '交易流水号',
                    '网银交易流水号', '银行流水号', '流水号', '核心流水号',
                ],
                '对方账号': [
                    '对方账号', '对方账户', '对方账户号码', '对方卡号',
                    '收(付)方账号', '收付方账号', '对方帐号',
                ],
                '交易时间': [
                    '交易时间', '交易日期', '记账日期', '交易日期时间',
                    '入账日期', '日期', '记帐日期', '交易时间/入账时间',
                ],
                '借贷标志': [
                    '借贷标志', '借贷', '收付标志', '借/贷', '借贷方向',
                    '收/付', '收付', '交易类型', '借贷类型',
                ],
                '对方单位': [
                    '对方单位', '对方户名', '对方账户名称', '交易对方',
                    '对方名称', '收(付)方名称', '收付方名称', '收(付)方户名',
                    '对方账户名', '对方客户名称',
                ],
                '对方行号': [
                    '对方行号', '对方开户行', '对方银行',
                    '对方开户行名称', '收(付)方开户行', '对方行名', '对方开户行名',
                ],
                '用途': ['用途', '交易用途', '交易描述'],
                '摘要': ['摘要', '交易摘要', '业务摘要', '摘要码'],
                '附言': ['附言', '备注', '交易备注', '交易附言'],
                '回单个性化信息': ['回单个性化信息', '回单信息'],
                '转入金额': [
                    '转入金额', '收入金额', '贷方发生额', '贷方金额', '收入',
                    '存入金额', '收方金额', '收入(贷方)',
                ],
                '转出金额': [
                    '转出金额', '支出金额', '借方发生额', '借方金额', '支出',
                    '付出金额', '付方金额', '支出(借方)',
                ],
                '交易金额': [
                    '交易金额', '金额', '发生额',
                ],
                '支付凭证种类': ['支付凭证种类', '凭证种类'],
                '余额': ['余额', '账户余额', '账面余额'],
            }

            # Build reverse lookup: alias (lower-cased, stripped) -> canonical name
            alias_to_canonical = {}
            for canonical, aliases in COLUMN_ALIASES.items():
                for alias in aliases:
                    alias_to_canonical[alias.strip()] = canonical

            def _normalize(v):
                """Read a cell, coerce to str, strip whitespace + zero-width chars."""
                if v is None:
                    return ''
                return str(v).replace('　', '').replace('﻿', '').strip()

            # Auto-detect the header row: scan the first 10 rows, pick the one
            # that contains the most known column headers (must have at least 3).
            header_row = None
            best_hit_count = 0
            scan_limit = min(10, sheet.max_row)
            for candidate_row in range(1, scan_limit + 1):
                hit = 0
                for col in range(1, sheet.max_column + 1):
                    cell_str = _normalize(sheet.cell(row=candidate_row, column=col).value)
                    if cell_str in alias_to_canonical:
                        hit += 1
                if hit > best_hit_count:
                    best_hit_count = hit
                    header_row = candidate_row

            if header_row is None or best_hit_count < 3:
                return Response(
                    {
                        'error': (
                            '未能识别银行流水表头。请确保 Excel 前 10 行内存在表头行，'
                            '且至少包含 3 个可识别列名。'
                            '支持的列名包括：交易时间/记账日期/入账日期、'
                            '转入金额/贷方发生额/收入、转出金额/借方发生额/支出、'
                            '交易金额/发生额、对方单位/对方户名/对方名称 等。'
                            '已兼容工商银行、华夏银行等主流银行导出格式。'
                        )
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Build col_map (canonical name -> 1-based column index)
            col_map = {canonical: None for canonical in COLUMN_ALIASES}
            for col in range(1, sheet.max_column + 1):
                cell_str = _normalize(sheet.cell(row=header_row, column=col).value)
                canonical = alias_to_canonical.get(cell_str)
                if canonical and col_map.get(canonical) is None:
                    col_map[canonical] = col

            # Detect split date/time columns (华夏银行: '交易日期' + '交易时间' as two cols)
            # If header has both '交易日期' and '交易时间' as separate columns,
            # store the time-only column index for merging later.
            time_only_col = None
            header_texts = []
            for col in range(1, sheet.max_column + 1):
                header_texts.append(_normalize(sheet.cell(row=header_row, column=col).value))
            if '交易日期' in header_texts and '交易时间' in header_texts:
                date_col_idx = header_texts.index('交易日期') + 1
                time_col_idx = header_texts.index('交易时间') + 1
                if date_col_idx != time_col_idx:
                    col_map['交易时间'] = date_col_idx
                    time_only_col = time_col_idx

            # Detect account holder name from header rows (for same-name transfer detection)
            account_holder = ''
            for r in range(1, header_row):
                for c in range(1, min(6, sheet.max_column + 1)):
                    cell_val = _normalize(sheet.cell(row=r, column=c).value)
                    if cell_val in ('户名：', '户名:', '账户名称：', '账户名称:'):
                        next_val = _normalize(sheet.cell(row=r, column=c + 1).value)
                        if next_val:
                            account_holder = next_val
                            break
                if account_holder:
                    break

            # Keywords to filter out internal/non-business transactions
            SKIP_KEYWORDS = [
                # 工资相关
                '工资', '薪资', '薪酬', '社保', '公积金', '个税', '代扣', '代缴',
                # 税费相关
                '税款', '税费', '增值税', '附加税', '所得税', '印花税', '税收收缴', '国库',
                # 内部往来
                '内部', '往来', '调拨', '备用金', '借款', '还款',
                # 银行费用
                '利息', '手续费', '银行费', '账户管理费',
                # 公用事业费
                '电费', '水费', '房租', '物业费', '供电局', '自来水', '燃气',
                # 其他
                '退款', '冲账', '更正', '报销费用', '差旅费'
            ]
            
            def is_internal_transaction(name, purpose, summary, postscript):
                """Check if transaction is internal/non-business that should be skipped."""
                all_text = f"{name or ''}{purpose or ''}{summary or ''}{postscript or ''}"
                for keyword in SKIP_KEYWORDS:
                    if keyword in all_text:
                        return True
                # Same-name transfer: counterparty is the account holder itself
                if account_holder and name and account_holder in str(name):
                    return True
                return False
            
            # Import records
            records = []
            errors = []
            skipped_count = 0
            duplicate_count = 0
            success_count = 0
            matched_count = 0
            debit_total = Decimal('0')
            credit_total = Decimal('0')
            
            # Track processed transactions within this import to avoid in-file duplicates
            processed_transactions = set()
            
            with transaction.atomic():
                for row in range(header_row + 1, sheet.max_row + 1):
                    try:
                        # Get cell values
                        def get_cell(col_name):
                            col_idx = col_map.get(col_name)
                            if col_idx:
                                return sheet.cell(row=row, column=col_idx).value
                            return None
                        
                        counterparty_name = get_cell('对方单位')
                        transaction_time = get_cell('交易时间')
                        
                        # Skip empty rows
                        if not counterparty_name and not transaction_time:
                            continue
                        
                        purpose = str(get_cell('用途') or '')
                        summary = str(get_cell('摘要') or '')
                        postscript = str(get_cell('附言') or '')
                        
                        # Skip internal/non-business transactions
                        if is_internal_transaction(counterparty_name, purpose, summary, postscript):
                            skipped_count += 1
                            continue
                        
                        voucher_no = str(get_cell('凭证号') or '')

                        # Create a unique key for duplicate detection
                        # Use voucher_no + transaction_time + counterparty_name + amounts
                        credit_amount_raw = self._parse_amount(get_cell('转入金额'))
                        debit_amount_raw = self._parse_amount(get_cell('转出金额'))

                        # Single-column amount mode (工商银行/华夏银行 etc.)
                        # When no separate credit/debit columns, derive from '交易金额'
                        if col_map.get('转入金额') is None and col_map.get('转出金额') is None:
                            raw_amount = self._parse_amount(get_cell('交易金额'))
                            debit_credit_flag = str(get_cell('借贷标志') or '').strip()
                            if debit_credit_flag in ('借', '付', '支出', 'D', 'DR'):
                                debit_amount_raw = abs(raw_amount)
                                credit_amount_raw = Decimal('0')
                            elif debit_credit_flag in ('贷', '收', '收入', 'C', 'CR'):
                                credit_amount_raw = abs(raw_amount)
                                debit_amount_raw = Decimal('0')
                            elif raw_amount < 0:
                                debit_amount_raw = abs(raw_amount)
                                credit_amount_raw = Decimal('0')
                            else:
                                credit_amount_raw = abs(raw_amount)
                                debit_amount_raw = Decimal('0')

                        parsed_time_raw = self._parse_datetime(get_cell('交易时间'))
                        # Merge separate time column if detected (华夏银行: 交易日期 + 交易时间)
                        if time_only_col and parsed_time_raw:
                            time_val = str(sheet.cell(row=row, column=time_only_col).value or '').strip()
                            if time_val and ':' in time_val:
                                try:
                                    parts = time_val.split(':')
                                    parsed_time_raw = parsed_time_raw.replace(
                                        hour=int(parts[0]), minute=int(parts[1]),
                                        second=int(parts[2]) if len(parts) > 2 else 0
                                    )
                                except (ValueError, IndexError):
                                    pass
                        time_str = parsed_time_raw.strftime('%Y-%m-%d %H:%M:%S') if parsed_time_raw else ''
                        
                        duplicate_key = f"{voucher_no}|{time_str}|{counterparty_name}|{credit_amount_raw}|{debit_amount_raw}"
                        
                        # Check for in-file duplicate
                        if duplicate_key in processed_transactions:
                            duplicate_count += 1
                            continue
                        processed_transactions.add(duplicate_key)
                        
                        # Check for existing record in database
                        # Treat all-zero or all-same-digit voucher numbers as placeholder (e.g. ICBC uses '000000000')
                        effective_voucher = voucher_no
                        if effective_voucher and len(set(effective_voucher)) <= 1:
                            effective_voucher = ''
                        if effective_voucher:
                            existing = BankStatement.objects.filter(
                                voucher_no=effective_voucher,
                                is_deleted=False
                            ).first()
                            if existing:
                                duplicate_count += 1
                                continue
                        elif parsed_time_raw:
                            # Check by time + counterparty + amount for records without valid voucher_no
                            existing = BankStatement.objects.filter(
                                transaction_time=parsed_time_raw,
                                counterparty_name=str(counterparty_name or ''),
                                credit_amount=credit_amount_raw,
                                debit_amount=debit_amount_raw,
                                is_deleted=False
                            ).first()
                            if existing:
                                duplicate_count += 1
                                continue
                        
                        # Parse transaction type from flag or infer from amounts
                        debit_credit = str(get_cell('借贷标志') or '').strip()
                        if debit_credit in ('借', '付', '支出', 'D', 'DR', '付款'):
                            transaction_type = 'DEBIT'
                        elif debit_credit in ('贷', '收', '收入', 'C', 'CR', '收款'):
                            transaction_type = 'CREDIT'
                        elif debit_amount_raw > 0:
                            transaction_type = 'DEBIT'
                        elif credit_amount_raw > 0:
                            transaction_type = 'CREDIT'
                        else:
                            transaction_type = 'CREDIT'

                        # Use already parsed amounts
                        credit_amount = credit_amount_raw
                        debit_amount = debit_amount_raw
                        balance = self._parse_amount(get_cell('余额'))
                        
                        # Use already parsed time
                        parsed_time = parsed_time_raw if parsed_time_raw else timezone.now()
                        
                        # Try to match supplier/customer BEFORE creating record
                        supplier = None
                        customer = None
                        match_confidence = Decimal('0')
                        
                        # Create a temporary object for matching
                        temp_statement = BankStatement(
                            counterparty_name=str(counterparty_name or ''),
                            transaction_type=transaction_type
                        )
                        
                        if transaction_type == 'DEBIT':
                            # Payment to supplier
                            supplier, confidence = temp_statement.auto_match_supplier()
                            if supplier:
                                match_confidence = Decimal(str(confidence))
                        else:
                            # Receipt from customer
                            customer, confidence = temp_statement.auto_match_customer()
                            if customer:
                                match_confidence = Decimal(str(confidence))
                        
                        # Create record (import all records, mark unmatched as PENDING)
                        statement = BankStatement(
                            import_batch=batch_no,
                            source_file=file_name,
                            bank_name=bank_name,
                            bank_account=bank_account,
                            voucher_no=voucher_no,
                            counterparty_account=str(get_cell('对方账号') or ''),
                            transaction_time=parsed_time,
                            transaction_type=transaction_type,
                            counterparty_name=str(counterparty_name or ''),
                            counterparty_bank=str(get_cell('对方行号') or ''),
                            purpose=purpose,
                            summary=summary,
                            postscript=postscript,
                            receipt_info=str(get_cell('回单个性化信息') or ''),
                            credit_amount=credit_amount,
                            debit_amount=debit_amount,
                            balance=balance,
                            payment_voucher_type=str(get_cell('支付凭证种类') or ''),
                            created_by=request.user,
                            supplier=supplier,
                            customer=customer,
                            match_confidence=match_confidence,
                            match_type='AP' if supplier else ('AR' if customer else None),
                            status='MATCHED' if match_confidence >= 70 else 'PENDING'
                        )
                        
                        if match_confidence >= 70:
                                    matched_count += 1
                        
                        statement.save()
                        records.append(statement)
                        success_count += 1
                        
                        debit_total += debit_amount
                        credit_total += credit_amount
                        
                    except Exception as e:
                        errors.append({
                            'row': row,
                            'error': str(e)
                        })
                
                # Create import log
                import_log = BankStatementImportLog.objects.create(
                    batch_no=batch_no,
                    file_name=file_name,
                    bank_name=bank_name,
                    bank_account=bank_account,
                    total_count=success_count + skipped_count + duplicate_count + len(errors),
                    success_count=success_count,
                    error_count=len(errors),
                    matched_count=matched_count,
                    debit_total=debit_total,
                    credit_total=credit_total,
                    error_details=errors,
                    notes=f'表头行={header_row}，跳过{skipped_count}条非业务流水，{duplicate_count}条重复记录',
                    created_by=request.user
                )

            return Response({
                'batch_no': batch_no,
                'header_row': header_row,
                'success_count': success_count,
                'skipped_count': skipped_count,
                'duplicate_count': duplicate_count,
                'error_count': len(errors),
                'matched_count': matched_count,
                'debit_total': float(debit_total),
                'credit_total': float(credit_total),
                'errors': errors[:10],  # Return first 10 errors
                'import_log_id': import_log.id
            })
            
        except Exception as e:
            return Response(
                {'error': f'导入失败: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'])
    def bank_names(self, request):
        """返回所有已导入的银行名称列表（用于筛选下拉）"""
        names = (
            BankStatement.objects.filter(is_deleted=False)
            .exclude(bank_name='')
            .values_list('bank_name', flat=True)
            .distinct()
            .order_by('bank_name')
        )
        return Response(list(names))

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """
        Export bank statements to Excel.
        """
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            
            # Get filters
            status_filter = request.query_params.get('status')
            transaction_type = request.query_params.get('transaction_type')
            supplier_id = request.query_params.get('supplier')
            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')
            
            queryset = self.get_queryset()
            
            if status_filter:
                queryset = queryset.filter(status=status_filter)
            if transaction_type:
                queryset = queryset.filter(transaction_type=transaction_type)
            if supplier_id:
                queryset = queryset.filter(supplier_id=supplier_id)
            if start_date:
                queryset = queryset.filter(transaction_time__date__gte=start_date)
            if end_date:
                queryset = queryset.filter(transaction_time__date__lte=end_date)
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = '银行流水'
            
            # Headers
            headers = [
                '交易时间', '借贷标志', '对方单位', '用途', '摘要', '附言',
                '转入金额', '转出金额', '余额', '匹配状态', '匹配供应商/客户',
                '关联单号', '导入批次'
            ]
            
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Data rows
            for row, statement in enumerate(queryset, 2):
                ws.cell(row=row, column=1, value=statement.transaction_time.strftime('%Y-%m-%d %H:%M:%S'))
                ws.cell(row=row, column=2, value='借' if statement.transaction_type == 'DEBIT' else '贷')
                ws.cell(row=row, column=3, value=statement.counterparty_name)
                ws.cell(row=row, column=4, value=statement.purpose)
                ws.cell(row=row, column=5, value=statement.summary)
                ws.cell(row=row, column=6, value=statement.postscript)
                ws.cell(row=row, column=7, value=float(statement.credit_amount))
                ws.cell(row=row, column=8, value=float(statement.debit_amount))
                ws.cell(row=row, column=9, value=float(statement.balance) if statement.balance else '')
                ws.cell(row=row, column=10, value=statement.get_status_display())
                
                # Matched entity
                if statement.supplier:
                    ws.cell(row=row, column=11, value=f'供应商: {statement.supplier.name}')
                elif statement.customer:
                    ws.cell(row=row, column=11, value=f'客户: {statement.customer.name}')
                else:
                    ws.cell(row=row, column=11, value='')
                
                # Related document
                if statement.related_ap:
                    ws.cell(row=row, column=12, value=statement.related_ap.ap_no)
                elif statement.related_ar:
                    ws.cell(row=row, column=12, value=statement.related_ar.ar_no)
                else:
                    ws.cell(row=row, column=12, value='')
                
                ws.cell(row=row, column=13, value=statement.import_batch)
            
            # Set column widths
            column_widths = [20, 8, 30, 12, 20, 30, 12, 12, 12, 10, 25, 15, 18]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(col)].width = width
            
            # Create response
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="bank_statements_{timezone.now().strftime("%Y%m%d")}.xlsx"'
            
            return response
            
        except Exception as e:
            return Response(
                {'error': f'导出失败: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'])
    def match(self, request, pk=None):
        """
        Manually match a bank statement to supplier/customer and AP/AR.
        """
        statement = self.get_object()
        serializer = BankStatementMatchSerializer(data=request.data)
        
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        data = serializer.validated_data
        match_type = data.get('match_type')

        with transaction.atomic():
            statement.match_type = match_type
            statement.match_notes = data.get('notes', '')

            if match_type == 'AP':
                supplier_id = data.get('supplier_id')
                if supplier_id:
                    statement.supplier = Supplier.objects.get(id=supplier_id)

                ap_id = data.get('ap_id')
                if ap_id:
                    statement.related_ap = AccountPayable.objects.get(id=ap_id)

                statement.customer = None
                statement.related_ar = None

            elif match_type == 'AR':
                customer_id = data.get('customer_id')
                if customer_id:
                    statement.customer = Customer.objects.get(id=customer_id)

                ar_id = data.get('ar_id')
                if ar_id:
                    statement.related_ar = AccountReceivable.objects.get(id=ar_id)

                statement.supplier = None
                statement.related_ap = None

            # Set project if provided
            project_id = data.get('project_id')
            if project_id:
                from apps.projects.models import Project
                statement.project = Project.objects.get(id=project_id)
            elif not statement.project:
                # Try auto-matching project
                project = statement.auto_match_project()
                if project:
                    statement.project = project

            statement.status = 'MATCHED'
            statement.match_confidence = Decimal('100')
            statement.save()

            # 命中具体 AR/AP 时落收/付款，使 amount_paid 与对账口径一致（不超核销）
            self._apply_statement_payment(statement, request.user)

        return Response(BankStatementSerializer(statement).data)

    def _apply_statement_payment(self, statement, user):
        """匹配到具体 AR/AP 时生成 Payment，更新已收/已付金额并重算状态。"""
        amount = statement.amount or Decimal('0')
        if amount <= 0:
            return

        # 用 notes 携带流水 id 做幂等标记，避免重复匹配重复记账
        marker = f'[BS#{statement.id}]'

        if statement.match_type == 'AR' and statement.related_ar_id:
            ar = AccountReceivable.objects.select_for_update().get(pk=statement.related_ar_id)
            if Payment.objects.filter(ar=ar, notes__contains=marker).exists():
                return
            remaining = ar.amount_due - ar.amount_paid
            pay_amount = min(amount, remaining)
            if pay_amount <= 0:
                return
            Payment.objects.create(
                payment_type='AR',
                ar=ar,
                payment_date=statement.transaction_time.date(),
                payment_method='BANK_TRANSFER',
                amount=pay_amount,
                currency=ar.currency,
                exchange_rate=ar.exchange_rate,
                notes=f'{marker} 银行流水匹配',
                created_by=user,
                updated_by=user,
            )
            ar.refresh_from_db(fields=['amount_paid'])
            ar.status = 'PAID' if ar.amount_paid >= ar.amount_due else 'PARTIAL'
            ar.save(update_fields=['status'])

        elif statement.match_type == 'AP' and statement.related_ap_id:
            ap = AccountPayable.objects.select_for_update().get(pk=statement.related_ap_id)
            if Payment.objects.filter(ap=ap, notes__contains=marker).exists():
                return
            remaining = ap.amount_due - ap.amount_paid
            pay_amount = min(amount, remaining)
            if pay_amount <= 0:
                return
            Payment.objects.create(
                payment_type='AP',
                ap=ap,
                payment_date=statement.transaction_time.date(),
                payment_method='BANK_TRANSFER',
                amount=pay_amount,
                currency=ap.currency,
                exchange_rate=ap.exchange_rate,
                notes=f'{marker} 银行流水匹配',
                created_by=user,
                updated_by=user,
            )
            ap.refresh_from_db(fields=['amount_paid'])
            ap.status = 'PAID' if ap.amount_paid >= ap.amount_due else 'PARTIAL'
            ap.save(update_fields=['status'])
    
    @action(detail=True, methods=['post'])
    def ignore(self, request, pk=None):
        """Mark a bank statement as ignored."""
        statement = self.get_object()
        statement.status = 'IGNORED'
        statement.match_notes = request.data.get('notes', '手动忽略')
        statement.save()
        return Response(BankStatementSerializer(statement).data)
    
    @action(detail=False, methods=['post'])
    def auto_match_all(self, request):
        """
        Auto-match all pending bank statements.
        Also attempts to match projects based on customer/supplier.
        """
        batch = request.data.get('import_batch')
        ids = request.data.get('ids')
        queryset = self.get_queryset().filter(status='PENDING')

        if batch:
            queryset = queryset.filter(import_batch=batch)

        # 支持仅对选中流水执行匹配（前端批量/单条匹配传入 ids）
        if ids:
            queryset = queryset.filter(id__in=ids)

        matched_count = 0
        project_matched_count = 0
        
        with transaction.atomic():
            for statement in queryset:
                matched = False
                
                if statement.transaction_type == 'DEBIT':
                    supplier, confidence = statement.auto_match_supplier()
                    if supplier and confidence >= 70:
                        statement.supplier = supplier
                        statement.match_confidence = Decimal(str(confidence))
                        statement.match_type = 'AP'
                        statement.status = 'MATCHED'
                        matched = True
                else:
                    customer, confidence = statement.auto_match_customer()
                    if customer and confidence >= 70:
                        statement.customer = customer
                        statement.match_confidence = Decimal(str(confidence))
                        statement.match_type = 'AR'
                        statement.status = 'MATCHED'
                        matched = True
                
                # Try to match project based on customer/supplier
                if matched:
                    project = statement.auto_match_project()
                    if project:
                        statement.project = project
                        project_matched_count += 1

                    statement.save()
                    matched_count += 1

                    # 尝试匹配付款计划
                    if statement.transaction_type == 'CREDIT' and statement.customer:
                        # 收款匹配销售付款计划
                        self._try_match_payment_schedule(statement)
                    elif statement.transaction_type == 'DEBIT' and statement.supplier:
                        # 付款匹配采购付款计划
                        self._try_match_purchase_payment_schedule(statement)
        
        return Response({
            'matched_count': matched_count,
            'project_matched_count': project_matched_count,
            'message': f'成功匹配 {matched_count} 条记录，其中 {project_matched_count} 条关联到项目'
        })
    
    def _try_match_payment_schedule(self, statement):
        """
        尝试将银行流水匹配到付款计划。
        根据客户和金额找到最匹配的付款计划。
        """
        from apps.finance.models import PaymentSchedule
        from decimal import Decimal
        
        customer = statement.customer
        amount = statement.credit_amount or Decimal('0')
        
        if not customer or amount <= 0:
            return None
        
        # 查找该客户待收款的付款计划
        # 优先匹配金额完全相等的
        exact_match = PaymentSchedule.objects.filter(
            sales_order__customer=customer,
            status__in=['PENDING', 'PARTIAL'],
            amount_due=amount,
            is_deleted=False
        ).order_by('due_date').first()
        
        if exact_match:
            exact_match.amount_paid += amount
            if exact_match.amount_paid >= exact_match.amount_due:
                exact_match.status = 'PAID'
                exact_match.actual_paid_date = statement.transaction_time.date()
                exact_match.reminder_status = 'COLLECTED'
            else:
                exact_match.status = 'PARTIAL'
            exact_match.save()
            exact_match.bank_statements.add(statement)
            
            # 更新流水的项目关联
            if exact_match.project:
                statement.project = exact_match.project
                statement.save()
            
            return exact_match
        
        # 如果没有完全匹配，查找剩余金额接近的
        pending_schedules = PaymentSchedule.objects.filter(
            sales_order__customer=customer,
            status__in=['PENDING', 'PARTIAL'],
            is_deleted=False
        ).order_by('due_date')
        
        for schedule in pending_schedules:
            remaining = schedule.amount_due - schedule.amount_paid
            # 如果付款金额在剩余金额的10%范围内，认为匹配
            if abs(remaining - amount) <= remaining * Decimal('0.1'):
                schedule.amount_paid += amount
                if schedule.amount_paid >= schedule.amount_due:
                    schedule.status = 'PAID'
                    schedule.actual_paid_date = statement.transaction_time.date()
                    schedule.reminder_status = 'COLLECTED'
                else:
                    schedule.status = 'PARTIAL'
                schedule.save()
                schedule.bank_statements.add(statement)
                
                if schedule.project:
                    statement.project = schedule.project
                    statement.save()
                
                return schedule
        
        return None
    
    def _try_match_purchase_payment_schedule(self, statement):
        """
        尝试将银行流水（付款）匹配到采购付款计划。
        根据供应商和金额找到最匹配的采购付款计划。
        """
        from apps.finance.models import PurchasePaymentSchedule
        from decimal import Decimal
        
        supplier = statement.supplier
        amount = statement.debit_amount or Decimal('0')
        
        if not supplier or amount <= 0:
            return None
        
        # 查找该供应商待付款的付款计划
        # 优先匹配金额完全相等的
        exact_match = PurchasePaymentSchedule.objects.filter(
            purchase_order__supplier=supplier,
            status__in=['PENDING', 'PARTIAL'],
            amount_due=amount,
            is_deleted=False
        ).order_by('due_date').first()
        
        if exact_match:
            exact_match.amount_paid += amount
            if exact_match.amount_paid >= exact_match.amount_due:
                exact_match.status = 'PAID'
                exact_match.actual_paid_date = statement.transaction_time.date()
                exact_match.reminder_status = 'PAID'
            else:
                exact_match.status = 'PARTIAL'
            exact_match.save()
            exact_match.bank_statements.add(statement)
            
            # 更新流水的项目关联
            if exact_match.project:
                statement.project = exact_match.project
                statement.save()
            
            return exact_match
        
        # 如果没有完全匹配，查找剩余金额接近的
        pending_schedules = PurchasePaymentSchedule.objects.filter(
            purchase_order__supplier=supplier,
            status__in=['PENDING', 'PARTIAL'],
            is_deleted=False
        ).order_by('due_date')
        
        for schedule in pending_schedules:
            remaining = schedule.amount_due - schedule.amount_paid
            # 如果付款金额在剩余金额的10%范围内，认为匹配
            if abs(remaining - amount) <= remaining * Decimal('0.1'):
                schedule.amount_paid += amount
                if schedule.amount_paid >= schedule.amount_due:
                    schedule.status = 'PAID'
                    schedule.actual_paid_date = statement.transaction_time.date()
                    schedule.reminder_status = 'PAID'
                else:
                    schedule.status = 'PARTIAL'
                schedule.save()
                schedule.bank_statements.add(statement)
                
                if schedule.project:
                    statement.project = schedule.project
                    statement.save()
                
                return schedule
        
        return None
    
    @action(detail=False, methods=['get'])
    def summary_by_supplier(self, request):
        """
        Get summary of bank statements grouped by supplier.
        """
        from django.db.models import Sum, Count
        
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        queryset = self.get_queryset().filter(
            transaction_type='DEBIT',
            supplier__isnull=False
        )
        
        if start_date:
            queryset = queryset.filter(transaction_time__date__gte=start_date)
        if end_date:
            queryset = queryset.filter(transaction_time__date__lte=end_date)
        
        summary = queryset.values(
            'supplier__id', 'supplier__name'
        ).annotate(
            total_amount=Sum('debit_amount'),
            transaction_count=Count('id')
        ).order_by('-total_amount')
        
        return Response(list(summary))
    
    @action(detail=False, methods=['get'])
    def unmatched_payments(self, request):
        """
        Get unmatched payments (debit transactions without AP match).
        """
        supplier_id = request.query_params.get('supplier_id')
        
        queryset = self.get_queryset().filter(
            transaction_type='DEBIT',
            status='PENDING'
        )
        
        if supplier_id:
            queryset = queryset.filter(supplier_id=supplier_id)
        
        serializer = self.get_serializer(queryset[:50], many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def import_logs(self, request):
        """Get import logs."""
        logs = BankStatementImportLog.objects.all()[:20]
        serializer = BankStatementImportLogSerializer(logs, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['post'])
    def bulk_delete(self, request):
        """Bulk delete bank statements."""
        ids = request.data.get('ids', [])
        if not ids:
            return Response({'error': '请选择要删除的记录'}, status=status.HTTP_400_BAD_REQUEST)
        
        from django.utils import timezone
        
        deleted_count = BankStatement.objects.filter(id__in=ids, is_deleted=False).update(
            is_deleted=True,
            deleted_at=timezone.now()
        )
        
        return Response({
            'success': True,
            'deleted_count': deleted_count
        })
    
    @action(detail=False, methods=['post'])
    def set_project(self, request):
        """
        批量设置银行流水的关联项目。
        用于将银行流水关联到项目进行成本核算。
        """
        statement_ids = request.data.get('statement_ids', [])
        project_id = request.data.get('project_id')
        
        if not statement_ids:
            return Response({'error': '请选择要关联的流水记录'}, status=status.HTTP_400_BAD_REQUEST)
        
        if not project_id:
            return Response({'error': '请选择项目'}, status=status.HTTP_400_BAD_REQUEST)
        
        from apps.projects.models import Project
        
        try:
            project = Project.objects.get(id=project_id, is_deleted=False)
        except Project.DoesNotExist:
            return Response({'error': '项目不存在'}, status=status.HTTP_400_BAD_REQUEST)
        
        updated_count = BankStatement.objects.filter(id__in=statement_ids).update(project=project)
        
        return Response({
            'success': True,
            'updated_count': updated_count,
            'project_code': project.code,
            'project_name': project.name
        })


class BankStatementImportLogViewSet(PermissionMixin, viewsets.ReadOnlyModelViewSet):

    permission_module = 'finance'
    permission_resource = 'bank_statement_import_log'
    """
    ViewSet for BankStatementImportLog (read-only).
    """
    queryset = BankStatementImportLog.objects.all()
    serializer_class = BankStatementImportLogSerializer
    filterset_fields = ['batch_no']
    search_fields = ['file_name', 'batch_no']
    ordering_fields = ['import_time', 'total_count']

