"""
Views for finance app.
"""
import logging
from decimal import Decimal
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.utils import timezone
from django.db import transaction
from django.db.models import Q, Sum, F, Count
from apps.core.mixins import SoftDeleteMixin, UserTrackingMixin
from apps.core.permission_mixin import PermissionMixin
from apps.core.workflow.mixins import WorkflowEnforcementMixin
from apps.projects.models import Project

logger = logging.getLogger(__name__)
from .models import (
    Currency, ExchangeRateHistory, Expense, Invoice,
    AccountReceivable, AccountPayable, Payment, PaymentSchedule, PurchasePaymentSchedule,
    SharedExpense, SharedExpenseAllocation, PaymentRequest
)
from .serializers import (
    CurrencySerializer, ExchangeRateHistorySerializer,
    ExpenseSerializer, InvoiceSerializer, AccountReceivableSerializer,
    AccountPayableSerializer, PaymentSerializer, PaymentScheduleSerializer,
    PurchasePaymentScheduleSerializer, SharedExpenseSerializer, SharedExpenseAllocationSerializer,
    PaymentRequestSerializer
)


def _normalize_payment_method(method):
    """归一化前端付款方式到 Payment.PAYMENT_METHOD_CHOICES。

    前端使用 BANK/CASH/CHECK，模型枚举为 CASH/BANK_TRANSFER/CREDIT_CARD/CHECK/OTHER。
    """
    if not method:
        return 'BANK_TRANSFER'
    mapping = {
        'BANK': 'BANK_TRANSFER',
        'BANK_TRANSFER': 'BANK_TRANSFER',
        'CASH': 'CASH',
        'CHECK': 'CHECK',
        'CREDIT_CARD': 'CREDIT_CARD',
        'OTHER': 'OTHER',
    }
    return mapping.get(str(method).upper(), 'OTHER')


class CurrencyViewSet(PermissionMixin, SoftDeleteMixin, UserTrackingMixin, viewsets.ModelViewSet):
    """
    ViewSet for Currency management.
    """
    queryset = Currency.objects.all()
    serializer_class = CurrencySerializer
    filterset_fields = ['is_active', 'is_base_currency']
    search_fields = ['code', 'name']

    permission_module = 'finance'
    permission_resource = 'currency'
    
    @action(detail=True, methods=['post'])
    def update_rate(self, request, pk=None):
        """Update exchange rate and record history."""
        currency = self.get_object()
        new_rate = request.data.get('exchange_rate')
        effective_date = request.data.get('effective_date', timezone.now().date())
        
        if not new_rate:
            return Response(
                {'error': '请提供汇率'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Record history
        ExchangeRateHistory.objects.create(
            currency=currency,
            exchange_rate=new_rate,
            effective_date=effective_date
        )
        
        # Update current rate
        currency.exchange_rate = new_rate
        currency.save()
        
        return Response(CurrencySerializer(currency).data)
    
    @action(detail=False, methods=['get'])
    def rate_history(self, request):
        """Get exchange rate history."""
        currency_code = request.query_params.get('currency')
        days = int(request.query_params.get('days', 30))
        
        from datetime import timedelta
        start_date = timezone.now().date() - timedelta(days=days)
        
        history = ExchangeRateHistory.objects.filter(
            effective_date__gte=start_date
        )
        
        if currency_code:
            history = history.filter(currency__code=currency_code)
        
        serializer = ExchangeRateHistorySerializer(history, many=True)
        return Response(serializer.data)


class PaymentViewSet(PermissionMixin, SoftDeleteMixin, UserTrackingMixin, viewsets.ModelViewSet):
    permission_module = 'finance'
    permission_resource = 'payment'
    """
    ViewSet for Payment management.
    财务敏感数据 - 仅财务人员和管理层可访问
    """
    queryset = Payment.objects.all()
    serializer_class = PaymentSerializer
    filterset_fields = ['payment_type', 'payment_method', 'ar', 'ap']
    search_fields = ['payment_no']
    ordering_fields = ['payment_date', 'amount', 'created_at']

    def perform_create(self, serializer):
        """三单匹配超付阻断:仅在 THREE_WAY_MATCH_ENFORCE=True 时对'有采购订单的应付付款'
        校验累计付款不超过已收货物可结算价值(默认关闭,避免影响未接三单匹配的存量付款/测试)。"""
        from django.conf import settings
        if getattr(settings, 'THREE_WAY_MATCH_ENFORCE', False):
            data = serializer.validated_data
            ap = data.get('ap')
            if data.get('payment_type') == 'AP' and ap is not None and getattr(ap, 'po_id', None):
                from apps.purchase.matching import assert_can_pay
                cumulative = (ap.amount_paid or 0) + (data.get('amount') or 0)
                assert_can_pay(ap.po, cumulative)
        super().perform_create(serializer)


class ExpenseViewSet(PermissionMixin, WorkflowEnforcementMixin, SoftDeleteMixin, UserTrackingMixin, viewsets.ModelViewSet):
    """
    ViewSet for Expense management.
    """
    queryset = Expense.objects.all()
    serializer_class = ExpenseSerializer
    filterset_fields = ['project', 'department', 'user', 'category', 'status', 'is_deleted']
    search_fields = ['expense_no', 'description']
    ordering_fields = ['expense_date', 'amount', 'created_at']

    permission_module = 'finance'
    permission_resource = 'expense'
    
    # Workflow configuration
    workflow_business_type = 'EXPENSE'
    workflow_amount_field = 'amount'
    workflow_no_field = 'expense_no'

    def perform_create(self, serializer):
        # 默认申请人为当前用户（前端未传 user 时），同时保留 created_by/updated_by 注入
        kwargs = {'created_by': self.request.user, 'updated_by': self.request.user}
        if not serializer.validated_data.get('user'):
            kwargs['user'] = self.request.user
        serializer.save(**kwargs)

    @action(detail=True, methods=['post'])
    def submit(self, request, pk=None):
        """Submit expense for approval with workflow."""
        expense = self.get_object()
        if expense.status != 'DRAFT':
            return Response(
                {'error': '只能提交草稿状态的报销单'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Try to start workflow
        try:
            from apps.core.workflow.services import WorkflowService
            
            instance, error = WorkflowService.start_workflow(
                business_type='EXPENSE',
                business_id=expense.id,
                business_no=expense.expense_no,
                submitter=request.user,
                amount=expense.amount
            )
            
            if instance:
                expense.status = 'SUBMITTED'
                expense.save()
                return Response({
                    **ExpenseSerializer(expense).data,
                    'workflow_started': True,
                    'workflow_id': instance.id
                })
            else:
                # No workflow configured, auto-approve
                expense.status = 'APPROVED'
                expense.save()
                logger.info(f'费用报销 {expense.expense_no} 自动批准（未配置审批流程）')
                return Response({
                    **ExpenseSerializer(expense).data,
                    'workflow_started': False,
                    'auto_approved': True,
                    'message': error or '未配置审批流程，已自动批准'
                })
                
        except Exception as e:
            # Workflow module not available, auto-approve
            logger.warning(f'工作流服务异常，费用报销 {expense.expense_no} 自动批准: {e}')
            expense.status = 'APPROVED'
            expense.save()
            return Response({
                **ExpenseSerializer(expense).data,
                'auto_approved': True,
                'message': '工作流服务不可用，已自动批准'
            })
    
    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """Approve expense - 仅在无活跃工作流时允许直接审批."""
        expense = self.get_object()
        if expense.status != 'SUBMITTED':
            return Response(
                {'error': '只能批准已提交的报销单'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # 检查是否有活跃工作流
        workflow_error = self.check_workflow_allows_direct_action(expense, '审批')
        if workflow_error:
            return workflow_error
        
        expense.status = 'APPROVED'
        expense.save()
        return Response(ExpenseSerializer(expense).data)
    
    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """Reject expense - 仅在无活跃工作流时允许直接拒绝."""
        expense = self.get_object()
        if expense.status != 'SUBMITTED':
            return Response(
                {'error': '只能拒绝已提交的报销单'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # 检查是否有活跃工作流
        workflow_error = self.check_workflow_allows_direct_action(expense, '拒绝')
        if workflow_error:
            return workflow_error
        
        expense.status = 'REJECTED'
        expense.save()
        return Response(ExpenseSerializer(expense).data)
    
    @action(detail=True, methods=['post'])
    def reimburse(self, request, pk=None):
        """已停用:报销付款统一由银行流水核销完成(付款核销工作台)。

        历史上 reimburse() 直接把报销单标 PAID、写报销日期,绕过统一核销台账,与
        合同付款 pay()、AP record_payment() 是同一类"双轨"问题。收口后员工报销
        →PAID 只经 settle→ExpensePayableSource.write_back 驱动(审批通过即经
        register_expense_payable 信号进台账候选)。
        """
        return Response(
            {'error': '报销付款已统一由银行流水核销完成:请在「付款核销工作台」核销对应银行流水。此接口已停用。'},
            status=status.HTTP_409_CONFLICT,
        )


class AccountReceivableViewSet(PermissionMixin, SoftDeleteMixin, UserTrackingMixin, viewsets.ModelViewSet):
    permission_module = 'finance'
    permission_resource = 'receivable'
    """
    ViewSet for AccountReceivable management.
    财务敏感数据 - 仅财务人员和管理层可访问
    """
    queryset = AccountReceivable.objects.all()
    serializer_class = AccountReceivableSerializer
    filterset_fields = ['customer', 'project', 'status', 'is_deleted']
    search_fields = ['ar_no', 'invoice_no']
    ordering_fields = ['invoice_date', 'due_date', 'created_at']
    
    @action(detail=True, methods=['post'])
    def record_payment(self, request, pk=None):
        """Record a payment.

        创建 Payment 记录并由 Payment.save() 统一更新 amount_paid，
        保证与对账（按 Payment 取收款明细）口径一致，避免重复记账。
        """
        payment_amount = request.data.get('amount')

        if not payment_amount or float(payment_amount) <= 0:
            return Response(
                {'error': '请提供有效的付款金额'},
                status=status.HTTP_400_BAD_REQUEST
            )

        from decimal import Decimal as D

        payment_amount = D(str(payment_amount))

        with transaction.atomic():
            ar = AccountReceivable.objects.select_for_update().get(pk=pk)
            if ar.amount_paid + payment_amount > ar.amount_due:
                return Response(
                    {'error': '付款金额超过应收金额'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # 创建收款记录，Payment.save() 会原子更新 ar.amount_paid
            Payment.objects.create(
                payment_type='AR',
                ar=ar,
                payment_date=request.data.get('payment_date') or timezone.now().date(),
                payment_method=_normalize_payment_method(request.data.get('payment_method')),
                amount=payment_amount,
                currency=ar.currency,
                exchange_rate=ar.exchange_rate,
                notes=request.data.get('notes', ''),
                created_by=request.user,
                updated_by=request.user,
            )

            # 重新计算状态（Payment.save 已 F() 自增 amount_paid，需刷新后判断）
            ar.refresh_from_db(fields=['amount_paid'])
            if ar.amount_paid >= ar.amount_due:
                ar.status = 'PAID'
            elif ar.amount_paid > 0:
                ar.status = 'PARTIAL'
            ar.save(update_fields=['status'])

        return Response(AccountReceivableSerializer(ar).data)
    
    @action(detail=False, methods=['get'])
    def overdue(self, request):
        """Get overdue receivables."""
        today = timezone.now().date()
        overdue_ars = self.get_queryset().filter(
            due_date__lt=today,
            status__in=['PENDING', 'PARTIAL']
        )
        serializer = self.get_serializer(overdue_ars, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def aging(self, request):
        """Get AR aging report."""
        today = timezone.now().date()
        ars = self.get_queryset().filter(status__in=['PENDING', 'PARTIAL'])
        
        aging_data = {
            'current': 0,
            '1-30_days': 0,
            '31-60_days': 0,
            '61-90_days': 0,
            'over_90_days': 0,
        }
        
        for ar in ars:
            days_overdue = (today - ar.due_date).days
            amount = float(ar.amount_remaining)
            
            if days_overdue <= 0:
                aging_data['current'] += amount
            elif days_overdue <= 30:
                aging_data['1-30_days'] += amount
            elif days_overdue <= 60:
                aging_data['31-60_days'] += amount
            elif days_overdue <= 90:
                aging_data['61-90_days'] += amount
            else:
                aging_data['over_90_days'] += amount
        
        return Response(aging_data)


class AccountPayableViewSet(PermissionMixin, SoftDeleteMixin, UserTrackingMixin, viewsets.ModelViewSet):
    permission_module = 'finance'
    permission_resource = 'payable'
    """
    ViewSet for AccountPayable management.
    财务敏感数据 - 仅财务人员和管理层可访问
    """
    queryset = AccountPayable.objects.all()
    serializer_class = AccountPayableSerializer
    filterset_fields = ['supplier', 'status', 'is_deleted']
    search_fields = ['ap_no', 'invoice_no']
    ordering_fields = ['invoice_date', 'due_date', 'created_at']
    
    @action(detail=True, methods=['post'])
    def record_payment(self, request, pk=None):
        """已停用:应付账款(AP)付款统一由银行流水核销完成(付款核销工作台)。

        历史上 record_payment 直接创建 Payment(payment_type='AP') 并手动重算
        amount_paid/status,绕过统一待付款项核销台账(PayableItem/PayableSettlement),
        与核销台账"两套并存"。收口后 AP 付款 →PAID 只经「付款核销工作台」的
        settle→write_back 驱动,此接口已停用,不再产生任何副作用。
        """
        return Response(
            {'error': '应付账款付款已统一由银行流水核销完成:请在「付款核销工作台」核销对应银行流水。此接口已停用。'},
            status=409,
        )
    
    @action(detail=False, methods=['get'])
    def overdue(self, request):
        """Get overdue payables."""
        today = timezone.now().date()
        overdue_aps = self.get_queryset().filter(
            due_date__lt=today,
            status__in=['PENDING', 'PARTIAL']
        )
        serializer = self.get_serializer(overdue_aps, many=True)
        return Response(serializer.data)


class InvoiceViewSet(PermissionMixin, SoftDeleteMixin, UserTrackingMixin, viewsets.ModelViewSet):
    permission_module = 'finance'
    permission_resource = 'invoice'
    """
    ViewSet for Invoice management.
    财务敏感数据 - 仅财务人员和管理层可访问
    """
    from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
    
    queryset = Invoice.objects.all()
    serializer_class = InvoiceSerializer
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    filterset_fields = ['invoice_type', 'status', 'reference_type', 'is_deleted']
    search_fields = ['invoice_no', 'digital_invoice_no', 'party_name']
    
    ordering_fields = ['invoice_date', 'total_amount', 'created_at']
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # 支持发票类型过滤
        invoice_type = self.request.query_params.get('invoice_type')
        if invoice_type:
            queryset = queryset.filter(invoice_type=invoice_type)
        
        # 支持发票号模糊查询
        invoice_no = self.request.query_params.get('invoice_no')
        if invoice_no:
            queryset = queryset.filter(
                Q(invoice_no__icontains=invoice_no) | 
                Q(digital_invoice_no__icontains=invoice_no)
            )
        
        # 支持状态过滤
        status_param = self.request.query_params.get('status')
        if status_param:
            queryset = queryset.filter(status=status_param)
        
        return queryset
    
    @action(detail=True, methods=['post'])
    def certify(self, request, pk=None):
        """Certify input invoice."""
        invoice = self.get_object()
        
        if invoice.invoice_type != 'INPUT':
            return Response(
                {'error': '只有进项发票需要认证'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if invoice.status == 'CERTIFIED':
            return Response(
                {'error': '发票已认证'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if invoice.status == 'VOID':
            return Response(
                {'error': '已作废的发票无法认证'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        invoice.status = 'CERTIFIED'
        invoice.save()
        return Response(InvoiceSerializer(invoice).data)
    
    @action(detail=True, methods=['post'])
    def void(self, request, pk=None):
        """Void invoice."""
        invoice = self.get_object()
        
        if invoice.status == 'VOID':
            return Response(
                {'error': '发票已作废'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        invoice.status = 'VOID'
        invoice.save()
        return Response(InvoiceSerializer(invoice).data)
    
    @action(detail=True, methods=['post'])
    def issue(self, request, pk=None):
        """开具销项发票：确认后事件驱动联动应收与'发货款'里程碑激活。"""
        invoice = self.get_object()
        if invoice.invoice_type != 'OUTPUT':
            return Response({'error': '只有销项发票可开具'}, status=status.HTTP_400_BAD_REQUEST)
        if invoice.status == 'VOID':
            return Response({'error': '已作废的发票无法开具'}, status=status.HTTP_400_BAD_REQUEST)
        with transaction.atomic():
            invoice.status = 'NORMAL'
            invoice.save(update_fields=['status', 'updated_at'])
            # 发货→开票→应收 链条收口：应收登记 + 发货款里程碑按事件激活（服务内幂等、防御式）
            from apps.finance.services import ensure_ar_and_activate_milestone
            ensure_ar_and_activate_milestone(invoice, request.user)
        return Response(InvoiceSerializer(invoice).data)

    @action(detail=False, methods=['post'])
    def auto_match(self, request):
        """自动匹配发票到销售订单/采购订单"""
        from apps.sales.models import SalesOrder
        from apps.purchase.models import PurchaseOrder
        from fuzzywuzzy import fuzz
        
        matched_count = 0
        invoices = Invoice.objects.filter(
            is_deleted=False,
            reference_type__isnull=True
        ).exclude(reference_id__isnull=False)
        
        for invoice in invoices:
            matched = False
            
            if invoice.invoice_type == 'OUTPUT':
                # 销项发票 - 匹配销售订单 (通过购买方/客户名称和金额)
                buyer_name = invoice.buyer_name or invoice.party_name
                if buyer_name:
                    # 查找客户匹配的销售订单
                    sales_orders = SalesOrder.objects.filter(
                        is_deleted=False,
                        status__in=['CONFIRMED', 'PARTIAL', 'COMPLETED']
                    ).select_related('customer')
                    
                    best_match = None
                    best_score = 0
                    
                    for so in sales_orders:
                        if so.customer:
                            score = fuzz.ratio(buyer_name, so.customer.name)
                            # 如果金额也接近，提高分数
                            if invoice.total_amount and so.total_with_tax:
                                amount_diff = abs(float(invoice.total_amount) - float(so.total_with_tax))
                                if amount_diff < 1:  # 完全匹配
                                    score += 30
                                elif amount_diff / float(so.total_with_tax) < 0.01:  # 1%以内
                                    score += 20
                            
                            if score > best_score and score >= 70:
                                best_score = score
                                best_match = so
                    
                    if best_match:
                        invoice.reference_type = 'SALES_ORDER'
                        invoice.reference_id = best_match.id
                        if best_match.project:
                            invoice.project = best_match.project
                        invoice.save()
                        matched = True
                        matched_count += 1
            
            elif invoice.invoice_type == 'INPUT':
                # 进项发票 - 匹配采购订单 (通过销售方/供应商名称和金额)
                seller_name = invoice.seller_name or invoice.party_name
                if seller_name:
                    # 查找供应商匹配的采购订单
                    purchase_orders = PurchaseOrder.objects.filter(
                        is_deleted=False,
                        status__in=['CONFIRMED', 'PARTIAL', 'COMPLETED']
                    ).select_related('supplier')
                    
                    best_match = None
                    best_score = 0
                    
                    for po in purchase_orders:
                        if po.supplier:
                            score = fuzz.ratio(seller_name, po.supplier.name)
                            # 如果金额也接近，提高分数
                            if invoice.total_amount and po.total_with_tax:
                                amount_diff = abs(float(invoice.total_amount) - float(po.total_with_tax))
                                if amount_diff < 1:  # 完全匹配
                                    score += 30
                                elif amount_diff / float(po.total_with_tax) < 0.01:  # 1%以内
                                    score += 20
                            
                            if score > best_score and score >= 70:
                                best_score = score
                                best_match = po
                    
                    if best_match:
                        invoice.reference_type = 'PURCHASE_ORDER'
                        invoice.reference_id = best_match.id
                        if best_match.project:
                            invoice.project = best_match.project
                        invoice.save()
                        matched = True
                        matched_count += 1
        
        return Response({
            'matched_count': matched_count,
            'message': f'成功匹配 {matched_count} 张发票'
        })
    
    @action(detail=True, methods=['post'])
    def match_order(self, request, pk=None):
        """手动匹配发票到指定订单"""
        invoice = self.get_object()
        order_type = request.data.get('order_type')  # 'SALES_ORDER' or 'PURCHASE_ORDER'
        order_id = request.data.get('order_id')
        
        if not order_type or not order_id:
            return Response({'error': '请指定订单类型和订单ID'}, status=status.HTTP_400_BAD_REQUEST)
        
        if order_type == 'SALES_ORDER':
            from apps.sales.models import SalesOrder
            try:
                so = SalesOrder.objects.get(id=order_id)
                invoice.reference_type = 'SALES_ORDER'
                invoice.reference_id = so.id
                if so.project:
                    invoice.project = so.project
                invoice.save()
            except SalesOrder.DoesNotExist:
                return Response({'error': '销售订单不存在'}, status=status.HTTP_400_BAD_REQUEST)
        elif order_type == 'PURCHASE_ORDER':
            from apps.purchase.models import PurchaseOrder
            try:
                po = PurchaseOrder.objects.get(id=order_id)
                invoice.reference_type = 'PURCHASE_ORDER'
                invoice.reference_id = po.id
                if po.project:
                    invoice.project = po.project
                invoice.save()
            except PurchaseOrder.DoesNotExist:
                return Response({'error': '采购订单不存在'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({'error': '无效的订单类型'}, status=status.HTTP_400_BAD_REQUEST)
        
        return Response(InvoiceSerializer(invoice).data)
    
    @action(detail=False, methods=['get'])
    def download_template(self, request):
        """Download invoice import template Excel file."""
        import io
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '发票导入模板'
        
        headers = [
            '发票类型', '发票号码', '数电发票号码', '开票日期',
            '销方识别号', '销方名称', '购方识别号', '购买方名称',
            '金额', '税额', '价税合计', '发票来源', '发票票种', '发票状态', '备注'
        ]
        
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        example = [
            '销项发票', 'INV-2025-001', '25952000000022399509', '2025-01-15',
            '91440300MA5GR95713', '深圳市示例公司', '911310267954776025', '客户公司名称',
            '100000', '13000', '113000', '电子发票服务平台', '数电发票（增值税专用发票）', '正常', ''
        ]
        for col, val in enumerate(example, 1):
            ws.cell(row=2, column=col, value=val)
        
        widths = [10, 15, 25, 12, 22, 30, 22, 30, 12, 12, 12, 20, 30, 10, 20]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="invoice_import_template.xlsx"'
        return response
    
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Export invoices to Excel."""
        import io
        import openpyxl
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        from django.utils import timezone
        
        queryset = self.get_queryset().filter(is_deleted=False)
        
        invoice_type = request.query_params.get('invoice_type')
        status_filter = request.query_params.get('status')
        if invoice_type:
            queryset = queryset.filter(invoice_type=invoice_type)
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '发票列表'
        
        headers = [
            '序号', '发票类型', '发票号码', '数电发票号码', '开票日期',
            '销方识别号', '销方名称', '购方识别号', '购买方名称',
            '金额', '税额', '价税合计', '发票来源', '发票票种', '状态', '备注'
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        for row, inv in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=row-1)
            ws.cell(row=row, column=2, value=inv.get_invoice_type_display())
            ws.cell(row=row, column=3, value=inv.invoice_no)
            ws.cell(row=row, column=4, value=inv.digital_invoice_no)
            ws.cell(row=row, column=5, value=inv.invoice_date.strftime('%Y-%m-%d %H:%M:%S') if inv.invoice_date else '')
            ws.cell(row=row, column=6, value=inv.seller_tax_no)
            ws.cell(row=row, column=7, value=inv.seller_name)
            ws.cell(row=row, column=8, value=inv.buyer_tax_no)
            ws.cell(row=row, column=9, value=inv.buyer_name)
            ws.cell(row=row, column=10, value=float(inv.amount_before_tax))
            ws.cell(row=row, column=11, value=float(inv.tax_amount))
            ws.cell(row=row, column=12, value=float(inv.total_amount))
            ws.cell(row=row, column=13, value=inv.invoice_source)
            ws.cell(row=row, column=14, value=inv.get_invoice_category_display() if inv.invoice_category else '')
            ws.cell(row=row, column=15, value=inv.get_status_display())
            ws.cell(row=row, column=16, value=inv.notes)
        
        widths = [6, 10, 15, 25, 20, 22, 30, 22, 30, 15, 15, 15, 20, 30, 10, 20]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="invoices_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        return response
    
    @action(detail=False, methods=['post'])
    def import_excel(self, request):
        """Import invoices from Excel file (supports 2025年开票明细.xlsx format)."""
        import io
        import openpyxl
        from django.db import transaction
        from decimal import Decimal, InvalidOperation
        from datetime import datetime
        from .models import InvoiceItem
        
        file = request.FILES.get('file')
        if not file:
            return Response({'error': '请上传文件'}, status=status.HTTP_400_BAD_REQUEST)
        
        if not file.name.endswith(('.xlsx', '.xls')):
            return Response({'error': '只支持Excel文件格式(.xlsx, .xls)'}, status=status.HTTP_400_BAD_REQUEST)
        
        def parse_decimal(value, default=Decimal('0')):
            """Safely parse decimal value."""
            if value is None:
                return default
            try:
                return Decimal(str(value).replace(',', ''))
            except (InvalidOperation, ValueError):
                return default
        
        def parse_datetime(value):
            """Parse datetime from various formats."""
            if value is None:
                return datetime.now()
            if isinstance(value, datetime):
                return value
            if isinstance(value, str):
                for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%Y/%m/%d %H:%M:%S', '%Y/%m/%d']:
                    try:
                        return datetime.strptime(value.strip(), fmt)
                    except ValueError:
                        continue
            return datetime.now()
        
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file.read()))
            
            success_count = 0
            update_count = 0
            item_count = 0
            errors = []
            
            # ========== Step 1: Import invoice headers from "发票基础信息" ==========
            if '发票基础信息' in wb.sheetnames:
                sheet = wb['发票基础信息']
            else:
                sheet = wb.active
            
            headers = [str(cell.value).strip() if cell.value else '' for cell in sheet[1]]
            
            # Map headers to fields based on the Excel format
            # Support multiple column name variations
            col_map = {
                '发票类型': 'invoice_type_str',
                '发票代码': 'invoice_code',
                '发票号码': 'invoice_no',
                '数电发票号码': 'digital_invoice_no',
                # 销方/销售方 variations
                '销方识别号': 'seller_tax_no',
                '销售方识别号': 'seller_tax_no',
                '销方纳税人识别号': 'seller_tax_no',
                '销售方纳税人识别号': 'seller_tax_no',
                '销方名称': 'seller_name',
                '销售方名称': 'seller_name',
                '销售方': 'seller_name',
                # 购方/购买方 variations
                '购方识别号': 'buyer_tax_no',
                '购买方识别号': 'buyer_tax_no',
                '购方纳税人识别号': 'buyer_tax_no',
                '购买方纳税人识别号': 'buyer_tax_no',
                '购方名称': 'buyer_name',
                '购买方名称': 'buyer_name',
                '购买方': 'buyer_name',
                # Other fields
                '开票日期': 'invoice_date',
                '金额': 'amount_before_tax',
                '不含税金额': 'amount_before_tax',
                '税额': 'tax_amount',
                '价税合计': 'total_amount',
                '合计金额': 'total_amount',
                '发票来源': 'invoice_source',
                '发票票种': 'invoice_category_str',
                '发票状态': 'status_str',
                '状态': 'status_str',
                '备注': 'notes',
            }
            
            header_to_field = {}
            for idx, header in enumerate(headers):
                if header in col_map:
                    header_to_field[idx] = col_map[header]
            
            # Store created/updated invoices for line item import
            invoice_map = {}  # invoice_no -> Invoice object
            # Track processed invoices in this import to avoid duplicates within the same file
            processed_invoice_nos = set()
            skip_count = 0
            
            with transaction.atomic():
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                    try:
                        data = {}
                        for col_idx, cell in enumerate(row):
                            if col_idx in header_to_field:
                                field = header_to_field[col_idx]
                                value = cell.value
                                if value is not None:
                                    data[field] = value
                        
                        # Skip empty rows
                        if not data:
                            continue
                        
                        # Determine invoice number - prefer digital_invoice_no as unique identifier
                        digital_invoice_no = str(data.get('digital_invoice_no', '')).strip()
                        invoice_no_raw = str(data.get('invoice_no', '')).strip()
                        
                        # Use digital_invoice_no as primary key if available
                        invoice_no = digital_invoice_no or invoice_no_raw
                        if not invoice_no:
                            continue
                        
                        # Check for duplicates within the same import file
                        if invoice_no in processed_invoice_nos:
                            skip_count += 1
                            continue
                        processed_invoice_nos.add(invoice_no)
                        
                        # Parse invoice date
                        invoice_date = parse_datetime(data.get('invoice_date'))
                        
                        # Parse amounts
                        amount_before_tax = parse_decimal(data.get('amount_before_tax'))
                        tax_amount = parse_decimal(data.get('tax_amount'))
                        total_amount = parse_decimal(data.get('total_amount'))
                        if not total_amount:
                            total_amount = amount_before_tax + tax_amount
                        
                        # Parse status
                        status_map = {'正常': 'NORMAL', '已登记': 'REGISTERED', '已认证': 'CERTIFIED', '已作废': 'VOID', '红冲': 'RED'}
                        status_str = str(data.get('status_str', '')).strip()
                        inv_status = status_map.get(status_str, 'REGISTERED')
                        
                        # Parse invoice category
                        category_map = {
                            '数电发票（增值税专用发票）': 'SPECIAL',
                            '数电发票（普通发票）': 'NORMAL',
                        }
                        category_str = str(data.get('invoice_category_str', '')).strip()
                        invoice_category = category_map.get(category_str, '')
                        
                        seller_name = str(data.get('seller_name', '')).strip()
                        buyer_name = str(data.get('buyer_name', '')).strip()
                        
                        # 获取系统配置的公司名称
                        from apps.core.models import SystemConfig
                        company_config = SystemConfig.get_config()
                        company_name = company_config.company_name
                        
                        # 根据公司名称判断进项/销项发票
                        # 如果购买方是本公司 → 进项发票
                        # 如果销方是本公司 → 销项发票
                        invoice_type_str = str(data.get('invoice_type_str', '')).strip()
                        
                        if company_name and buyer_name == company_name:
                            # 购买方是本公司，说明是进项发票
                            invoice_type = 'INPUT'
                            party_name = seller_name  # 对方是销方（供应商）
                        elif company_name and seller_name == company_name:
                            # 销方是本公司，说明是销项发票
                            invoice_type = 'OUTPUT'
                            party_name = buyer_name  # 对方是购方（客户）
                        elif '进项' in invoice_type_str or '进' in invoice_type_str:
                            # 根据发票类型字段判断
                            invoice_type = 'INPUT'
                            party_name = seller_name
                        elif '销项' in invoice_type_str or '销' in invoice_type_str:
                            invoice_type = 'OUTPUT'
                            party_name = buyer_name or seller_name
                        else:
                            # 默认为销项发票
                            invoice_type = 'OUTPUT'
                            party_name = buyer_name or seller_name
                        
                        # Check if invoice exists (including soft-deleted) - check both invoice_no and digital_invoice_no
                        existing = Invoice.objects.filter(
                            Q(invoice_no=invoice_no) | 
                            Q(digital_invoice_no=invoice_no) |
                            (Q(digital_invoice_no=digital_invoice_no) & Q(digital_invoice_no__isnull=False) & ~Q(digital_invoice_no=''))
                        ).first()
                        if existing:
                            # Restore if was deleted
                            if existing.is_deleted:
                                existing.is_deleted = False
                                existing.deleted_at = None
                            existing.digital_invoice_no = str(data.get('digital_invoice_no', '')).strip()
                            existing.invoice_date = invoice_date
                            existing.seller_tax_no = str(data.get('seller_tax_no', '')).strip()
                            existing.seller_name = seller_name
                            existing.buyer_tax_no = str(data.get('buyer_tax_no', '')).strip()
                            existing.buyer_name = buyer_name
                            existing.party_name = party_name
                            existing.amount_before_tax = amount_before_tax
                            existing.tax_amount = tax_amount
                            existing.total_amount = total_amount
                            existing.invoice_source = str(data.get('invoice_source', '')).strip()
                            existing.invoice_category = invoice_category
                            existing.status = inv_status
                            existing.save()
                            # Soft-delete existing items for reimport
                            existing.items.filter(is_deleted=False).update(
                                is_deleted=True, deleted_at=timezone.now()
                            )
                            invoice_map[invoice_no] = existing
                            update_count += 1
                        else:
                            inv = Invoice.objects.create(
                                invoice_type=invoice_type,
                                invoice_no=invoice_no,
                                invoice_code=str(data.get('invoice_code', '')).strip(),
                                digital_invoice_no=str(data.get('digital_invoice_no', '')).strip(),
                                invoice_date=invoice_date,
                                seller_tax_no=str(data.get('seller_tax_no', '')).strip(),
                                seller_name=seller_name,
                                buyer_tax_no=str(data.get('buyer_tax_no', '')).strip(),
                                buyer_name=buyer_name,
                                party_name=party_name,
                                amount_before_tax=amount_before_tax,
                                tax_amount=tax_amount,
                                total_amount=total_amount,
                                invoice_source=str(data.get('invoice_source', '')).strip(),
                                invoice_category=invoice_category,
                                status=inv_status,
                            )
                            invoice_map[invoice_no] = inv
                            success_count += 1
                    
                    except Exception as e:
                        errors.append({'row': row_idx, 'sheet': '发票基础信息', 'error': str(e)})
                
                # ========== Step 2: Import line items from "信息汇总表" ==========
                if '信息汇总表' in wb.sheetnames:
                    detail_sheet = wb['信息汇总表']
                    detail_headers = [str(cell.value).strip() if cell.value else '' for cell in detail_sheet[1]]
                    
                    # Map detail sheet columns
                    detail_col_map = {
                        '数电发票号码': 'invoice_no',
                        '税收分类编码': 'tax_category_code',
                        '特定业务类型': 'business_type',
                        '货物或应税劳务名称': 'item_name',
                        '规格型号': 'specification',
                        '单位': 'unit',
                        '数量': 'quantity',
                        '单价': 'unit_price',
                        '金额': 'amount',
                        '税率': 'tax_rate',
                        '税额': 'tax_amount',
                    }
                    
                    detail_header_to_field = {}
                    for idx, header in enumerate(detail_headers):
                        if header in detail_col_map:
                            detail_header_to_field[idx] = detail_col_map[header]
                    
                    # Track line numbers per invoice
                    invoice_line_counts = {}
                    
                    for row_idx, row in enumerate(detail_sheet.iter_rows(min_row=2), start=2):
                        try:
                            data = {}
                            for col_idx, cell in enumerate(row):
                                if col_idx in detail_header_to_field:
                                    field = detail_header_to_field[col_idx]
                                    value = cell.value
                                    if value is not None:
                                        data[field] = value
                            
                            if not data:
                                continue
                            
                            invoice_no = str(data.get('invoice_no', '')).strip()
                            if not invoice_no:
                                continue
                            
                            # Find the invoice (from map or database)
                            invoice = invoice_map.get(invoice_no)
                            if not invoice:
                                invoice = Invoice.objects.filter(
                                    Q(invoice_no=invoice_no) | Q(digital_invoice_no=invoice_no)
                                ).first()
                            
                            if not invoice:
                                continue
                            
                            # Increment line number
                            if invoice_no not in invoice_line_counts:
                                invoice_line_counts[invoice_no] = 0
                            invoice_line_counts[invoice_no] += 1
                            line_no = invoice_line_counts[invoice_no]
                            
                            # Create line item
                            item_name = str(data.get('item_name', '')).strip()
                            if not item_name:
                                continue
                            
                            InvoiceItem.objects.create(
                                invoice=invoice,
                                line_no=line_no,
                                tax_category_code=str(data.get('tax_category_code', '')).strip(),
                                business_type=str(data.get('business_type', '')).strip(),
                                item_name=item_name,
                                specification=str(data.get('specification', '')).strip() if data.get('specification') else '',
                                unit=str(data.get('unit', '')).strip() if data.get('unit') else '',
                                quantity=parse_decimal(data.get('quantity'), None),
                                unit_price=parse_decimal(data.get('unit_price'), None),
                                amount=parse_decimal(data.get('amount')),
                                tax_rate=parse_decimal(data.get('tax_rate'), None),
                                tax_amount=parse_decimal(data.get('tax_amount'), None),
                            )
                            item_count += 1
                        
                        except Exception as e:
                            errors.append({'row': row_idx, 'sheet': '信息汇总表', 'error': str(e)})
            
            return Response({
                'success_count': success_count,
                'update_count': update_count,
                'skip_count': skip_count,
                'item_count': item_count,
                'error_count': len(errors),
                'errors': errors[:10]
            })
        
        except Exception as e:
            return Response({'error': f'导入失败: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def perform_destroy(self, instance):
        """删除发票时同时删除关联的附件"""
        from apps.core.models import Attachment
        
        # 删除关联的附件文件
        attachments = Attachment.objects.filter(related_model='Invoice', related_id=instance.id)
        for attachment in attachments:
            # 删除物理文件
            if attachment.file:
                attachment.file.delete(save=False)
            attachment.delete()
        
        # 删除发票（物理删除）
        instance.delete()
    
    @action(detail=False, methods=['post'])
    def bulk_delete(self, request):
        """Bulk delete invoices."""
        ids = request.data.get('ids', [])
        if not ids:
            return Response({'error': '请选择要删除的发票'}, status=status.HTTP_400_BAD_REQUEST)
        
        from apps.core.models import Attachment
        from .models import InvoiceItem
        
        # 删除关联的附件文件
        attachments = Attachment.objects.filter(related_model='Invoice', related_id__in=ids)
        for attachment in attachments:
            if attachment.file:
                attachment.file.delete(save=False)
            attachment.delete()
        
        # 软删除发票明细
        InvoiceItem.objects.filter(invoice_id__in=ids, is_deleted=False).update(
            is_deleted=True, deleted_at=timezone.now()
        )

        # 软删除发票
        deleted_count = Invoice.objects.filter(id__in=ids, is_deleted=False).update(
            is_deleted=True, deleted_at=timezone.now()
        )
        
        return Response({
            'success': True,
            'deleted_count': deleted_count
        })
    
    @action(detail=False, methods=['post'])
    def set_project(self, request):
        """
        批量设置发票的关联项目。
        用于将发票关联到项目进行成本核算。
        """
        invoice_ids = request.data.get('invoice_ids', [])
        project_id = request.data.get('project_id')
        
        if not invoice_ids:
            return Response({'error': '请选择要关联的发票'}, status=status.HTTP_400_BAD_REQUEST)
        
        if not project_id:
            return Response({'error': '请选择项目'}, status=status.HTTP_400_BAD_REQUEST)
        
        from apps.projects.models import Project
        
        try:
            project = Project.objects.get(id=project_id, is_deleted=False)
        except Project.DoesNotExist:
            return Response({'error': '项目不存在'}, status=status.HTTP_400_BAD_REQUEST)
        
        updated_count = Invoice.objects.filter(id__in=invoice_ids).update(project=project)
        
        return Response({
            'success': True,
            'updated_count': updated_count,
            'project_code': project.code,
            'project_name': project.name
        })
    
    @action(detail=False, methods=['post'])
    def import_pdf(self, request):
        """
        批量导入发票PDF文件。
        PDF文件名应包含数电发票号码，系统将自动匹配并关联到对应发票。
        支持的文件名格式:
        - 25952000000272801788.pdf
        - 发票_25952000000272801788.pdf
        - 任何包含20位数电发票号码的文件名
        """
        import re
        import os
        from apps.core.models import Attachment
        
        # 支持多文件上传(files)和单文件上传(file)
        files = request.FILES.getlist('files')
        if not files:
            # el-upload 的 multiple 模式会逐个发送文件，字段名为 'file'
            single_file = request.FILES.get('file')
            if single_file:
                files = [single_file]
        
        if not files:
            return Response({'error': '请选择要上传的PDF文件'}, status=status.HTTP_400_BAD_REQUEST)
        
        # 数电发票号码通常是20位数字
        invoice_no_pattern = re.compile(r'(\d{20})')
        
        results = {
            'total': len(files),
            'matched': 0,
            'unmatched': 0,
            'matched_files': [],
            'unmatched_files': []
        }
        
        for file in files:
            filename = file.name
            file_ext = os.path.splitext(filename)[1].lower()
            
            # 只接受PDF文件
            if file_ext != '.pdf':
                results['unmatched'] += 1
                results['unmatched_files'].append({
                    'filename': filename,
                    'reason': '不是PDF文件'
                })
                continue
            
            # 从文件名中提取数电发票号码
            match = invoice_no_pattern.search(filename)
            if not match:
                results['unmatched'] += 1
                results['unmatched_files'].append({
                    'filename': filename,
                    'reason': '文件名中未找到20位数电发票号码'
                })
                continue
            
            invoice_no = match.group(1)
            
            # 查找对应的发票
            invoice = Invoice.objects.filter(
                Q(digital_invoice_no=invoice_no) | Q(invoice_no=invoice_no),
                is_deleted=False
            ).first()
            
            if not invoice:
                results['unmatched'] += 1
                results['unmatched_files'].append({
                    'filename': filename,
                    'reason': f'未找到发票号码 {invoice_no}'
                })
                continue
            
            # 检查是否已存在相同的附件
            existing = Attachment.objects.filter(
                related_model='Invoice',
                related_id=invoice.id,
                original_name=filename
            ).exists()
            
            if existing:
                results['unmatched'] += 1
                results['unmatched_files'].append({
                    'filename': filename,
                    'reason': '该附件已存在'
                })
                continue
            
            # 创建附件记录
            attachment = Attachment.objects.create(
                file=file,
                original_name=filename,
                file_size=file.size,
                file_type='application/pdf',
                category='INVOICE',
                related_model='Invoice',
                related_id=invoice.id,
                description=f'发票PDF - {invoice_no}',
                uploaded_by=request.user
            )
            
            results['matched'] += 1
            results['matched_files'].append({
                'filename': filename,
                'invoice_no': invoice_no,
                'invoice_id': invoice.id,
                'attachment_id': attachment.id
            })
        
        return Response(results)
    
    @action(detail=True, methods=['get'])
    def attachments(self, request, pk=None):
        """获取发票的附件列表"""
        from apps.core.models import Attachment
        from apps.core.serializers import AttachmentSerializer
        
        invoice = self.get_object()
        attachments = Attachment.objects.filter(
            related_model='Invoice',
            related_id=invoice.id
        )
        
        serializer = AttachmentSerializer(attachments, many=True, context={'request': request})
        return Response(serializer.data)


class SharedExpenseViewSet(PermissionMixin, SoftDeleteMixin, UserTrackingMixin, viewsets.ModelViewSet):
    """
    ViewSet for SharedExpense management.
    """
    queryset = SharedExpense.objects.all()
    serializer_class = SharedExpenseSerializer
    filterset_fields = ['category', 'status', 'allocation_method', 'is_deleted']
    search_fields = ['expense_no', 'name']
    ordering_fields = ['expense_date', 'amount', 'created_at']

    permission_module = 'finance'
    permission_resource = 'shared_expense'
    
    @action(detail=True, methods=['post'])
    def calculate_allocation(self, request, pk=None):
        """
        Calculate allocation ratios based on allocation method.
        Returns preview without saving.
        """
        shared_expense = self.get_object()
        project_ids = request.data.get('project_ids', [])
        custom_ratios = request.data.get('custom_ratios', {})
        
        if not project_ids:
            return Response(
                {'error': '请选择至少一个项目'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        projects = Project.objects.filter(id__in=project_ids, is_deleted=False)
        
        if not projects.exists():
            return Response(
                {'error': '未找到有效项目'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        allocations = self._calculate_allocations(
            shared_expense, 
            projects, 
            custom_ratios
        )
        
        return Response({
            'expense_id': shared_expense.id,
            'expense_no': shared_expense.expense_no,
            'total_amount': float(shared_expense.amount),
            'allocation_method': shared_expense.allocation_method,
            'allocations': allocations
        })
    
    @action(detail=True, methods=['post'])
    def allocate(self, request, pk=None):
        """
        Perform the actual allocation to projects.
        """
        shared_expense = self.get_object()
        
        if shared_expense.status == 'ALLOCATED':
            return Response(
                {'error': '该费用已分摊'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        project_ids = request.data.get('project_ids', [])
        custom_ratios = request.data.get('custom_ratios', {})
        
        if not project_ids:
            return Response(
                {'error': '请选择至少一个项目'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        projects = Project.objects.filter(id__in=project_ids, is_deleted=False)
        
        if not projects.exists():
            return Response(
                {'error': '未找到有效项目'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        allocations_data = self._calculate_allocations(
            shared_expense, 
            projects, 
            custom_ratios
        )
        
        with transaction.atomic():
            # Delete existing allocations
            shared_expense.allocations.all().delete()
            
            # Create new allocations
            for alloc_data in allocations_data:
                SharedExpenseAllocation.objects.create(
                    shared_expense=shared_expense,
                    project_id=alloc_data['project_id'],
                    allocation_ratio=Decimal(str(alloc_data['ratio'])),
                    allocated_amount=Decimal(str(alloc_data['amount'])),
                    created_by=request.user
                )
            
            # Update status
            shared_expense.status = 'ALLOCATED'
            shared_expense.allocated_at = timezone.now()
            shared_expense.allocated_by = request.user
            shared_expense.save()
        
        return Response(SharedExpenseSerializer(shared_expense).data)
    
    @action(detail=True, methods=['post'])
    def cancel_allocation(self, request, pk=None):
        """Cancel allocation and reset to pending status."""
        shared_expense = self.get_object()
        
        if shared_expense.status != 'ALLOCATED':
            return Response(
                {'error': '只能取消已分摊的费用'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        with transaction.atomic():
            # Delete allocations
            shared_expense.allocations.all().delete()
            
            # Reset status
            shared_expense.status = 'PENDING'
            shared_expense.allocated_at = None
            shared_expense.allocated_by = None
            shared_expense.save()
        
        return Response(SharedExpenseSerializer(shared_expense).data)
    
    @action(detail=False, methods=['get'])
    def allocation_methods(self, request):
        """Get available allocation methods."""
        return Response([
            {'value': 'EQUAL', 'label': '平均分摊'},
            {'value': 'REVENUE', 'label': '按收入比例'},
            {'value': 'HEADCOUNT', 'label': '按人员数量'},
            {'value': 'LABOR_HOURS', 'label': '按工时比例'},
            {'value': 'BUDGET', 'label': '按预算比例'},
            {'value': 'CUSTOM', 'label': '自定义比例'},
        ])
    
    @action(detail=False, methods=['get'])
    def project_allocation_summary(self, request):
        """Get total allocated expenses for each project."""
        project_id = request.query_params.get('project_id')
        
        queryset = SharedExpenseAllocation.objects.filter(
            shared_expense__status='ALLOCATED',
            shared_expense__is_deleted=False
        )
        
        if project_id:
            queryset = queryset.filter(project_id=project_id)
        
        summary = queryset.values(
            'project__id', 'project__code', 'project__name'
        ).annotate(
            total_allocated=Sum('allocated_amount'),
            allocation_count=Count('id')
        ).order_by('-total_allocated')
        
        return Response(list(summary))
    
    def _calculate_allocations(self, shared_expense, projects, custom_ratios=None):
        """Calculate allocation based on method."""
        method = shared_expense.allocation_method
        total_amount = shared_expense.amount
        project_list = list(projects)
        n = len(project_list)
        
        if method == 'EQUAL':
            # Equal split — 用 Decimal 计算，舍入尾差补到最后一项，保证合计=总额
            ratio = Decimal('1') / n
            per = (total_amount / n).quantize(Decimal('0.01'))
            result = []
            allocated = Decimal('0')
            for idx, p in enumerate(project_list):
                amount = total_amount - allocated if idx == n - 1 else per
                allocated += amount
                result.append(
                    {
                        'project_id': p.id,
                        'project_code': p.code,
                        'project_name': p.name,
                        'ratio': float(ratio),
                        'amount': float(amount),
                    }
                )
            return result

        elif method == 'REVENUE':
            # By revenue ratio
            from apps.sales.models import SalesOrder
            revenues = {}
            for p in project_list:
                total = SalesOrder.objects.filter(
                    project=p,
                    status__in=['CONFIRMED', 'SHIPPED', 'COMPLETED'],
                    is_deleted=False
                ).aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
                revenues[p.id] = total
            
            total_revenue = sum(revenues.values()) or Decimal('1')
            
            return [
                {
                    'project_id': p.id,
                    'project_code': p.code,
                    'project_name': p.name,
                    'ratio': float(revenues[p.id] / total_revenue),
                    'amount': float(total_amount * revenues[p.id] / total_revenue)
                }
                for p in project_list
            ]
        
        elif method == 'HEADCOUNT':
            # By headcount
            from apps.projects.models import ProjectMember
            headcounts = {}
            for p in project_list:
                count = ProjectMember.objects.filter(
                    project=p,
                    is_deleted=False
                ).count() or 1
                headcounts[p.id] = count
            
            total_headcount = sum(headcounts.values()) or 1
            
            return [
                {
                    'project_id': p.id,
                    'project_code': p.code,
                    'project_name': p.name,
                    'ratio': float(headcounts[p.id] / total_headcount),
                    'amount': float(total_amount * headcounts[p.id] / total_headcount)
                }
                for p in project_list
            ]
        
        elif method == 'LABOR_HOURS':
            # By labor hours
            from apps.projects.models import ProjectMember
            hours = {}
            for p in project_list:
                total = ProjectMember.objects.filter(
                    project=p,
                    is_deleted=False
                ).aggregate(total=Sum('actual_hours'))['total'] or Decimal('1')
                hours[p.id] = total
            
            total_hours = sum(hours.values()) or Decimal('1')
            
            return [
                {
                    'project_id': p.id,
                    'project_code': p.code,
                    'project_name': p.name,
                    'ratio': float(hours[p.id] / total_hours),
                    'amount': float(total_amount * hours[p.id] / total_hours)
                }
                for p in project_list
            ]
        
        elif method == 'BUDGET':
            # By budget ratio
            budgets = {}
            for p in project_list:
                budgets[p.id] = p.budget_total or Decimal('0')
            
            total_budget = sum(budgets.values()) or Decimal('1')
            
            return [
                {
                    'project_id': p.id,
                    'project_code': p.code,
                    'project_name': p.name,
                    'ratio': float(budgets[p.id] / total_budget),
                    'amount': float(total_amount * budgets[p.id] / total_budget)
                }
                for p in project_list
            ]
        
        elif method == 'CUSTOM' and custom_ratios:
            # Custom ratios provided — 校验比例合计>0，避免全 0 时各项金额为 0 仍标记已分摊
            total_ratio = sum(Decimal(str(v)) for v in custom_ratios.values())
            if total_ratio <= 0:
                from rest_framework.exceptions import ValidationError

                raise ValidationError({'custom_ratios': '自定义比例合计必须大于 0'})

            result = []
            allocated = Decimal('0')
            for idx, p in enumerate(project_list):
                share = Decimal(str(custom_ratios.get(str(p.id), 0)))
                ratio = share / total_ratio
                amount = (total_amount * share / total_ratio).quantize(Decimal('0.01'))
                if idx == n - 1:
                    amount = total_amount - allocated
                allocated += amount
                result.append(
                    {
                        'project_id': p.id,
                        'project_code': p.code,
                        'project_name': p.name,
                        'ratio': float(ratio),
                        'amount': float(amount),
                    }
                )
            return result

        # Default to equal — 同 EQUAL，Decimal 计算并把尾差补到最后一项
        ratio = Decimal('1') / n
        per = (total_amount / n).quantize(Decimal('0.01'))
        result = []
        allocated = Decimal('0')
        for idx, p in enumerate(project_list):
            amount = total_amount - allocated if idx == n - 1 else per
            allocated += amount
            result.append(
                {
                    'project_id': p.id,
                    'project_code': p.code,
                    'project_name': p.name,
                    'ratio': float(ratio),
                    'amount': float(amount),
                }
            )
        return result


class SharedExpenseAllocationViewSet(PermissionMixin, viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for SharedExpenseAllocation (read-only).
    """

    permission_module = 'finance'
    permission_resource = 'shared_expense_allocation'
    queryset = SharedExpenseAllocation.objects.all()
    serializer_class = SharedExpenseAllocationSerializer
    filterset_fields = ['shared_expense', 'project']
    
    @action(detail=False, methods=['get'])
    def by_project(self, request):
        """Get all allocations for a specific project."""
        project_id = request.query_params.get('project_id')
        
        if not project_id:
            return Response(
                {'error': '请提供项目ID'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        allocations = self.get_queryset().filter(
            project_id=project_id,
            shared_expense__status='ALLOCATED'
        ).select_related('shared_expense')
        
        data = []
        for alloc in allocations:
            data.append({
                'id': alloc.id,
                'expense_no': alloc.shared_expense.expense_no,
                'expense_name': alloc.shared_expense.name,
                'category': alloc.shared_expense.category,
                'expense_date': alloc.shared_expense.expense_date,
                'total_amount': float(alloc.shared_expense.amount),
                'allocation_ratio': float(alloc.allocation_ratio),
                'allocated_amount': float(alloc.allocated_amount),
            })
        
        return Response(data)


class PaymentScheduleViewSet(PermissionMixin, SoftDeleteMixin, UserTrackingMixin, viewsets.ModelViewSet):
    """
    ViewSet for PaymentSchedule management.
    用于管理和跟踪销售订单的付款计划。

    permission_module = 'finance'
    permission_resource = 'payment_schedule'
    """
    queryset = PaymentSchedule.objects.filter(is_deleted=False).select_related(
        'sales_order', 'sales_order__customer', 'project', 'account_receivable'
    )
    serializer_class = PaymentScheduleSerializer
    filterset_fields = ['sales_order', 'project', 'status', 'milestone_type', 'reminder_status']
    search_fields = ['schedule_no', 'milestone_name', 'sales_order__order_no']
    ordering_fields = ['due_date', 'amount_due', 'created_at']
    
    @action(detail=False, methods=['get'])
    def summary(self, request):
        """Get payment schedule summary statistics."""
        from datetime import timedelta
        
        queryset = self.get_queryset()
        
        # 过滤条件
        project_id = request.query_params.get('project_id')
        customer_id = request.query_params.get('customer_id')
        
        if project_id:
            queryset = queryset.filter(project_id=project_id)
        if customer_id:
            queryset = queryset.filter(sales_order__customer_id=customer_id)
        
        # 统计数据
        stats = queryset.aggregate(
            total_amount=Sum('amount_due'),
            total_paid=Sum('amount_paid'),
        )
        
        total_amount = stats['total_amount'] or Decimal('0')
        total_paid = stats['total_paid'] or Decimal('0')
        total_remaining = total_amount - total_paid
        overall_progress = float(total_paid / total_amount * 100) if total_amount > 0 else 0
        
        # 按状态统计
        status_counts = queryset.values('status').annotate(count=Count('id'))
        counts = {s['status']: s['count'] for s in status_counts}
        
        # 即将到期的付款（7天内）
        today = timezone.now().date()
        upcoming = queryset.filter(
            status__in=['PENDING', 'PARTIAL'],
            due_date__gte=today,
            due_date__lte=today + timedelta(days=7)
        ).order_by('due_date')[:10]
        
        # 已逾期的付款
        overdue = queryset.filter(
            status__in=['PENDING', 'PARTIAL'],
            due_date__lt=today
        ).order_by('due_date')[:10]
        
        return Response({
            'total_amount': total_amount,
            'total_paid': total_paid,
            'total_remaining': total_remaining,
            'overall_progress': round(overall_progress, 2),
            'pending_count': counts.get('PENDING', 0),
            'partial_count': counts.get('PARTIAL', 0),
            'paid_count': counts.get('PAID', 0),
            'overdue_count': counts.get('OVERDUE', 0),
            'upcoming_payments': PaymentScheduleSerializer(upcoming, many=True).data,
            'overdue_payments': PaymentScheduleSerializer(overdue, many=True).data,
        })
    
    @action(detail=False, methods=['post'])
    def generate_for_order(self, request):
        """Generate payment schedules for a sales order."""
        from apps.sales.models import SalesOrder
        
        sales_order_id = request.data.get('sales_order_id')
        if not sales_order_id:
            return Response({'error': '请提供销售订单ID'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            sales_order = SalesOrder.objects.get(id=sales_order_id, is_deleted=False)
        except SalesOrder.DoesNotExist:
            return Response({'error': '销售订单不存在'}, status=status.HTTP_404_NOT_FOUND)
        
        schedules = PaymentSchedule.generate_from_sales_order(sales_order)
        
        return Response({
            'success': True,
            'count': len(schedules),
            'schedules': PaymentScheduleSerializer(schedules, many=True).data
        })
    
    @action(detail=True, methods=['post'])
    def record_payment(self, request, pk=None):
        """Record a payment against this schedule."""
        schedule = self.get_object()
        amount = request.data.get('amount')
        payment_date = request.data.get('payment_date', timezone.now().date())
        
        if not amount:
            return Response({'error': '请提供付款金额'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            amount = Decimal(str(amount))
        except:
            return Response({'error': '金额格式错误'}, status=status.HTTP_400_BAD_REQUEST)
        
        from django.db.models import F as DbF
        from django.db import transaction
        with transaction.atomic():
            PaymentSchedule.objects.filter(pk=schedule.pk).update(
                amount_paid=DbF('amount_paid') + amount
            )
            schedule.refresh_from_db()
            if schedule.amount_paid >= schedule.amount_due:
                schedule.status = 'PAID'
                schedule.actual_paid_date = payment_date
                schedule.reminder_status = 'COLLECTED'
            elif schedule.amount_paid > 0:
                schedule.status = 'PARTIAL'
            schedule.save(update_fields=['status', 'actual_paid_date', 'reminder_status'])

        return Response(PaymentScheduleSerializer(schedule).data)
    
    @action(detail=False, methods=['get'])
    def reminders(self, request):
        """Get payment schedules that need reminders."""
        from datetime import timedelta
        
        today = timezone.now().date()
        
        # 找出需要提醒的付款计划
        # 1. 即将到期的（在提醒天数内）
        # 2. 已逾期但未提醒的
        reminders = []
        
        queryset = self.get_queryset().filter(
            status__in=['PENDING', 'PARTIAL'],
            reminder_status='PENDING'
        )
        
        for schedule in queryset:
            remind_date = schedule.due_date - timedelta(days=schedule.reminder_days_before)
            
            if today >= remind_date:
                reminders.append({
                    'schedule': PaymentScheduleSerializer(schedule).data,
                    'reminder_type': 'OVERDUE' if schedule.is_overdue else 'UPCOMING',
                    'days_until_due': schedule.days_until_due,
                    'urgency': 'HIGH' if schedule.is_overdue else ('MEDIUM' if schedule.days_until_due <= 3 else 'LOW')
                })
        
        # 按紧急程度和到期日排序
        reminders.sort(key=lambda x: (
            0 if x['urgency'] == 'HIGH' else (1 if x['urgency'] == 'MEDIUM' else 2),
            x['days_until_due']
        ))
        
        return Response({
            'total_count': len(reminders),
            'reminders': reminders
        })
    
    @action(detail=True, methods=['post'])
    def mark_reminded(self, request, pk=None):
        """Mark a payment schedule as reminded."""
        schedule = self.get_object()
        schedule.reminder_status = 'REMINDED'
        schedule.last_reminded_at = timezone.now()
        schedule.save()
        
        return Response(PaymentScheduleSerializer(schedule).data)
    
    @action(detail=False, methods=['post'])
    def match_bank_statement(self, request):
        """Match a bank statement to a payment schedule."""
        from .bank_statement_models import BankStatement
        
        schedule_id = request.data.get('schedule_id')
        bank_statement_id = request.data.get('bank_statement_id')
        
        if not schedule_id or not bank_statement_id:
            return Response(
                {'error': '请提供付款计划ID和银行流水ID'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            schedule = PaymentSchedule.objects.get(id=schedule_id)
            bank_statement = BankStatement.objects.get(id=bank_statement_id)
        except (PaymentSchedule.DoesNotExist, BankStatement.DoesNotExist):
            return Response({'error': '记录不存在'}, status=status.HTTP_404_NOT_FOUND)
        
        # 更新付款进度
        amount = bank_statement.credit_amount
        schedule.amount_paid += amount
        
        if schedule.amount_paid >= schedule.amount_due:
            schedule.status = 'PAID'
            schedule.actual_paid_date = bank_statement.transaction_time.date()
            schedule.reminder_status = 'COLLECTED'
        elif schedule.amount_paid > 0:
            schedule.status = 'PARTIAL'
        
        schedule.save()
        
        # 关联银行流水到付款计划
        schedule.bank_statements.add(bank_statement)
        
        # 更新银行流水状态
        bank_statement.status = 'MATCHED'
        bank_statement.match_type = 'AR'
        bank_statement.project = schedule.project
        bank_statement.save()
        
        return Response({
            'success': True,
            'schedule': PaymentScheduleSerializer(schedule).data,
            'message': f'成功匹配付款 ¥{amount}，当前进度 {schedule.payment_progress}%'
        })


class PurchasePaymentScheduleViewSet(PermissionMixin, SoftDeleteMixin, UserTrackingMixin, viewsets.ModelViewSet):
    """
    ViewSet for PurchasePaymentSchedule management.
    用于管理和跟踪采购订单的付款计划。
    """
    queryset = PurchasePaymentSchedule.objects.filter(is_deleted=False).select_related(
        'purchase_order', 'purchase_order__supplier', 'project', 'account_payable'
    )
    serializer_class = PurchasePaymentScheduleSerializer
    filterset_fields = ['purchase_order', 'project', 'status', 'milestone_type', 'reminder_status']
    search_fields = ['schedule_no', 'milestone_name', 'purchase_order__order_no']
    ordering_fields = ['due_date', 'amount_due', 'created_at']

    permission_module = 'finance'
    permission_resource = 'purchase_payment_schedule'
    
    @action(detail=False, methods=['get'])
    def summary(self, request):
        """Get purchase payment schedule summary statistics."""
        from datetime import timedelta
        
        queryset = self.get_queryset()
        
        # 过滤条件
        project_id = request.query_params.get('project_id')
        supplier_id = request.query_params.get('supplier_id')
        
        if project_id:
            queryset = queryset.filter(project_id=project_id)
        if supplier_id:
            queryset = queryset.filter(purchase_order__supplier_id=supplier_id)
        
        # 统计数据
        stats = queryset.aggregate(
            total_amount=Sum('amount_due'),
            total_paid=Sum('amount_paid'),
        )
        
        total_amount = stats['total_amount'] or Decimal('0')
        total_paid = stats['total_paid'] or Decimal('0')
        total_remaining = total_amount - total_paid
        overall_progress = float(total_paid / total_amount * 100) if total_amount > 0 else 0
        
        # 按状态统计
        status_counts = queryset.values('status').annotate(count=Count('id'))
        counts = {s['status']: s['count'] for s in status_counts}
        
        # 即将到期的付款（7天内）
        today = timezone.now().date()
        upcoming = queryset.filter(
            status__in=['PENDING', 'PARTIAL'],
            due_date__gte=today,
            due_date__lte=today + timedelta(days=7)
        ).order_by('due_date')[:10]
        
        # 已逾期的付款
        overdue = queryset.filter(
            status__in=['PENDING', 'PARTIAL'],
            due_date__lt=today
        ).order_by('due_date')[:10]
        
        return Response({
            'total_amount': total_amount,
            'total_paid': total_paid,
            'total_remaining': total_remaining,
            'overall_progress': round(overall_progress, 2),
            'pending_count': counts.get('PENDING', 0),
            'partial_count': counts.get('PARTIAL', 0),
            'paid_count': counts.get('PAID', 0),
            'overdue_count': counts.get('OVERDUE', 0),
            'upcoming_payments': PurchasePaymentScheduleSerializer(upcoming, many=True).data,
            'overdue_payments': PurchasePaymentScheduleSerializer(overdue, many=True).data,
        })
    
    @action(detail=False, methods=['post'])
    def generate_for_order(self, request):
        """Generate payment schedules for a purchase order."""
        from apps.purchase.models import PurchaseOrder
        
        purchase_order_id = request.data.get('purchase_order_id')
        if not purchase_order_id:
            return Response({'error': '请提供采购订单ID'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            purchase_order = PurchaseOrder.objects.get(id=purchase_order_id, is_deleted=False)
        except PurchaseOrder.DoesNotExist:
            return Response({'error': '采购订单不存在'}, status=status.HTTP_404_NOT_FOUND)
        
        schedules = PurchasePaymentSchedule.generate_from_purchase_order(purchase_order)
        
        return Response({
            'success': True,
            'count': len(schedules),
            'schedules': PurchasePaymentScheduleSerializer(schedules, many=True).data
        })
    
    @action(detail=True, methods=['post'])
    def record_payment(self, request, pk=None):
        """已停用:采购付款计划的里程碑付款统一由银行流水核销驱动。

        历史上本 action 直接对里程碑 amount_paid 累加,是绕过统一核销台账、且不回写
        所属 AP 的"双轨"付款(与合同 pay()、AP record_payment、报销 reimburse 同类)。
        收口后:一个 PO 只有一张 AccountPayable(真实应付,已在核销台账),里程碑的
        已付/进度改由该 AP 的核销额按 milestone_order 顺序自动派生(见 finance/signals.py
        的 sync_purchase_schedules_from_ap),付款计划回归"计划 + 催付 + 进度展示"职能。
        """
        return Response(
            {'error': '采购付款已统一由银行流水核销完成:请在「付款核销工作台」核销对应银行流水。此接口已停用。'},
            status=status.HTTP_409_CONFLICT,
        )
    
    @action(detail=False, methods=['get'])
    def reminders(self, request):
        """Get purchase payment schedules that need reminders."""
        from datetime import timedelta
        
        today = timezone.now().date()
        
        reminders = []
        
        queryset = self.get_queryset().filter(
            status__in=['PENDING', 'PARTIAL'],
            reminder_status='PENDING'
        )
        
        for schedule in queryset:
            remind_date = schedule.due_date - timedelta(days=schedule.reminder_days_before)
            
            if today >= remind_date:
                reminders.append({
                    'schedule': PurchasePaymentScheduleSerializer(schedule).data,
                    'reminder_type': 'OVERDUE' if schedule.is_overdue else 'UPCOMING',
                    'days_until_due': schedule.days_until_due,
                    'urgency': 'HIGH' if schedule.is_overdue else ('MEDIUM' if schedule.days_until_due <= 3 else 'LOW')
                })
        
        reminders.sort(key=lambda x: (
            0 if x['urgency'] == 'HIGH' else (1 if x['urgency'] == 'MEDIUM' else 2),
            x['days_until_due']
        ))
        
        return Response({
            'total_count': len(reminders),
            'reminders': reminders
        })
    
    @action(detail=True, methods=['post'])
    def mark_reminded(self, request, pk=None):
        """Mark a purchase payment schedule as reminded."""
        schedule = self.get_object()
        schedule.reminder_status = 'REMINDED'
        schedule.last_reminded_at = timezone.now()
        schedule.save()
        
        return Response(PurchasePaymentScheduleSerializer(schedule).data)
    
    @action(detail=False, methods=['post'])
    def match_bank_statement(self, request):
        """Match a bank statement to a purchase payment schedule."""
        from .bank_statement_models import BankStatement
        
        schedule_id = request.data.get('schedule_id')
        bank_statement_id = request.data.get('bank_statement_id')
        
        if not schedule_id or not bank_statement_id:
            return Response(
                {'error': '请提供付款计划ID和银行流水ID'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            schedule = PurchasePaymentSchedule.objects.get(id=schedule_id)
            bank_statement = BankStatement.objects.get(id=bank_statement_id)
        except (PurchasePaymentSchedule.DoesNotExist, BankStatement.DoesNotExist):
            return Response({'error': '记录不存在'}, status=status.HTTP_404_NOT_FOUND)
        
        # 更新付款进度
        amount = bank_statement.debit_amount
        schedule.amount_paid += amount
        
        if schedule.amount_paid >= schedule.amount_due:
            schedule.status = 'PAID'
            schedule.actual_paid_date = bank_statement.transaction_time.date()
            schedule.reminder_status = 'PAID'
        elif schedule.amount_paid > 0:
            schedule.status = 'PARTIAL'
        
        schedule.save()
        
        # 关联银行流水到付款计划
        schedule.bank_statements.add(bank_statement)
        
        # 更新银行流水状态
        bank_statement.status = 'MATCHED'
        bank_statement.match_type = 'AP'
        bank_statement.project = schedule.project
        bank_statement.save()
        
        return Response({
            'success': True,
            'schedule': PurchasePaymentScheduleSerializer(schedule).data,
            'message': f'成功匹配付款 ¥{amount}，当前进度 {schedule.payment_progress}%'
        })


class PaymentRequestViewSet(PermissionMixin, WorkflowEnforcementMixin, SoftDeleteMixin, UserTrackingMixin, viewsets.ModelViewSet):
    """
    ViewSet for PaymentRequest management - 付款申请管理.
    """
    queryset = PaymentRequest.objects.all()
    serializer_class = PaymentRequestSerializer
    filterset_fields = ['supplier', 'project', 'status', 'payment_type', 'applicant', 'is_deleted']
    search_fields = ['request_no', 'title', 'reason']
    ordering_fields = ['expected_date', 'amount', 'created_at']

    permission_module = 'finance'
    permission_resource = 'payment_request'

    # Workflow configuration
    workflow_business_type = 'PAYMENT'
    workflow_amount_field = 'amount'
    workflow_no_field = 'request_no'
    
    def perform_create(self, serializer):
        """创建时自动设置申请人"""
        serializer.save(applicant=self.request.user)
    
    @action(detail=True, methods=['post'])
    def submit(self, request, pk=None):
        """提交付款申请审批 - 审批步骤由流程配置决定"""
        payment_req = self.get_object()
        if payment_req.status not in ['DRAFT', 'REJECTED']:
            return Response(
                {'error': '只能提交草稿或已拒绝状态的付款申请'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        amount = payment_req.amount or 0
        
        try:
            from apps.core.workflow.services import WorkflowService
            
            instance, error = WorkflowService.start_workflow(
                business_type='PAYMENT',
                business_id=payment_req.id,
                business_no=payment_req.request_no,
                submitter=request.user,
                amount=amount
            )
            
            if instance:
                payment_req.status = 'PENDING'
                payment_req.save()
                return Response({
                    **PaymentRequestSerializer(payment_req).data,
                    'workflow_started': True,
                    'workflow_id': instance.id,
                    'message': '已提交审批，请在审批中心查看审批进度'
                })
            else:
                # 未配置审批流程，直接批准
                payment_req.status = 'APPROVED'
                payment_req.approved_by = request.user
                payment_req.approved_at = timezone.now()
                payment_req.save()
                return Response({
                    **PaymentRequestSerializer(payment_req).data,
                    'workflow_started': False,
                    'message': error or '未配置审批流程，付款申请已直接批准'
                })
                
        except Exception as e:
            # 审批模块不可用，直接批准
            payment_req.status = 'APPROVED'
            payment_req.approved_by = request.user
            payment_req.approved_at = timezone.now()
            payment_req.save()
            return Response({
                **PaymentRequestSerializer(payment_req).data,
                'workflow_started': False,
                'message': f'付款申请已批准，但工作流服务异常: {e}'
            })
    
    @action(detail=True, methods=['post'])
    def pay(self, request, pk=None):
        """执行付款 - 审批通过后可执行"""
        payment_req = self.get_object()
        if payment_req.status != 'APPROVED':
            return Response(
                {'error': '只能对已批准的付款申请执行付款'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        with transaction.atomic():
            # 创建付款记录
            payment = Payment.objects.create(
                payment_no=f'P{payment_req.request_no[3:]}',
                payment_type='AP',
                ap=payment_req.ap,
                amount=payment_req.amount,
                payment_date=timezone.now().date(),
                payment_method=request.data.get('payment_method', 'BANK_TRANSFER'),
                notes=request.data.get('notes', f'付款申请: {payment_req.request_no}'),
                created_by=request.user
            )
            
            # 更新付款申请状态
            payment_req.status = 'PAID'
            payment_req.paid_at = timezone.now()
            payment_req.payment = payment
            payment_req.save()
            
            # 更新应付账款:上面 Payment.objects.create(payment_type='AP', ap=...) 已经过
            # Payment.save() 原子 F() 递增了 ap.amount_paid 一次;这里不再重复递增
            # (历史 bug:曾在此处对同一笔付款再手动 F() 一次,导致 amount_paid 被双记,
            # 见 apps.finance.tests.test_payable_ledger.PaymentRequestPayNoDoubleCountTest),
            # 只需刷新后让 AccountPayable.save() 按最新 amount_paid 重算 status。
            if payment_req.ap:
                payment_req.ap.refresh_from_db(fields=['amount_paid'])
                payment_req.ap.save(update_fields=['status'])
        
        return Response({
            **PaymentRequestSerializer(payment_req).data,
            'payment': PaymentSerializer(payment).data,
            'message': '付款成功'
        })
    
    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        """取消付款申请"""
        payment_req = self.get_object()
        if payment_req.status in ['PAID', 'CANCELLED']:
            return Response(
                {'error': '已付款或已取消的申请无法取消'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        payment_req.status = 'CANCELLED'
        payment_req.save()
        
        return Response(PaymentRequestSerializer(payment_req).data)
    
    @action(detail=False, methods=['get'])
    def my_requests(self, request):
        """获取我的付款申请"""
        queryset = self.get_queryset().filter(applicant=request.user)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def pending_approval(self, request):
        """获取待审批的付款申请（供审批人使用）"""
        queryset = self.get_queryset().filter(status='PENDING')
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
