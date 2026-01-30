"""
Views for sales app.
"""
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.http import HttpResponse
from django.db import models, transaction
from apps.core.mixins import SoftDeleteMixin, UserTrackingMixin
from apps.core.data_permission import DataPermissionMixin
from .models import (
    SalesQuotation, SalesQuotationLine,
    SalesOrder, SalesOrderLine,
    DeliveryOrder, DeliveryOrderLine,
    SalesContract
)
from .serializers import (
    SalesQuotationSerializer, SalesQuotationLineSerializer,
    SalesOrderSerializer, SalesOrderLineSerializer,
    DeliveryOrderSerializer, DeliveryOrderLineSerializer,
    SalesContractSerializer
)


class SalesQuotationViewSet(SoftDeleteMixin, UserTrackingMixin, DataPermissionMixin, viewsets.ModelViewSet):
    """
    ViewSet for SalesQuotation management.
    """
    queryset = SalesQuotation.objects.all()
    serializer_class = SalesQuotationSerializer
    filterset_fields = ['customer', 'project', 'status', 'is_deleted']
    search_fields = ['quote_no']
    ordering_fields = ['quote_date', 'created_at']
    
    @action(detail=True, methods=['post'])
    def create_new_version(self, request, pk=None):
        """Create a new version of the quotation."""
        old_quotation = self.get_object()
        
        with transaction.atomic():
            new_quotation = SalesQuotation.objects.create(
                customer=old_quotation.customer,
                project=old_quotation.project,
                valid_until=request.data.get('valid_until', old_quotation.valid_until),
                version=old_quotation.version + 1,
                created_by=request.user
            )
            
            for line in old_quotation.lines.filter(is_deleted=False):
                SalesQuotationLine.objects.create(
                    quotation=new_quotation,
                    item=line.item,
                    qty=line.qty,
                    unit_price=line.unit_price,
                    notes=line.notes,
                    created_by=request.user
                )
            
            # Update total
            from django.db.models import Sum
            total = new_quotation.lines.aggregate(Sum('line_amount'))['line_amount__sum'] or 0
            new_quotation.total_amount = total
            new_quotation.save()
        
        return Response(SalesQuotationSerializer(new_quotation).data, status=status.HTTP_201_CREATED)
    
    @action(detail=True, methods=['post'])
    def submit(self, request, pk=None):
        """提交报价单审批 - 审批步骤由流程配置决定"""
        quotation = self.get_object()
        if quotation.status not in ['DRAFT', 'REJECTED']:
            return Response(
                {'error': '只能提交草稿或已拒绝状态的报价单'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        amount = quotation.total_amount or 0
        
        try:
            from apps.core.workflow.services import WorkflowService
            
            instance, error = WorkflowService.start_workflow(
                business_type='QUOTATION',
                business_id=quotation.id,
                business_no=quotation.quote_no,
                submitter=request.user,
                amount=amount
            )
            
            if instance:
                quotation.status = 'PENDING'
                quotation.save()
                return Response({
                    **SalesQuotationSerializer(quotation).data,
                    'workflow_started': True,
                    'workflow_id': instance.id,
                    'message': '已提交审批，请在审批中心查看审批进度'
                })
            else:
                quotation.status = 'APPROVED'
                quotation.save()
                return Response({
                    **SalesQuotationSerializer(quotation).data,
                    'workflow_started': False,
                    'message': error or '未配置审批流程，报价单已直接通过'
                })
                
        except Exception as e:
            quotation.status = 'APPROVED'
            quotation.save()
            return Response({
                **SalesQuotationSerializer(quotation).data,
                'workflow_started': False,
                'message': f'报价单已通过，但工作流服务异常: {e}'
            })
    
    @action(detail=True, methods=['post'])
    def convert_to_so(self, request, pk=None):
        """Convert quotation to sales order."""
        quotation = self.get_object()
        
        if not quotation.project:
            return Response(
                {'error': '报价单必须关联项目才能转换为订单'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        with transaction.atomic():
            so = SalesOrder.objects.create(
                customer=quotation.customer,
                project=quotation.project,
                delivery_date=request.data.get('delivery_date'),
                created_by=request.user
            )
            
            for line in quotation.lines.filter(is_deleted=False):
                SalesOrderLine.objects.create(
                    so=so,
                    item=line.item,
                    qty=line.qty,
                    unit_price=line.unit_price,
                    created_by=request.user
                )
            
            # Update total
            from django.db.models import Sum
            total = so.lines.aggregate(Sum('line_amount'))['line_amount__sum'] or 0
            so.total_amount = total
            so.save()
            
            quotation.status = 'ACCEPTED'
            quotation.save()
        
        return Response(SalesOrderSerializer(so).data, status=status.HTTP_201_CREATED)


class SalesQuotationLineViewSet(SoftDeleteMixin, UserTrackingMixin, viewsets.ModelViewSet):
    """
    ViewSet for SalesQuotationLine management.
    """
    queryset = SalesQuotationLine.objects.all()
    serializer_class = SalesQuotationLineSerializer
    filterset_fields = ['quotation', 'item', 'is_deleted']
    search_fields = ['item__sku', 'item__name']


class SalesOrderViewSet(SoftDeleteMixin, UserTrackingMixin, DataPermissionMixin, viewsets.ModelViewSet):
    """
    ViewSet for SalesOrder management.
    
    销售订单审批流程由审批中心的流程配置决定。
    """
    queryset = SalesOrder.objects.all()
    serializer_class = SalesOrderSerializer
    filterset_fields = ['customer', 'project', 'status', 'is_deleted']
    search_fields = ['order_no']
    ordering_fields = ['order_date', 'created_at']
    
    @action(detail=False, methods=['get'])
    def for_linking(self, request):
        """获取可用于关联的销售订单（不受数据权限限制）"""
        from apps.projects.models import Project
        
        # 仅返回：已确认/部分发货的订单
        allowed_status = ['CONFIRMED', 'PARTIAL']  # 已确认且未完成
        
        # 排除已被项目关联的订单（非取消状态的项目）
        linked_order_ids = Project.objects.filter(
            is_deleted=False,
            sales_order__isnull=False,
            status__in=['DRAFT', 'PLANNING', 'PENDING', 'IN_PROGRESS', 'ACTIVE', 'PAUSED', 'COMPLETED']
        ).values_list('sales_order_id', flat=True)
        
        queryset = SalesOrder.objects.filter(
            is_deleted=False,
            status__in=allowed_status
        ).exclude(
            id__in=linked_order_ids
        ).order_by('-created_at')
        
        # 支持搜索
        search = request.query_params.get('search', '')
        if search:
            queryset = queryset.filter(
                models.Q(order_no__icontains=search) |
                models.Q(customer__name__icontains=search)
            )
        
        # 返回详细数据，包含明细信息供项目经理参考
        data = []
        for so in queryset[:100]:  # 限制返回数量
            # 获取订单明细
            lines = []
            for line in so.lines.filter(is_deleted=False):
                lines.append({
                    'id': line.id,
                    'product_name': line.item.name if line.item else line.custom_name,
                    'spec': line.item.spec if line.item else line.custom_spec,
                    'qty': float(line.qty),
                    'unit_price': float(line.unit_price),
                    'line_amount': float(line.line_amount),
                })
            
            data.append({
                'id': so.id,
                'order_no': so.order_no,
                'customer_order_no': so.customer_order_no or '',
                'customer': so.customer_id,
                'customer_name': so.customer.name if so.customer else '',
                'status': so.status,
                'order_date': so.order_date.strftime('%Y-%m-%d') if so.order_date else '',
                'delivery_date': so.delivery_date.strftime('%Y-%m-%d') if so.delivery_date else '',
                'total_amount': float(so.total_amount or 0),
                'total_with_tax': float(so.total_with_tax or 0),
                'lines': lines,
            })
        
        return Response(data)
    
    @action(detail=True, methods=['post'])
    def submit(self, request, pk=None):
        """提交销售订单审批 - 审批步骤由流程配置决定"""
        so = self.get_object()
        if so.status not in ['DRAFT', 'REJECTED']:
            return Response(
                {'error': '只能提交草稿或已拒绝状态的订单'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # 计算订单金额用于流程路由
        amount = so.total_with_tax or so.total_amount or 0
        
        try:
            from apps.core.workflow.services import WorkflowService
            
            instance, error = WorkflowService.start_workflow(
                business_type='SALES_ORDER',
                business_id=so.id,
                business_no=so.order_no,
                submitter=request.user,
                amount=amount
            )
            
            if instance:
                so.status = 'PENDING'
                so.save()
                return Response({
                    **SalesOrderSerializer(so).data,
                    'workflow_started': True,
                    'workflow_id': instance.id,
                    'message': '已提交审批，请在审批中心查看审批进度'
                })
            else:
                # 未配置审批流程，直接确认
                return self._do_confirm(so, request)
                
        except Exception as e:
            # 审批模块不可用，直接确认
            return self._do_confirm(so, request)
    
    @action(detail=True, methods=['post'])
    def confirm(self, request, pk=None):
        """直接确认销售订单（跳过审批）"""
        so = self.get_object()
        if so.status not in ['DRAFT', 'PENDING']:
            return Response(
                {'error': '只能确认草稿或待审批状态的订单'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        return self._do_confirm(so, request)
    
    def _do_confirm(self, so, request):
        """执行订单确认逻辑"""
        so.status = 'CONFIRMED'
        so.save()
        
        # Auto-create AR - 使用含税金额（避免重复创建）
        from apps.finance.models import AccountReceivable, PaymentSchedule
        
        # 检查是否已存在该SO的应收账款
        existing_ar = AccountReceivable.objects.filter(so=so, is_deleted=False).first()
        if not existing_ar:
            AccountReceivable.objects.create(
                customer=so.customer,
                so=so,
                project=so.project,
                invoice_date=so.order_date,
                amount_due=so.total_with_tax or so.total_amount,
                due_date=request.data.get('due_date', so.delivery_date),
                created_by=request.user
            )
        
        # 自动生成付款计划（避免重复创建）
        existing_schedules = PaymentSchedule.objects.filter(sales_order=so, is_deleted=False).exists()
        if existing_schedules:
            schedules = []
        else:
            schedules = PaymentSchedule.generate_from_sales_order(so)
        
        response_data = SalesOrderSerializer(so).data
        response_data['payment_schedules_count'] = len(schedules)
        
        return Response(response_data)
    
    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        """Cancel sales order."""
        so = self.get_object()
        if so.status in ['COMPLETED', 'CANCELLED']:
            return Response(
                {'error': '无法取消已完成或已取消的订单'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        so.status = 'CANCELLED'
        so.save()
        return Response(SalesOrderSerializer(so).data)
    
    @action(detail=True, methods=['post'], url_path='return_to_draft')
    def return_to_draft(self, request, pk=None):
        """将已确认的订单退回草稿状态（用于补充明细）"""
        so = self.get_object()
        
        # 只能退回确认状态的订单
        if so.status != 'CONFIRMED':
            return Response(
                {'error': '只能将已确认状态的订单退回草稿'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # 检查是否已有发货记录
        if so.deliveries.filter(is_deleted=False).exists():
            return Response(
                {'error': '订单已有发货记录，无法退回草稿'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        so.status = 'DRAFT'
        so.save()
        return Response(SalesOrderSerializer(so).data)
    
    @action(detail=True, methods=['get'])
    def generate_invoice(self, request, pk=None):
        """Generate PDF invoice for sales order."""
        so = self.get_object()
        
        try:
            from apps.finance.invoice_generator import InvoiceGenerator
            pdf_buffer = InvoiceGenerator.generate_invoice(so)
            
            response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="invoice_{so.order_no}.pdf"'
            return response
        except Exception as e:
            return Response(
                {'error': f'发票生成失败: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'], url_path='download-template')
    def download_template(self, request):
        """下载销售订单导入模板"""
        import io
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '销售订单'
        
        # 表头样式
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 表头 - 与创建表单保持一致
        headers = [
            '销售订单号', '客户订单号', '客户名称*', '关联项目',
            '订单日期', '交货日期*', '税率(%)', 
            '付款条款', '付款方式', '付款说明', '备注'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # 设置列宽
        widths = [15, 15, 25, 20, 12, 12, 10, 25, 12, 25, 30]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # 示例数据
        sample_data = [
            '', 'CUST-001', '示例客户', '示例项目',
            '2025-01-01', '2025-02-01', 13, 
            '30%预付/60%验收/10%质保', '电汇', '分3期付款', '示例备注'
        ]
        for col, value in enumerate(sample_data, 1):
            ws.cell(row=2, column=col, value=value)
        
        # 创建明细表
        ws2 = wb.create_sheet(title='订单明细')
        detail_headers = [
            '行号', '产品名称*', '规格型号', '单位', '数量*', '单价*', '备注'
        ]
        
        for col, header in enumerate(detail_headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # 设置列宽
        detail_widths = [8, 30, 25, 10, 12, 12, 30]
        for i, width in enumerate(detail_widths, 1):
            ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # 示例数据（非标定制产品，无需物料编码）
        detail_sample = [
            1, '定制产品A', '按图纸加工', '件', 1, 50000.00, ''
        ]
        for col, value in enumerate(detail_sample, 1):
            ws2.cell(row=2, column=col, value=value)
        
        # 创建说明表
        ws3 = wb.create_sheet(title='填写说明')
        instructions = [
            ['销售订单导入模板说明（适用于非标定制企业）'],
            [''],
            ['【销售订单表】'],
            ['- 销售订单号: 可选，留空则系统自动生成'],
            ['- 客户订单号: 可选，客户的订单编号，用于对账'],
            ['- 客户名称*: 必填，必须与系统中的客户名称完全匹配'],
            ['- 关联项目: 可选，后续可通过项目管理关联'],
            ['- 订单日期: 格式 YYYY-MM-DD，留空使用当天'],
            ['- 交货日期*: 必填，格式 YYYY-MM-DD'],
            ['- 税率(%): 0/1/3/6/9/13，默认13'],
            ['- 付款条款/付款方式/付款说明: 可选'],
            ['- 备注: 可选'],
            [''],
            ['【订单明细表】'],
            ['- 行号: 可选，用于一次导入多个订单时区分明细归属'],
            ['  * 如果主表只有1个订单：可不填，所有明细自动关联'],
            ['  * 如果主表有多个订单：按主表行号对应填写(1,2,3...)'],
            ['- 产品名称*: 必填，定制产品名称'],
            ['- 规格型号: 可选，如"按图纸加工"'],
            ['- 单位: 可选，默认"件"'],
            ['- 数量*: 必填，正数'],
            ['- 单价*: 必填，不含税单价'],
            ['- 备注: 可选'],
            [''],
            ['【非标定制企业使用说明】'],
            ['1. 销售订单阶段无需物料编码，直接填写产品名称'],
            ['2. 订单号留空，系统自动生成'],
            ['3. 后续流程：销售订单 → 创建项目 → 项目BOM → 采购申请'],
            ['4. 只需填写客户名称、交货日期、产品明细即可'],
            ['5. 一次导入多个订单时，明细表用"行号"区分归属'],
        ]
        for row_idx, row_data in enumerate(instructions, 1):
            for col_idx, value in enumerate(row_data, 1):
                ws3.cell(row=row_idx, column=col_idx, value=value)
        ws3.column_dimensions['A'].width = 60
        
        # 输出
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="sales_order_template.xlsx"'
        return response
    
    @action(detail=False, methods=['get'], url_path='export')
    def export_excel(self, request):
        """导出销售订单到Excel"""
        import io
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from django.http import HttpResponse
        
        # 获取查询参数过滤的数据
        queryset = self.filter_queryset(self.get_queryset())
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '销售订单'
        
        # 表头样式
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 表头 - 与创建表单保持一致
        headers = [
            '销售订单号', '客户订单号', '客户名称', '客户编码', '关联项目',
            '订单日期', '交货日期', '状态', '税率(%)', 
            '不含税金额', '税额', '含税总额',
            '付款条款', '付款方式', '付款说明', '备注', '创建时间'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # 数据行
        for row_idx, so in enumerate(queryset, 2):
            data = [
                so.order_no,
                so.customer_order_no or '',
                so.customer.name if so.customer else '',
                so.customer.code if so.customer else '',
                so.project.name if so.project else '',
                so.order_date.strftime('%Y-%m-%d') if so.order_date else '',
                so.delivery_date.strftime('%Y-%m-%d') if so.delivery_date else '',
                so.get_status_display(),
                so.tax_rate,
                float(so.total_amount),
                float(so.tax_amount),
                float(so.total_with_tax),
                so.get_payment_terms_display() if so.payment_terms else '',
                so.get_payment_method_display() if so.payment_method else '',
                so.payment_terms_detail or '',
                so.notes or '',
                so.created_at.strftime('%Y-%m-%d %H:%M') if so.created_at else ''
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
        
        # 设置列宽
        widths = [15, 15, 25, 15, 20, 12, 12, 10, 8, 15, 12, 15, 25, 12, 20, 30, 18]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # 创建明细表
        ws2 = wb.create_sheet(title='订单明细')
        detail_headers = [
            '订单号', '物料编码', '物料名称', '产品名称', '规格型号',
            '单位', '数量', '单价', '行金额', '已发货', '备注'
        ]
        
        for col, header in enumerate(detail_headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        detail_row = 2
        for so in queryset:
            for line in so.lines.filter(is_deleted=False):
                data = [
                    so.order_no,
                    line.item.sku if line.item else '',
                    line.item.name if line.item else '',
                    line.custom_name or '',
                    line.custom_spec or (line.item.spec if line.item else ''),
                    line.custom_unit or (line.item.unit if line.item else '件'),
                    float(line.qty),
                    float(line.unit_price),
                    float(line.line_amount),
                    float(line.delivered_qty),
                    line.notes or ''
                ]
                for col, value in enumerate(data, 1):
                    cell = ws2.cell(row=detail_row, column=col, value=value)
                    cell.border = thin_border
                detail_row += 1
        
        # 设置列宽
        detail_widths = [15, 15, 20, 20, 20, 8, 12, 12, 15, 12, 25]
        for i, width in enumerate(detail_widths, 1):
            ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # 输出
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="sales_orders_export.xlsx"'
        return response
    
    @action(detail=False, methods=['post'], url_path='import')
    def import_excel(self, request):
        """导入销售订单"""
        import io
        import openpyxl
        from django.db import transaction
        from decimal import Decimal, InvalidOperation
        from datetime import datetime
        from apps.masterdata.models import Customer, Item
        
        file = request.FILES.get('file')
        if not file:
            return Response({'error': '请上传文件'}, status=status.HTTP_400_BAD_REQUEST)
        
        if not file.name.endswith(('.xlsx', '.xls')):
            return Response({'error': '只支持Excel文件格式'}, status=status.HTTP_400_BAD_REQUEST)
        
        def parse_decimal(value, default=Decimal('0')):
            if value is None:
                return default
            try:
                return Decimal(str(value).replace(',', ''))
            except (InvalidOperation, ValueError):
                return default
        
        def parse_date(value):
            if value is None:
                return None
            if isinstance(value, datetime):
                return value.date()
            if isinstance(value, str):
                for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d']:
                    try:
                        return datetime.strptime(value.strip(), fmt).date()
                    except ValueError:
                        continue
            return None
        
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file.read()))
            
            success_count = 0
            update_count = 0
            skip_count = 0
            line_count = 0
            errors = []
            
            # 付款条款映射
            payment_terms_map = {
                '全款预付': 'FULL_PREPAY',
                '货到付款': 'COD',
                '月结30天': 'NET30',
                '月结60天': 'NET60',
                '月结90天': 'NET90',
                '30%预付/70%发货前': 'M_30_70',
                '30%预付/30%发货前/40%验收后': 'M_30_30_40',
                '30%预付/30%发货前/30%验收/10%质保': 'M_30_30_30_10',
                '30%预付/60%验收/10%质保': 'M_30_60_10',
                '50%预付/40%验收/10%质保': 'M_50_40_10',
                '40%预付/50%验收/10%质保': 'M_40_50_10',
                '20%预付/70%验收/10%质保': 'M_20_70_10',
                '自定义分期': 'CUSTOM',
                '其他': 'OTHER',
            }
            
            # 付款方式映射
            payment_method_map = {
                '电汇': 'WIRE',
                '承兑汇票': 'ACCEPTANCE',
                '支票': 'CHECK',
                '现金': 'CASH',
                '信用证': 'LC',
                '其他': 'OTHER',
            }
            
            # 缓存客户、项目和物料
            customer_cache = {c.name: c for c in Customer.objects.filter(is_deleted=False)}
            item_cache = {i.sku: i for i in Item.objects.filter(is_deleted=False)}
            
            # 项目缓存（按名称和编码）
            from apps.projects.models import Project
            project_cache = {}
            for p in Project.objects.filter(is_deleted=False):
                project_cache[p.name] = p
                if p.code:
                    project_cache[p.code] = p
            
            # 跟踪已处理的订单号
            processed_orders = set()
            order_map = {}  # order_no -> SalesOrder
            
            # Step 1: 导入订单头
            if '销售订单' in wb.sheetnames:
                sheet = wb['销售订单']
            else:
                sheet = wb.active
            
            headers = [str(cell.value).strip() if cell.value else '' for cell in sheet[1]]
            
            col_map = {
                '销售订单号': 'order_no',
                '订单号': 'order_no',
                '订单号*': 'order_no',
                '客户订单号': 'customer_order_no',
                '客户名称': 'customer_name',
                '客户名称*': 'customer_name',
                '关联项目': 'project_name',
                '订单日期': 'order_date',
                '交货日期': 'delivery_date',
                '交货日期*': 'delivery_date',
                '税率': 'tax_rate',
                '税率(%)': 'tax_rate',
                '付款条款': 'payment_terms',
                '付款方式': 'payment_method',
                '付款说明': 'payment_terms_detail',
                '备注': 'notes',
            }
            
            header_to_field = {}
            for idx, header in enumerate(headers):
                if header in col_map:
                    header_to_field[idx] = col_map[header]
            
            # 用于按主表行号索引订单（支持明细表用行号关联）
            order_by_row_idx = {}
            main_row_counter = 0
            
            with transaction.atomic():
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                    try:
                        data = {}
                        for col_idx, cell in enumerate(row):
                            if col_idx in header_to_field:
                                field = header_to_field[col_idx]
                                value = cell.value
                                if value is not None:
                                    data[field] = value
                        
                        # 跳过空行
                        if not data or all(v is None or str(v).strip() == '' for v in data.values()):
                            continue
                        
                        main_row_counter += 1  # 记录有效行号
                        
                        order_no = str(data.get('order_no', '')).strip()
                        
                        # 订单号可选，留空则后面自动生成
                        # 检查重复（只在有订单号时检查）
                        if order_no and order_no in processed_orders:
                            skip_count += 1
                            continue
                        if order_no:
                            processed_orders.add(order_no)
                        
                        customer_name = str(data.get('customer_name', '')).strip()
                        if not customer_name:
                            errors.append({'row': row_idx, 'error': '客户名称不能为空'})
                            continue
                        
                        customer = customer_cache.get(customer_name)
                        if not customer:
                            errors.append({'row': row_idx, 'error': f'客户 "{customer_name}" 不存在'})
                            continue
                        
                        delivery_date = parse_date(data.get('delivery_date'))
                        if not delivery_date:
                            errors.append({'row': row_idx, 'error': '交货日期格式错误'})
                            continue
                        
                        order_date = parse_date(data.get('order_date'))
                        
                        tax_rate = int(data.get('tax_rate', 13))
                        if tax_rate not in [0, 1, 3, 6, 9, 13]:
                            tax_rate = 13
                        
                        payment_terms_str = str(data.get('payment_terms', '')).strip()
                        payment_terms = payment_terms_map.get(payment_terms_str, 'M_30_30_30_10')
                        
                        payment_method_str = str(data.get('payment_method', '')).strip()
                        payment_method = payment_method_map.get(payment_method_str, 'WIRE')
                        
                        notes = str(data.get('notes', '')).strip()
                        
                        # 新增字段
                        customer_order_no = str(data.get('customer_order_no', '')).strip()
                        payment_terms_detail = str(data.get('payment_terms_detail', '')).strip()
                        
                        # 关联项目
                        project_name = str(data.get('project_name', '')).strip()
                        project = project_cache.get(project_name) if project_name else None
                        
                        # 检查是否已存在（只在有订单号时检查）
                        existing = SalesOrder.objects.filter(order_no=order_no).first() if order_no else None
                        if existing:
                            if existing.is_deleted:
                                existing.is_deleted = False
                                existing.deleted_at = None
                            existing.customer = customer
                            existing.customer_order_no = customer_order_no
                            existing.project = project
                            existing.delivery_date = delivery_date
                            existing.tax_rate = tax_rate
                            existing.payment_terms = payment_terms
                            existing.payment_method = payment_method
                            existing.payment_terms_detail = payment_terms_detail
                            existing.notes = notes
                            existing.save()
                            # 删除旧明细，等待重新导入
                            existing.lines.all().delete()
                            order_map[existing.order_no] = existing
                            order_by_row_idx[main_row_counter] = existing
                            update_count += 1
                        else:
                            so = SalesOrder.objects.create(
                                order_no=order_no if order_no else None,  # 留空自动生成
                                customer_order_no=customer_order_no,
                                customer=customer,
                                project=project,
                                delivery_date=delivery_date,
                                tax_rate=tax_rate,
                                payment_terms=payment_terms,
                                payment_method=payment_method,
                                payment_terms_detail=payment_terms_detail,
                                notes=notes,
                                created_by=request.user
                            )
                            # 更新订单日期
                            if order_date:
                                SalesOrder.objects.filter(pk=so.pk).update(order_date=order_date)
                            order_map[so.order_no] = so
                            order_by_row_idx[main_row_counter] = so
                            success_count += 1
                    
                    except Exception as e:
                        errors.append({'row': row_idx, 'error': str(e)})
                
                # Step 2: 导入订单明细
                # 兼容多种工作表名称
                detail_sheet_names = ['订单明细', 'Sheet2', '明细', 'Details', 'Lines']
                detail_sheet = None
                for sheet_name in detail_sheet_names:
                    if sheet_name in wb.sheetnames:
                        detail_sheet = wb[sheet_name]
                        break
                
                # 如果没有找到明细sheet，尝试从主sheet的额外列读取明细
                # 检查主sheet是否有明细列（产品名称、数量、单价）
                if not detail_sheet and len(order_map) > 0:
                    # 检查主sheet是否包含明细列
                    main_detail_cols = {'产品名称', '产品名称*', '数量', '数量*', '单价', '单价*'}
                    has_detail_cols = any(h in main_detail_cols for h in headers)
                    
                    if has_detail_cols:
                        # 主sheet包含明细列，尝试从主sheet读取明细
                        main_detail_col_map = {
                            '产品名称': 'product_name',
                            '产品名称*': 'product_name',
                            '规格型号': 'spec',
                            '单位': 'unit',
                            '数量': 'qty',
                            '数量*': 'qty',
                            '单价': 'unit_price',
                            '单价*': 'unit_price',
                        }
                        
                        main_detail_header_to_field = {}
                        for idx, header in enumerate(headers):
                            if header in main_detail_col_map:
                                main_detail_header_to_field[idx] = main_detail_col_map[header]
                        
                        if main_detail_header_to_field:
                            # 重新遍历主sheet获取明细
                            row_counter = 0
                            for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                                try:
                                    # 检查是否为空行
                                    row_data = {}
                                    for col_idx, cell in enumerate(row):
                                        if col_idx in header_to_field:
                                            if cell.value is not None:
                                                row_data[header_to_field[col_idx]] = cell.value
                                    
                                    if not row_data or all(v is None or str(v).strip() == '' for v in row_data.values()):
                                        continue
                                    
                                    row_counter += 1
                                    so = order_by_row_idx.get(row_counter)
                                    if not so:
                                        continue
                                    
                                    # 获取明细数据
                                    detail_data = {}
                                    for col_idx, cell in enumerate(row):
                                        if col_idx in main_detail_header_to_field:
                                            field = main_detail_header_to_field[col_idx]
                                            if cell.value is not None:
                                                detail_data[field] = cell.value
                                    
                                    if not detail_data:
                                        continue
                                    
                                    qty = parse_decimal(detail_data.get('qty'))
                                    unit_price = parse_decimal(detail_data.get('unit_price'))
                                    product_name = str(detail_data.get('product_name', '')).strip()
                                    spec = str(detail_data.get('spec', '')).strip()
                                    unit = str(detail_data.get('unit', '件')).strip()
                                    
                                    if qty > 0 and unit_price > 0 and product_name:
                                        line_amount = qty * unit_price
                                        SalesOrderLine.objects.create(
                                            so=so,
                                            item=None,
                                            custom_name=product_name,
                                            custom_spec=spec,
                                            custom_unit=unit,
                                            qty=qty,
                                            unit_price=unit_price,
                                            line_amount=line_amount,
                                            notes='',
                                            created_by=request.user
                                        )
                                        line_count += 1
                                except Exception as e:
                                    errors.append({'row': row_idx, 'error': f'读取明细失败: {str(e)}'})
                    else:
                        # 没有明细sheet，也没有明细列，给出警告
                        errors.append({
                            'row': 0, 
                            'error': f'警告：未找到订单明细工作表（支持的名称：订单明细、Sheet2、明细），已创建 {len(order_map)} 个订单但没有产品明细。请手动添加明细或重新导入包含明细sheet的文件。'
                        })
                
                if detail_sheet:
                    detail_headers = [str(cell.value).strip() if cell.value else '' for cell in detail_sheet[1]]
                    
                    detail_col_map = {
                        '行号': 'line_no',
                        '订单号': 'order_no',
                        '订单号*': 'order_no',
                        '销售订单号': 'order_no',
                        '物料编码': 'item_sku',
                        '产品名称': 'product_name',
                        '产品名称*': 'product_name',
                        '规格型号': 'spec',
                        '单位': 'unit',
                        '数量': 'qty',
                        '数量*': 'qty',
                        '单价': 'unit_price',
                        '单价*': 'unit_price',
                        '备注': 'notes',
                    }
                    
                    detail_header_to_field = {}
                    for idx, header in enumerate(detail_headers):
                        if header in detail_col_map:
                            detail_header_to_field[idx] = detail_col_map[header]
                    
                    for row_idx, row in enumerate(detail_sheet.iter_rows(min_row=2), start=2):
                        try:
                            data = {}
                            for col_idx, cell in enumerate(row):
                                if col_idx in detail_header_to_field:
                                    field = detail_header_to_field[col_idx]
                                    value = cell.value
                                    if value is not None:
                                        data[field] = value
                            
                            if not data or all(v is None or str(v).strip() == '' for v in data.values()):
                                continue
                            
                            # 获取关联订单：优先行号，其次订单号，最后默认第一个
                            line_no_str = str(data.get('line_no', '')).strip()
                            order_no = str(data.get('order_no', '')).strip()
                            so = None
                            
                            # 方法1：按行号关联
                            if line_no_str:
                                try:
                                    line_no = int(line_no_str)
                                    so = order_by_row_idx.get(line_no)
                                except ValueError:
                                    pass
                            
                            # 方法2：按订单号关联
                            if not so and order_no:
                                so = order_map.get(order_no)
                                if not so:
                                    so = SalesOrder.objects.filter(order_no=order_no).first()
                            
                            # 方法3：默认关联到第一个订单（仅当只有一个订单时）
                            if not so and len(order_map) == 1:
                                so = list(order_map.values())[0]
                            
                            if not so:
                                errors.append({'row': row_idx, 'sheet': '订单明细', 'error': '无法关联订单，请填写行号或订单号'})
                                continue
                            
                            qty = parse_decimal(data.get('qty'))
                            unit_price = parse_decimal(data.get('unit_price'))
                            
                            if qty <= 0:
                                errors.append({'row': row_idx, 'sheet': '订单明细', 'error': '数量必须大于0'})
                                continue
                            
                            if unit_price <= 0:
                                errors.append({'row': row_idx, 'sheet': '订单明细', 'error': '单价必须大于0'})
                                continue
                            
                            item_sku = str(data.get('item_sku', '')).strip()
                            item = item_cache.get(item_sku) if item_sku else None
                            
                            product_name = str(data.get('product_name', '')).strip()
                            spec = str(data.get('spec', '')).strip()
                            unit = str(data.get('unit', '件')).strip()
                            notes = str(data.get('notes', '')).strip()
                            
                            # 如果没有物料也没有产品名称，跳过
                            if not item and not product_name:
                                errors.append({'row': row_idx, 'sheet': '订单明细', 'error': '产品名称不能为空'})
                                continue
                            
                            line_amount = qty * unit_price
                            
                            SalesOrderLine.objects.create(
                                so=so,
                                item=item,
                                custom_name=product_name if not item else '',
                                custom_spec=spec if not item else '',
                                custom_unit=unit if not item else '',
                                qty=qty,
                                unit_price=unit_price,
                                line_amount=line_amount,
                                notes=notes,
                                created_by=request.user
                            )
                            line_count += 1
                        
                        except Exception as e:
                            errors.append({'row': row_idx, 'sheet': '订单明细', 'error': str(e)})
                
                # 更新订单金额
                for order_no, so in order_map.items():
                    total_amount = sum(
                        line.line_amount for line in so.lines.filter(is_deleted=False)
                    )
                    so.total_amount = total_amount
                    so.tax_amount = total_amount * so.tax_rate / 100
                    so.total_with_tax = total_amount + so.tax_amount
                    so.save()
            
            return Response({
                'success_count': success_count,
                'update_count': update_count,
                'skip_count': skip_count,
                'line_count': line_count,
                'errors': errors[:20] if errors else []
            })
        
        except Exception as e:
            return Response(
                {'error': f'导入失败: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['post'], url_path='bulk-delete')
    def bulk_delete(self, request):
        """批量删除销售订单"""
        ids = request.data.get('ids', [])
        if not ids:
            return Response({'error': '请选择要删除的订单'}, status=status.HTTP_400_BAD_REQUEST)
        
        # 只删除草稿和已取消状态的订单
        deletable_orders = SalesOrder.objects.filter(
            id__in=ids,
            status__in=['DRAFT', 'CANCELLED', 'REJECTED'],
            is_deleted=False
        )
        
        delete_count = deletable_orders.count()
        skip_count = len(ids) - delete_count
        
        # 软删除
        deletable_orders.update(is_deleted=True)
        
        message = f'成功删除 {delete_count} 个订单'
        if skip_count > 0:
            message += f'，跳过 {skip_count} 个（只能删除草稿/已取消/已拒绝状态的订单）'
        
        return Response({
            'message': message,
            'delete_count': delete_count,
            'skip_count': skip_count
        })


class SalesOrderLineViewSet(SoftDeleteMixin, UserTrackingMixin, viewsets.ModelViewSet):
    """
    ViewSet for SalesOrderLine management.
    """
    queryset = SalesOrderLine.objects.all()
    serializer_class = SalesOrderLineSerializer
    filterset_fields = ['so', 'item', 'is_deleted']
    search_fields = ['item__sku', 'item__name']


class DeliveryOrderViewSet(SoftDeleteMixin, UserTrackingMixin, DataPermissionMixin, viewsets.ModelViewSet):
    """
    ViewSet for DeliveryOrder management.
    
    发货流程:
    1. 销售发货通知 (DRAFT) - 销售创建发货单
    2. 提交审批 (PENDING) - 进入审批中心，审批步骤由流程配置决定
    3. 审批通过后进入操作流程:
       - 仓库备货 (PREPARING)
       - 采购预约物流 (LOGISTICS_BOOKING)
       - 客户签署送货单 (CUSTOMER_SIGNING)
       - 采购上传送货单 (UPLOADING_RECEIPT)
       - 项目确认 (PROJECT_CONFIRMING)
    4. 完成 (COMPLETED)
    
    注意: 审批步骤由审批中心的流程配置决定，修改流程配置会自动影响审批流程
    """
    queryset = DeliveryOrder.objects.all()
    serializer_class = DeliveryOrderSerializer
    filterset_fields = ['so', 'warehouse', 'status', 'is_deleted']
    search_fields = ['delivery_no']
    ordering_fields = ['delivery_date', 'created_at']
    
    def _calculate_delivery_amount(self, delivery):
        """计算发货单总金额"""
        total = 0
        for line in delivery.lines.filter(is_deleted=False):
            total += line.qty * line.so_line.unit_price
        return total
    
    # 提交审批 - 使用审批中心的流程配置
    @action(detail=True, methods=['post'])
    def submit(self, request, pk=None):
        """提交发货申请 - 进入审批中心，审批步骤由流程配置决定"""
        delivery = self.get_object()
        if delivery.status not in ['DRAFT', 'REJECTED']:
            return Response(
                {'error': '只能提交草稿或已拒绝状态的发货单'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # 计算金额用于流程路由
        amount = self._calculate_delivery_amount(delivery)
        
        try:
            from apps.core.workflow.services import WorkflowService
            
            instance, error = WorkflowService.start_workflow(
                business_type='DELIVERY_ORDER',
                business_id=delivery.id,
                business_no=delivery.delivery_no,
                submitter=request.user,
                amount=amount
            )
            
            if instance:
                delivery.status = 'PENDING'
                delivery.rejection_reason = ''  # 清除之前的拒绝原因
                delivery.save()
                return Response({
                    **DeliveryOrderSerializer(delivery).data,
                    'workflow_started': True,
                    'workflow_id': instance.id,
                    'message': '已提交审批，请在审批中心查看审批进度'
                })
            else:
                # 未配置审批流程，直接进入备货
                delivery.status = 'PREPARING'
                delivery.save()
                return Response({
                    **DeliveryOrderSerializer(delivery).data,
                    'workflow_started': False,
                    'message': error or '未配置审批流程，已直接进入备货环节'
                })
                
        except Exception as e:
            # 审批模块不可用，直接进入备货
            delivery.status = 'PREPARING'
            delivery.save()
            return Response({
                **DeliveryOrderSerializer(delivery).data,
                'workflow_started': False,
                'message': f'提交成功，但工作流服务异常: {e}'
            })
    
    # 获取当前审批状态
    @action(detail=True, methods=['get'])
    def workflow_status(self, request, pk=None):
        """获取发货单的审批状态"""
        delivery = self.get_object()
        
        try:
            from apps.core.workflow.models import WorkflowInstance, WorkflowTask
            
            instance = WorkflowInstance.objects.filter(
                business_type='DELIVERY_ORDER',
                business_id=delivery.id,
                is_deleted=False
            ).order_by('-created_at').first()
            
            if instance:
                pending_task = instance.tasks.filter(status='PENDING', is_deleted=False).first()
                return Response({
                    'has_workflow': True,
                    'workflow_id': instance.id,
                    'workflow_status': instance.status,
                    'current_step': instance.current_step,
                    'pending_task': {
                        'id': pending_task.id,
                        'step_name': pending_task.step.name,
                        'assignee': pending_task.assignee.username if pending_task.assignee else None,
                    } if pending_task else None
                })
            else:
                return Response({
                    'has_workflow': False,
                    'message': '无审批流程'
                })
        except Exception as e:
            return Response({
                'has_workflow': False,
                'error': str(e)
            })
    
    # 仓库确认备货完成
    @action(detail=True, methods=['post'])
    def confirm_prepared(self, request, pk=None):
        """仓库确认备货完成"""
        delivery = self.get_object()
        # 允许从 APPROVED 或 PREPARING 状态确认备货
        if delivery.status not in ['APPROVED', 'PREPARING']:
            return Response(
                {'error': '当前状态不能确认备货'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # 创建出库记录
        with transaction.atomic():
            from apps.inventory.models import StockMove
            from django.utils import timezone
            
            for line in delivery.lines.filter(is_deleted=False):
                StockMove.objects.create(
                    item=line.item,
                    warehouse_from=delivery.warehouse,
                    qty=line.qty,
                    unit_cost=line.so_line.unit_price,
                    move_type='OUT_SALES',
                    reference_type='DeliveryOrder',
                    reference_id=delivery.id,
                    project=delivery.so.project,
                    move_date=timezone.now().date(),
                    status='COMPLETED',
                    created_by=request.user
                )
                
                # 更新销售订单发货数量
                line.so_line.delivered_qty += line.qty
                line.so_line.save()
        
        delivery.status = 'LOGISTICS_BOOKING'
        delivery.save()
        
        return Response({
            **DeliveryOrderSerializer(delivery).data,
            'message': '备货完成，请采购预约物流'
        })
    
    # 采购确认物流
    @action(detail=True, methods=['post'])
    def confirm_logistics(self, request, pk=None):
        """采购确认物流信息"""
        delivery = self.get_object()
        if delivery.status != 'LOGISTICS_BOOKING':
            return Response(
                {'error': '当前状态不能确认物流'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # 更新物流信息
        logistics_company = request.data.get('logistics_company')
        tracking_number = request.data.get('tracking_number')
        logistics_cost = request.data.get('logistics_cost')
        
        if logistics_company:
            delivery.logistics_company = logistics_company
        if tracking_number:
            delivery.tracking_number = tracking_number
        if logistics_cost:
            delivery.logistics_cost = logistics_cost
        
        delivery.status = 'CUSTOMER_SIGNING'
        delivery.save()
        
        return Response({
            **DeliveryOrderSerializer(delivery).data,
            'message': '物流已预约，等待客户签收'
        })
    
    # Step 6 -> 7: 客户签收
    @action(detail=True, methods=['post'])
    def confirm_signed(self, request, pk=None):
        """确认客户已签收"""
        delivery = self.get_object()
        if delivery.status != 'CUSTOMER_SIGNING':
            return Response(
                {'error': '当前状态不能确认签收'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        signed_by = request.data.get('signed_by')
        signed_date = request.data.get('signed_date')
        
        if signed_by:
            delivery.signed_by = signed_by
        if signed_date:
            delivery.signed_date = signed_date
        
        delivery.status = 'UPLOADING_RECEIPT'
        delivery.save()
        
        return Response({
            **DeliveryOrderSerializer(delivery).data,
            'message': '已确认签收，请上传送货单'
        })
    
    # Step 7 -> 8: 上传送货单
    @action(detail=True, methods=['post'])
    def upload_receipt(self, request, pk=None):
        """上传签收单据"""
        delivery = self.get_object()
        if delivery.status != 'UPLOADING_RECEIPT':
            return Response(
                {'error': '当前状态不能上传送货单'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if 'signed_receipt' in request.FILES:
            delivery.signed_receipt = request.FILES['signed_receipt']
        
        delivery.status = 'PROJECT_CONFIRMING'
        delivery.save()
        
        return Response({
            **DeliveryOrderSerializer(delivery).data,
            'message': '送货单已上传，等待项目确认'
        })
    
    # Step 8 -> 9: 项目确认
    @action(detail=True, methods=['post'])
    def project_confirm(self, request, pk=None):
        """项目确认完成"""
        delivery = self.get_object()
        if delivery.status != 'PROJECT_CONFIRMING':
            return Response(
                {'error': '当前状态不能进行项目确认'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        from django.utils import timezone
        
        delivery.status = 'COMPLETED'
        delivery.actual_delivery_date = timezone.now().date()
        delivery.save()
        
        # 更新销售订单状态
        so = delivery.so
        all_delivered = all(
            line.delivered_qty >= line.qty
            for line in so.lines.filter(is_deleted=False)
        )
        if all_delivered:
            so.status = 'COMPLETED'
        else:
            so.status = 'PARTIAL'
        so.save()
        
        return Response({
            **DeliveryOrderSerializer(delivery).data,
            'message': '发货流程已完成'
        })
    
    # 拒绝操作（可在任意审批环节拒绝）
    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """拒绝发货申请"""
        delivery = self.get_object()
        if delivery.status in ['DRAFT', 'COMPLETED', 'REJECTED']:
            return Response(
                {'error': '当前状态不能拒绝'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        reason = request.data.get('reason', '')
        delivery.status = 'REJECTED'
        delivery.rejection_reason = reason
        delivery.save()
        
        return Response({
            **DeliveryOrderSerializer(delivery).data,
            'message': '已拒绝'
        })


class DeliveryOrderLineViewSet(SoftDeleteMixin, UserTrackingMixin, viewsets.ModelViewSet):
    """
    ViewSet for DeliveryOrderLine management.
    """
    queryset = DeliveryOrderLine.objects.all()
    serializer_class = DeliveryOrderLineSerializer
    filterset_fields = ['delivery', 'item', 'is_deleted']
    search_fields = ['item__sku', 'item__name']


class SalesContractViewSet(SoftDeleteMixin, UserTrackingMixin, viewsets.ModelViewSet):
    """
    ViewSet for SalesContract management.
    """
    queryset = SalesContract.objects.all()
    serializer_class = SalesContractSerializer
    filterset_fields = ['so', 'customer', 'project', 'status', 'is_deleted']
    search_fields = ['contract_no', 'title']
    ordering_fields = ['contract_date', 'created_at']
    
    @action(detail=False, methods=['post'])
    def create_from_so(self, request):
        """从销售订单创建合同."""
        so_id = request.data.get('so_id')
        if not so_id:
            return Response(
                {'error': '请选择销售订单'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            so = SalesOrder.objects.get(id=so_id, is_deleted=False)
        except SalesOrder.DoesNotExist:
            return Response(
                {'error': '销售订单不存在'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # 检查是否已有合同
        existing = SalesContract.objects.filter(so=so, is_deleted=False).first()
        if existing:
            return Response(
                {'error': f'该销售订单已存在合同: {existing.contract_no}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        from django.utils import timezone
        
        contract = SalesContract.objects.create(
            so=so,
            customer=so.customer,
            project=so.project,
            title=f'{so.customer.name}销售合同',
            contract_date=timezone.now().date(),
            total_amount=so.total_amount,
            tax_rate=so.tax_rate,
            tax_amount=so.tax_amount,
            total_with_tax=so.total_with_tax,
            payment_terms=self._get_payment_terms_text(so),
            delivery_terms=f'交货日期：{so.delivery_date}',
            created_by=request.user
        )
        
        return Response(SalesContractSerializer(contract).data, status=status.HTTP_201_CREATED)
    
    def _get_payment_terms_text(self, so):
        """生成付款条款文本."""
        terms_map = {
            'FULL_PREPAY': '全款预付',
            'COD': '货到付款',
            'NET30': '月结30天',
            'NET60': '月结60天',
            'NET90': '月结90天',
            'M_30_70': '30%预付/70%发货前',
            'M_30_30_40': '30%预付/30%发货前/40%验收后',
            'M_30_30_30_10': '30%预付/30%发货前/30%验收/10%质保',
            'M_30_60_10': '30%预付/60%验收/10%质保',
            'M_50_40_10': '50%预付/40%验收/10%质保',
            'M_40_50_10': '40%预付/50%验收/10%质保',
            'M_20_70_10': '20%预付/70%验收/10%质保',
            'CUSTOM': '自定义分期',
            'OTHER': '其他',
        }
        base_text = terms_map.get(so.payment_terms, '30%预付/30%发货前/30%验收/10%质保')
        if so.payment_terms_detail:
            base_text += f'（{so.payment_terms_detail}）'
        return base_text
    
    @action(detail=True, methods=['post'])
    def submit(self, request, pk=None):
        """提交合同审批 - 审批步骤由流程配置决定"""
        contract = self.get_object()
        if contract.status not in ['DRAFT', 'REJECTED']:
            return Response(
                {'error': '只能提交草稿或已拒绝状态的合同'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        amount = contract.total_with_tax or contract.total_amount or 0
        
        try:
            from apps.core.workflow.services import WorkflowService
            
            instance, error = WorkflowService.start_workflow(
                business_type='SALES_CONTRACT',
                business_id=contract.id,
                business_no=contract.contract_no,
                submitter=request.user,
                amount=amount
            )
            
            if instance:
                contract.status = 'PENDING'
                contract.save()
                return Response({
                    **SalesContractSerializer(contract).data,
                    'workflow_started': True,
                    'workflow_id': instance.id,
                    'message': '已提交审批，请在审批中心查看审批进度'
                })
            else:
                contract.status = 'ACTIVE'
                contract.save()
                return Response({
                    **SalesContractSerializer(contract).data,
                    'workflow_started': False,
                    'message': error or '未配置审批流程，合同已直接生效'
                })
                
        except Exception as e:
            contract.status = 'ACTIVE'
            contract.save()
            return Response({
                **SalesContractSerializer(contract).data,
                'workflow_started': False,
                'message': f'合同已生效，但工作流服务异常: {e}'
            })
    
    @action(detail=True, methods=['get'])
    def print_preview(self, request, pk=None):
        """获取合同打印预览数据."""
        contract = self.get_object()
        so = contract.so
        
        # 获取公司信息
        from apps.core.models import SystemConfig
        company_config = SystemConfig.get_config('company', {})
        
        # 获取订单明细
        lines = []
        for line in so.lines.filter(is_deleted=False):
            lines.append({
                'item_sku': line.item.sku if line.item else '',
                'item_name': line.item.name if line.item else line.custom_name,
                'specification': (line.item.spec if line.item else line.custom_spec) or '',
                'unit': (line.item.get_unit_display() if line.item else line.custom_unit) or '件',
                'qty': float(line.qty),
                'unit_price': float(line.unit_price),
                'line_amount': float(line.line_amount),
            })
        
        return Response({
            'contract': SalesContractSerializer(contract).data,
            'company': {
                'name': company_config.get('name', ''),
                'address': company_config.get('address', ''),
                'phone': company_config.get('phone', ''),
                'fax': company_config.get('fax', ''),
                'bank_name': company_config.get('bank_name', ''),
                'bank_account': company_config.get('bank_account', ''),
            },
            'customer': {
                'name': contract.customer.name,
                'address': contract.customer.address or '',
                'contact': contract.customer.contact or '',
                'phone': contract.customer.phone or '',
            },
            'lines': lines,
        })
    
    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """审批合同."""
        contract = self.get_object()
        if contract.status not in ['DRAFT', 'PENDING']:
            return Response(
                {'error': '只能审批草稿或待审批状态的合同'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        contract.status = 'APPROVED'
        contract.save()
        return Response(SalesContractSerializer(contract).data)
    
    @action(detail=True, methods=['post'])
    def sign(self, request, pk=None):
        """签署合同."""
        contract = self.get_object()
        if contract.status != 'APPROVED':
            return Response(
                {'error': '只能签署已审批的合同'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        from django.utils import timezone
        
        contract.buyer_signer = request.data.get('buyer_signer', '')
        contract.seller_signer = request.data.get('seller_signer', '')
        contract.signed_date = timezone.now().date()
        contract.status = 'SIGNED'
        contract.save()
        return Response(SalesContractSerializer(contract).data)
