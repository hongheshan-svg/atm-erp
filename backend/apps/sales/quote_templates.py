"""
报价单模板管理
Quote Template Management

功能：支持灵活的报价单模板配置和生成

非标自动化行业特点：
- 每个项目报价内容差异大，建议创建灵活的列配置
- 可按客户/行业创建不同模板
- 支持自定义表头、表尾、备注条款
- 模板仅作为格式参考，具体内容由项目报价决定

建议配置：
- 设备清单模板（设备名称、规格、数量、单价等）
- 工程服务模板（服务项目、工时、单价等）
- 综合报价模板（设备+安装+调试+培训等）
"""
import os
from datetime import datetime
from decimal import Decimal

from django.conf import settings
from django.db import models
from rest_framework import serializers, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from apps.core.mixins import SoftDeleteMixin, UserTrackingMixin
from apps.core.models import BaseModel


class QuoteTemplate(BaseModel):
    """
    报价单模板
    """
    TEMPLATE_FORMATS = [
        ('EXCEL', 'Excel模板'),
        ('PDF', 'PDF模板'),
        ('WORD', 'Word模板'),
        ('HTML', 'HTML模板'),
    ]

    code = models.CharField(max_length=50, unique=True, verbose_name='模板编码')
    name = models.CharField(max_length=100, verbose_name='模板名称')
    format = models.CharField(
        max_length=20,
        choices=TEMPLATE_FORMATS,
        default='EXCEL',
        verbose_name='模板格式'
    )
    template_file = models.FileField(
        upload_to='templates/quotes/',
        blank=True,
        null=True,
        verbose_name='模板文件'
    )
    html_content = models.TextField(blank=True, verbose_name='HTML模板内容')
    header_config = models.JSONField(
        default=dict,
        blank=True,
        verbose_name='表头配置',
        help_text='公司名称、Logo、联系方式等'
    )
    footer_config = models.JSONField(
        default=dict,
        blank=True,
        verbose_name='表尾配置',
        help_text='备注、条款、签章等'
    )
    column_config = models.JSONField(
        default=list,
        blank=True,
        verbose_name='列配置',
        help_text='报价明细列定义'
    )
    style_config = models.JSONField(
        default=dict,
        blank=True,
        verbose_name='样式配置'
    )
    is_default = models.BooleanField(default=False, verbose_name='默认模板')
    is_enabled = models.BooleanField(default=True, verbose_name='是否启用')
    description = models.TextField(blank=True, verbose_name='模板说明')

    class Meta:
        db_table = 'sales_quote_template'
        verbose_name = '报价单模板'
        verbose_name_plural = verbose_name
        ordering = ['code']

    def __str__(self):
        return f'{self.code} - {self.name}'

    def save(self, *args, **kwargs):
        # 如果设为默认，取消其他默认
        if self.is_default:
            QuoteTemplate.objects.filter(is_default=True).exclude(pk=self.pk).update(is_default=False)
        super().save(*args, **kwargs)


class QuoteHistory(BaseModel):
    """
    报价单生成历史
    """
    template = models.ForeignKey(
        QuoteTemplate,
        on_delete=models.SET_NULL,
        null=True,
        related_name='histories',
        verbose_name='使用模板'
    )
    quote_no = models.CharField(max_length=50, verbose_name='报价单号')
    customer_name = models.CharField(max_length=200, verbose_name='客户名称')
    project_name = models.CharField(max_length=200, blank=True, verbose_name='项目名称')
    total_amount = models.DecimalField(
        max_digits=18,
        decimal_places=2,
        default=0,
        verbose_name='报价总额'
    )
    currency = models.CharField(max_length=10, default='CNY', verbose_name='币种')
    valid_until = models.DateField(null=True, blank=True, verbose_name='有效期至')
    quote_data = models.JSONField(default=dict, verbose_name='报价数据')
    generated_file = models.FileField(
        upload_to='quotes/generated/',
        blank=True,
        null=True,
        verbose_name='生成的文件'
    )
    status = models.CharField(
        max_length=20,
        choices=[
            ('DRAFT', '草稿'),
            ('SENT', '已发送'),
            ('ACCEPTED', '已接受'),
            ('REJECTED', '已拒绝'),
            ('EXPIRED', '已过期'),
        ],
        default='DRAFT',
        verbose_name='状态'
    )

    class Meta:
        db_table = 'sales_quote_history'
        verbose_name = '报价单历史'
        verbose_name_plural = verbose_name
        ordering = ['-created_at']

    def __str__(self):
        return f'{self.quote_no} - {self.customer_name}'


# =====================
# Quote Generation Service
# =====================

class QuoteGenerationService:
    """报价单生成服务"""

    def __init__(self, template: QuoteTemplate):
        self.template = template

    def generate_excel(self, quote_data: dict, output_path: str = None) -> str:
        """生成Excel报价单"""
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError('openpyxl 库未安装')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '报价单'

        # 样式定义
        header_font = Font(name='微软雅黑', size=16, bold=True)
        title_font = Font(name='微软雅黑', size=11, bold=True)
        normal_font = Font(name='微软雅黑', size=10)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')

        row = 1

        # 公司信息头
        header_config = self.template.header_config or {}
        company_name = header_config.get('company_name', '公司名称')
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = company_name
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 1

        # 报价单标题
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = '报 价 单'
        ws[f'A{row}'].font = Font(name='微软雅黑', size=14, bold=True)
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 2

        # 基本信息
        info_items = [
            ('报价单号', quote_data.get('quote_no', '')),
            ('客户名称', quote_data.get('customer_name', '')),
            ('项目名称', quote_data.get('project_name', '')),
            ('报价日期', quote_data.get('quote_date', datetime.now().strftime('%Y-%m-%d'))),
            ('有效期至', quote_data.get('valid_until', '')),
            ('联系人', quote_data.get('contact_person', '')),
            ('联系电话', quote_data.get('contact_phone', '')),
        ]

        for label, value in info_items:
            if value:
                ws[f'A{row}'] = label + '：'
                ws[f'A{row}'].font = title_font
                ws[f'B{row}'] = value
                ws[f'B{row}'].font = normal_font
                row += 1

        row += 1

        # 列配置
        column_config = self.template.column_config or [
            {'key': 'index', 'label': '序号', 'width': 8},
            {'key': 'name', 'label': '名称', 'width': 30},
            {'key': 'spec', 'label': '规格型号', 'width': 25},
            {'key': 'unit', 'label': '单位', 'width': 10},
            {'key': 'quantity', 'label': '数量', 'width': 12},
            {'key': 'unit_price', 'label': '单价', 'width': 15},
            {'key': 'amount', 'label': '金额', 'width': 18},
            {'key': 'remark', 'label': '备注', 'width': 20},
        ]

        # 设置列宽
        for idx, col in enumerate(column_config):
            ws.column_dimensions[get_column_letter(idx + 1)].width = col.get('width', 15)

        # 表头
        header_row = row
        for idx, col in enumerate(column_config):
            cell = ws.cell(row=row, column=idx + 1, value=col['label'])
            cell.font = title_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            cell.fill = header_fill
        row += 1

        # 明细行
        items = quote_data.get('items', [])
        total_amount = Decimal('0')

        for item_idx, item in enumerate(items):
            for col_idx, col in enumerate(column_config):
                key = col['key']
                if key == 'index':
                    value = item_idx + 1
                else:
                    value = item.get(key, '')

                cell = ws.cell(row=row, column=col_idx + 1, value=value)
                cell.font = normal_font
                cell.border = thin_border

                if key in ('quantity', 'unit_price', 'amount'):
                    cell.alignment = Alignment(horizontal='right')
                    if key == 'amount' and value:
                        total_amount += Decimal(str(value))
                else:
                    cell.alignment = Alignment(horizontal='center' if key == 'index' else 'left')

            row += 1

        # 合计行
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = '合计'
        ws[f'A{row}'].font = title_font
        ws[f'A{row}'].alignment = Alignment(horizontal='right')
        ws[f'A{row}'].border = thin_border

        ws[f'G{row}'] = float(total_amount)
        ws[f'G{row}'].font = title_font
        ws[f'G{row}'].alignment = Alignment(horizontal='right')
        ws[f'G{row}'].border = thin_border
        ws[f'G{row}'].number_format = '#,##0.00'

        for col_idx in range(1, len(column_config) + 1):
            ws.cell(row=row, column=col_idx).border = thin_border

        row += 2

        # 备注
        footer_config = self.template.footer_config or {}
        remarks = footer_config.get('remarks', quote_data.get('remarks', ''))
        if remarks:
            ws[f'A{row}'] = '备注：'
            ws[f'A{row}'].font = title_font
            row += 1
            ws.merge_cells(f'A{row}:H{row}')
            ws[f'A{row}'] = remarks
            ws[f'A{row}'].font = normal_font
            row += 2

        # 条款
        terms = footer_config.get('terms', [])
        if terms:
            ws[f'A{row}'] = '条款与条件：'
            ws[f'A{row}'].font = title_font
            row += 1
            for term in terms:
                ws.merge_cells(f'A{row}:H{row}')
                ws[f'A{row}'] = f'• {term}'
                ws[f'A{row}'].font = normal_font
                row += 1
            row += 1

        # 签章区
        row += 1
        ws[f'A{row}'] = '供方（盖章）：'
        ws[f'A{row}'].font = title_font
        ws[f'E{row}'] = '需方（盖章）：'
        ws[f'E{row}'].font = title_font
        row += 3
        ws[f'A{row}'] = '日期：'
        ws[f'E{row}'] = '日期：'

        # 保存文件
        if not output_path:
            output_dir = os.path.join(settings.MEDIA_ROOT, 'quotes', 'generated')
            os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(
                output_dir,
                f"{quote_data.get('quote_no', 'quote')}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
            )

        wb.save(output_path)
        return output_path

    def generate_html(self, quote_data: dict) -> str:
        """生成HTML报价单"""
        from django.template import Context, Template

        if self.template.html_content:
            tpl = Template(self.template.html_content)
        else:
            # 默认HTML模板
            tpl = Template(self._get_default_html_template())

        # 计算合计
        items = quote_data.get('items', [])
        total_amount = sum(Decimal(str(item.get('amount', 0))) for item in items)

        context = Context({
            'quote': quote_data,
            'header': self.template.header_config or {},
            'footer': self.template.footer_config or {},
            'items': items,
            'total_amount': total_amount,
            'generated_at': datetime.now(),
        })

        return tpl.render(context)

    def _get_default_html_template(self) -> str:
        return """
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>报价单 - {{ quote.quote_no }}</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { font-family: "Microsoft YaHei", sans-serif; padding: 40px; background: #f5f5f5; }
        .quote-container { max-width: 900px; margin: 0 auto; background: white; padding: 40px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
        .header { text-align: center; margin-bottom: 30px; padding-bottom: 20px; border-bottom: 2px solid #333; }
        .company-name { font-size: 24px; font-weight: bold; color: #333; }
        .quote-title { font-size: 20px; margin-top: 15px; color: #666; }
        .info-section { display: flex; flex-wrap: wrap; margin-bottom: 25px; }
        .info-item { width: 50%; padding: 8px 0; }
        .info-label { font-weight: bold; color: #666; }
        .items-table { width: 100%; border-collapse: collapse; margin-bottom: 25px; }
        .items-table th, .items-table td { border: 1px solid #ddd; padding: 12px; text-align: left; }
        .items-table th { background: #f0f0f0; font-weight: bold; }
        .items-table .num { text-align: right; }
        .total-row { font-weight: bold; background: #f9f9f9; }
        .footer-section { margin-top: 30px; padding-top: 20px; border-top: 1px solid #ddd; }
        .remarks { margin-bottom: 20px; }
        .terms { margin-bottom: 20px; }
        .terms li { margin: 5px 0; color: #666; }
        .signature-area { display: flex; justify-content: space-between; margin-top: 40px; }
        .signature-box { width: 45%; }
        .signature-line { border-top: 1px solid #333; margin-top: 60px; padding-top: 10px; }
        @media print { body { padding: 0; background: white; } .quote-container { box-shadow: none; } }
    </style>
</head>
<body>
    <div class="quote-container">
        <div class="header">
            <div class="company-name">{{ header.company_name|default:"公司名称" }}</div>
            <div class="quote-title">报 价 单</div>
        </div>
        
        <div class="info-section">
            <div class="info-item"><span class="info-label">报价单号：</span>{{ quote.quote_no }}</div>
            <div class="info-item"><span class="info-label">报价日期：</span>{{ quote.quote_date }}</div>
            <div class="info-item"><span class="info-label">客户名称：</span>{{ quote.customer_name }}</div>
            <div class="info-item"><span class="info-label">有效期至：</span>{{ quote.valid_until }}</div>
            <div class="info-item"><span class="info-label">项目名称：</span>{{ quote.project_name }}</div>
            <div class="info-item"><span class="info-label">联 系 人：</span>{{ quote.contact_person }}</div>
        </div>
        
        <table class="items-table">
            <thead>
                <tr>
                    <th style="width:50px;">序号</th>
                    <th>名称</th>
                    <th>规格型号</th>
                    <th style="width:60px;">单位</th>
                    <th style="width:80px;" class="num">数量</th>
                    <th style="width:100px;" class="num">单价</th>
                    <th style="width:120px;" class="num">金额</th>
                    <th>备注</th>
                </tr>
            </thead>
            <tbody>
                {% for item in items %}
                <tr>
                    <td>{{ forloop.counter }}</td>
                    <td>{{ item.name }}</td>
                    <td>{{ item.spec }}</td>
                    <td>{{ item.unit }}</td>
                    <td class="num">{{ item.quantity }}</td>
                    <td class="num">{{ item.unit_price }}</td>
                    <td class="num">{{ item.amount }}</td>
                    <td>{{ item.remark }}</td>
                </tr>
                {% endfor %}
                <tr class="total-row">
                    <td colspan="6" style="text-align:right;">合计：</td>
                    <td class="num">¥ {{ total_amount }}</td>
                    <td></td>
                </tr>
            </tbody>
        </table>
        
        <div class="footer-section">
            {% if quote.remarks %}
            <div class="remarks">
                <strong>备注：</strong>{{ quote.remarks }}
            </div>
            {% endif %}
            
            {% if footer.terms %}
            <div class="terms">
                <strong>条款与条件：</strong>
                <ul>
                    {% for term in footer.terms %}
                    <li>{{ term }}</li>
                    {% endfor %}
                </ul>
            </div>
            {% endif %}
        </div>
        
        <div class="signature-area">
            <div class="signature-box">
                <div>供方（盖章）：</div>
                <div class="signature-line">日期：</div>
            </div>
            <div class="signature-box">
                <div>需方（盖章）：</div>
                <div class="signature-line">日期：</div>
            </div>
        </div>
    </div>
</body>
</html>
"""


# =====================
# Serializers
# =====================

class QuoteTemplateSerializer(serializers.ModelSerializer):
    format_display = serializers.CharField(source='get_format_display', read_only=True)

    class Meta:
        model = QuoteTemplate
        fields = '__all__'
        read_only_fields = ['created_by', 'updated_by']


class QuoteTemplateListSerializer(serializers.ModelSerializer):
    format_display = serializers.CharField(source='get_format_display', read_only=True)

    class Meta:
        model = QuoteTemplate
        fields = [
            'id', 'code', 'name', 'format', 'format_display',
            'is_default', 'is_enabled', 'created_at'
        ]


class QuoteHistorySerializer(serializers.ModelSerializer):
    template_name = serializers.CharField(source='template.name', read_only=True)
    status_display = serializers.CharField(source='get_status_display', read_only=True)

    class Meta:
        model = QuoteHistory
        fields = '__all__'
        read_only_fields = ['created_by', 'updated_by']


class QuoteGenerateSerializer(serializers.Serializer):
    """报价单生成请求"""
    template_id = serializers.IntegerField(required=False, help_text='模板ID，不传则使用默认模板')
    quote_no = serializers.CharField(max_length=50)
    customer_name = serializers.CharField(max_length=200)
    project_name = serializers.CharField(max_length=200, required=False, allow_blank=True)
    quote_date = serializers.DateField(required=False)
    valid_until = serializers.DateField(required=False)
    contact_person = serializers.CharField(max_length=100, required=False, allow_blank=True)
    contact_phone = serializers.CharField(max_length=50, required=False, allow_blank=True)
    remarks = serializers.CharField(required=False, allow_blank=True)
    items = serializers.ListField(
        child=serializers.DictField(),
        help_text='报价明细列表'
    )
    output_format = serializers.ChoiceField(
        choices=['excel', 'html', 'pdf'],
        default='excel'
    )


# =====================
# ViewSets
# =====================

class QuoteTemplateViewSet(SoftDeleteMixin, UserTrackingMixin, viewsets.ModelViewSet):
    """
    报价单模板管理
    """
    queryset = QuoteTemplate.objects.filter(is_deleted=False)
    permission_classes = [IsAuthenticated]
    filterset_fields = ['format', 'is_default', 'is_enabled']
    search_fields = ['code', 'name', 'description']
    ordering_fields = ['code', 'name', 'created_at']

    def get_serializer_class(self):
        if self.action == 'list':
            return QuoteTemplateListSerializer
        return QuoteTemplateSerializer

    @action(detail=True, methods=['post'])
    def set_default(self, request, pk=None):
        """设为默认模板"""
        instance = self.get_object()
        instance.is_default = True
        instance.save()
        return Response(self.get_serializer(instance).data)

    @action(detail=True, methods=['post'])
    def toggle_enable(self, request, pk=None):
        """切换启用状态"""
        instance = self.get_object()
        instance.is_enabled = not instance.is_enabled
        instance.save()
        return Response(self.get_serializer(instance).data)

    @action(detail=False, methods=['post'])
    def generate(self, request):
        """生成报价单"""
        serializer = QuoteGenerateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        # 获取模板
        template_id = data.get('template_id')
        if template_id:
            try:
                template = QuoteTemplate.objects.get(id=template_id, is_enabled=True, is_deleted=False)
            except QuoteTemplate.DoesNotExist:
                return Response({'error': '模板不存在'}, status=404)
        else:
            template = QuoteTemplate.objects.filter(
                is_default=True, is_enabled=True, is_deleted=False
            ).first()
            if not template:
                # 创建默认模板
                template = QuoteTemplate.objects.create(
                    code='DEFAULT',
                    name='默认报价单模板',
                    format='EXCEL',
                    is_default=True,
                    column_config=[
                        {'key': 'index', 'label': '序号', 'width': 8},
                        {'key': 'name', 'label': '名称', 'width': 30},
                        {'key': 'spec', 'label': '规格型号', 'width': 25},
                        {'key': 'unit', 'label': '单位', 'width': 10},
                        {'key': 'quantity', 'label': '数量', 'width': 12},
                        {'key': 'unit_price', 'label': '单价', 'width': 15},
                        {'key': 'amount', 'label': '金额', 'width': 18},
                        {'key': 'remark', 'label': '备注', 'width': 20},
                    ],
                    header_config={
                        'company_name': '非标自动化设备有限公司',
                    },
                    footer_config={
                        'terms': [
                            '以上报价有效期30天',
                            '付款方式：预付30%，发货前付60%，验收后付10%',
                            '交货期：合同签订后45个工作日',
                        ]
                    },
                    created_by=request.user
                )

        # 准备报价数据
        quote_data = {
            'quote_no': data['quote_no'],
            'customer_name': data['customer_name'],
            'project_name': data.get('project_name', ''),
            'quote_date': data.get('quote_date', datetime.now().date()).strftime('%Y-%m-%d') if data.get('quote_date') else datetime.now().strftime('%Y-%m-%d'),
            'valid_until': data.get('valid_until').strftime('%Y-%m-%d') if data.get('valid_until') else '',
            'contact_person': data.get('contact_person', ''),
            'contact_phone': data.get('contact_phone', ''),
            'remarks': data.get('remarks', ''),
            'items': data.get('items', []),
        }

        # 生成报价单
        service = QuoteGenerationService(template)
        output_format = data.get('output_format', 'excel')

        try:
            if output_format == 'excel':
                file_path = service.generate_excel(quote_data)
                # 记录历史
                total_amount = sum(Decimal(str(item.get('amount', 0))) for item in quote_data['items'])
                history = QuoteHistory.objects.create(
                    template=template,
                    quote_no=quote_data['quote_no'],
                    customer_name=quote_data['customer_name'],
                    project_name=quote_data.get('project_name', ''),
                    total_amount=total_amount,
                    valid_until=data.get('valid_until'),
                    quote_data=quote_data,
                    generated_file=file_path.replace(settings.MEDIA_ROOT, '').lstrip('/'),
                    created_by=request.user
                )
                return Response({
                    'success': True,
                    'history_id': history.id,
                    'download_url': f'/media/{history.generated_file}'
                })

            elif output_format == 'html':
                html_content = service.generate_html(quote_data)
                return Response({
                    'success': True,
                    'html_content': html_content
                })

            else:
                return Response({'error': '不支持的输出格式'}, status=400)

        except Exception as e:
            return Response({'error': str(e)}, status=500)

    @action(detail=True, methods=['post'])
    def preview(self, request, pk=None):
        """预览模板效果"""
        template = self.get_object()

        # 使用示例数据预览
        sample_data = {
            'quote_no': 'QT-2025-SAMPLE',
            'customer_name': '示例客户有限公司',
            'project_name': '自动化装配线项目',
            'quote_date': datetime.now().strftime('%Y-%m-%d'),
            'valid_until': '2025-02-28',
            'contact_person': '张先生',
            'contact_phone': '13800138000',
            'remarks': '以上价格均含税（增值税13%）',
            'items': [
                {'name': '装配工位', 'spec': 'ZP-001', 'unit': '套', 'quantity': 2, 'unit_price': 50000, 'amount': 100000, 'remark': '含工装夹具'},
                {'name': '输送线体', 'spec': 'SS-200', 'unit': '米', 'quantity': 20, 'unit_price': 3000, 'amount': 60000, 'remark': '倍速链'},
                {'name': '控制系统', 'spec': 'PLC-S7', 'unit': '套', 'quantity': 1, 'unit_price': 35000, 'amount': 35000, 'remark': '西门子'},
            ]
        }

        service = QuoteGenerationService(template)

        if template.format == 'HTML':
            html_content = service.generate_html(sample_data)
            return Response({'html_content': html_content})
        else:
            # 其他格式返回基本配置信息
            return Response({
                'template': QuoteTemplateSerializer(template).data,
                'sample_data': sample_data
            })


class QuoteHistoryViewSet(viewsets.ReadOnlyModelViewSet):
    """
    报价单历史（只读）
    """
    queryset = QuoteHistory.objects.all()
    serializer_class = QuoteHistorySerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['status', 'template']
    search_fields = ['quote_no', 'customer_name', 'project_name']
    ordering_fields = ['created_at', 'total_amount']

    @action(detail=True, methods=['post'])
    def update_status(self, request, pk=None):
        """更新状态"""
        history = self.get_object()
        new_status = request.data.get('status')

        if new_status not in dict(QuoteHistory._meta.get_field('status').choices):
            return Response({'error': '无效的状态'}, status=400)

        history.status = new_status
        history.save()
        return Response(self.get_serializer(history).data)

    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """报价统计"""
        from django.db.models import Count, Sum
        from django.db.models.functions import TruncMonth

        qs = self.get_queryset()

        # 按状态统计
        by_status = qs.values('status').annotate(
            count=Count('id'),
            total=Sum('total_amount')
        )

        # 按月统计
        by_month = qs.annotate(
            month=TruncMonth('created_at')
        ).values('month').annotate(
            count=Count('id'),
            total=Sum('total_amount')
        ).order_by('month')

        # 转化率
        total = qs.count()
        accepted = qs.filter(status='ACCEPTED').count()

        return Response({
            'total_count': total,
            'total_amount': qs.aggregate(Sum('total_amount'))['total_amount__sum'] or 0,
            'accepted_count': accepted,
            'conversion_rate': round(accepted / total * 100, 2) if total > 0 else 0,
            'by_status': list(by_status),
            'by_month': list(by_month)
        })
