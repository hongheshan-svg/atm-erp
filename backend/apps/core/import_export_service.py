"""
数据导入导出服务
Data Import/Export Service
"""

import csv
import io
from datetime import datetime

from django.db import transaction
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class ExportService:
    """数据导出服务"""

    @staticmethod
    def export_to_excel(data, columns, filename='export', sheet_name='Sheet1'):
        """
        导出数据到Excel

        :param data: 数据列表 [dict, ...]
        :param columns: 列定义 [{'field': 'name', 'title': '名称', 'width': 20}, ...]
        :param filename: 文件名
        :param sheet_name: 工作表名称
        :return: HttpResponse
        """
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # 样式定义
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='409EFF', end_color='409EFF', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')
        )

        # 写入表头
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col['title'])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

            # 设置列宽
            ws.column_dimensions[get_column_letter(col_idx)].width = col.get('width', 15)

        # 写入数据
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, col in enumerate(columns, 1):
                field = col['field']
                value = ExportService._get_nested_value(row_data, field)

                # 格式化日期
                if isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')

                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')

        # 冻结首行
        ws.freeze_panes = 'A2'

        # 生成响应
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = (
            f'attachment; filename="{filename}_{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx"'
        )

        wb.save(response)
        return response

    @staticmethod
    def export_to_csv(data, columns, filename='export'):
        """
        导出数据到CSV
        """
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = (
            f'attachment; filename="{filename}_{datetime.now().strftime("%Y%m%d%H%M%S")}.csv"'
        )

        writer = csv.writer(response)

        # 写入表头
        writer.writerow([col['title'] for col in columns])

        # 写入数据
        for row_data in data:
            row = []
            for col in columns:
                value = ExportService._get_nested_value(row_data, col['field'])
                if isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                row.append(value)
            writer.writerow(row)

        return response

    @staticmethod
    def _get_nested_value(data, field):
        """获取嵌套字段值"""
        if isinstance(data, dict):
            keys = field.split('.')
            value = data
            for key in keys:
                if isinstance(value, dict):
                    value = value.get(key, '')
                else:
                    return ''
            return value
        return getattr(data, field, '')


class ImportService:
    """数据导入服务"""

    @staticmethod
    def parse_excel(file, column_mapping, start_row=2):
        """
        解析Excel文件

        :param file: 上传的文件对象
        :param column_mapping: 列映射 {'A': 'name', 'B': 'code', ...}
        :param start_row: 数据起始行（默认跳过表头）
        :return: 解析后的数据列表
        """
        try:
            wb = load_workbook(file, data_only=True)
            ws = wb.active

            data = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=start_row), start_row):
                row_data = {}
                for col_letter, field in column_mapping.items():
                    col_idx = ord(col_letter.upper()) - ord('A') + 1
                    if col_idx <= len(row):
                        value = row[col_idx - 1].value
                        row_data[field] = value

                # 跳过空行
                if any(row_data.values()):
                    row_data['_row'] = row_idx  # 记录行号用于错误提示
                    data.append(row_data)

            return {'success': True, 'data': data}
        except Exception as e:
            return {'success': False, 'error': str(e)}

    @staticmethod
    def parse_csv(file, column_mapping):
        """
        解析CSV文件

        :param file: 上传的文件对象
        :param column_mapping: 列映射 {'列名1': 'field1', '列名2': 'field2', ...}
        :return: 解析后的数据列表
        """
        try:
            # 读取内容
            content = file.read()
            # 尝试不同编码
            for encoding in ['utf-8-sig', 'utf-8', 'gbk', 'gb2312']:
                try:
                    text = content.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            else:
                return {'success': False, 'error': '无法识别文件编码'}

            reader = csv.DictReader(io.StringIO(text))

            data = []
            for row_idx, row in enumerate(reader, 2):
                row_data = {}
                for csv_col, field in column_mapping.items():
                    row_data[field] = row.get(csv_col, '')

                if any(row_data.values()):
                    row_data['_row'] = row_idx
                    data.append(row_data)

            return {'success': True, 'data': data}
        except Exception as e:
            return {'success': False, 'error': str(e)}

    @staticmethod
    def validate_data(data, validators):
        """
        验证导入数据

        :param data: 数据列表
        :param validators: 验证器字典 {'field': [validator_func, ...], ...}
        :return: {'valid': [...], 'errors': [{'row': 1, 'field': 'name', 'message': '...'}]}
        """
        valid_data = []
        errors = []

        for row_data in data:
            row_errors = []
            for field, field_validators in validators.items():
                value = row_data.get(field)
                for validator in field_validators:
                    result = validator(value, row_data)
                    if result is not True:
                        row_errors.append({'row': row_data.get('_row', 0), 'field': field, 'message': result})
                        break

            if row_errors:
                errors.extend(row_errors)
            else:
                valid_data.append(row_data)

        return {'valid': valid_data, 'errors': errors}

    @staticmethod
    def import_data(model_class, data, field_mapping, user=None, batch_size=100):
        """
        批量导入数据

        :param model_class: Django模型类
        :param data: 数据列表
        :param field_mapping: 字段映射 {'import_field': 'model_field', ...}
        :param user: 当前用户（用于设置created_by等）
        :param batch_size: 批量创建大小
        :return: {'created': count, 'errors': [...]}
        """
        created_count = 0
        errors = []

        try:
            with transaction.atomic():
                objects_to_create = []

                for row_data in data:
                    try:
                        obj_data = {}
                        for import_field, model_field in field_mapping.items():
                            if import_field in row_data:
                                obj_data[model_field] = row_data[import_field]

                        if user:
                            obj_data['created_by'] = user

                        obj = model_class(**obj_data)
                        objects_to_create.append(obj)

                        if len(objects_to_create) >= batch_size:
                            model_class.objects.bulk_create(objects_to_create)
                            created_count += len(objects_to_create)
                            objects_to_create = []

                    except Exception as e:
                        errors.append({'row': row_data.get('_row', 0), 'error': str(e)})

                # 创建剩余对象
                if objects_to_create:
                    model_class.objects.bulk_create(objects_to_create)
                    created_count += len(objects_to_create)

        except Exception as e:
            return {'created': 0, 'errors': [{'error': str(e)}]}

        return {'created': created_count, 'errors': errors}


# 常用验证器
def required(value, row_data):
    """必填验证"""
    if value is None or value == '':
        return '此字段必填'
    return True


def max_length(max_len):
    """最大长度验证"""

    def validator(value, row_data):
        if value and len(str(value)) > max_len:
            return f'长度不能超过{max_len}个字符'
        return True

    return validator


def is_number(value, row_data):
    """数字验证"""
    if value is None or value == '':
        return True
    try:
        float(value)
        return True
    except (ValueError, TypeError):
        return '必须是数字'


def is_positive(value, row_data):
    """正数验证"""
    if value is None or value == '':
        return True
    try:
        if float(value) < 0:
            return '必须是正数'
        return True
    except (ValueError, TypeError):
        return '必须是数字'


def unique_in_db(model_class, field):
    """数据库唯一性验证"""

    def validator(value, row_data):
        if value and model_class.objects.filter(**{field: value}).exists():
            return f'"{value}" 已存在'
        return True

    return validator
