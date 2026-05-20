"""
文件预览服务
File Preview Service
支持PDF、Office文档、图片等常见格式的在线预览
"""

import hashlib
import mimetypes
import os
from pathlib import Path

from django.conf import settings
from django.http import FileResponse
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView


class FilePreviewService:
    """文件预览服务"""

    # 支持直接预览的MIME类型
    DIRECT_PREVIEW_TYPES = {
        # 图片
        'image/jpeg',
        'image/png',
        'image/gif',
        'image/webp',
        'image/svg+xml',
        # PDF
        'application/pdf',
        # 文本
        'text/plain',
        'text/html',
        'text/css',
        'text/javascript',
        'text/csv',
        'text/markdown',
        # 代码
        'application/json',
        'application/xml',
    }

    # 需要转换的Office文档类型
    OFFICE_TYPES = {
        'application/msword',
        'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        'application/vnd.ms-excel',
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'application/vnd.ms-powerpoint',
        'application/vnd.openxmlformats-officedocument.presentationml.presentation',
    }

    # 文件扩展名映射
    EXTENSION_MAP = {
        '.doc': 'application/msword',
        '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        '.xls': 'application/vnd.ms-excel',
        '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        '.ppt': 'application/vnd.ms-powerpoint',
        '.pptx': 'application/vnd.openxmlformats-officedocument.presentationml.presentation',
        '.pdf': 'application/pdf',
        '.txt': 'text/plain',
        '.csv': 'text/csv',
        '.md': 'text/markdown',
        '.json': 'application/json',
        '.xml': 'application/xml',
    }

    def __init__(self):
        self.preview_cache_dir = Path(settings.MEDIA_ROOT) / 'preview_cache'
        self.preview_cache_dir.mkdir(parents=True, exist_ok=True)

    def get_mime_type(self, file_path: str) -> str:
        """获取文件MIME类型"""
        ext = os.path.splitext(file_path)[1].lower()
        if ext in self.EXTENSION_MAP:
            return self.EXTENSION_MAP[ext]
        mime_type, _ = mimetypes.guess_type(file_path)
        return mime_type or 'application/octet-stream'

    def get_preview_info(self, file_path: str) -> dict:
        """获取文件预览信息"""
        if not os.path.exists(file_path):
            return {'error': '文件不存在', 'previewable': False}

        file_stat = os.stat(file_path)
        mime_type = self.get_mime_type(file_path)
        file_name = os.path.basename(file_path)
        file_ext = os.path.splitext(file_name)[1].lower()

        # 判断是否可预览
        can_preview = False
        preview_type = 'download'  # download, image, pdf, office, text, video, audio

        if mime_type in self.DIRECT_PREVIEW_TYPES:
            can_preview = True
            if mime_type.startswith('image/'):
                preview_type = 'image'
            elif mime_type == 'application/pdf':
                preview_type = 'pdf'
            elif mime_type.startswith('text/') or mime_type in ('application/json', 'application/xml'):
                preview_type = 'text'
        elif mime_type in self.OFFICE_TYPES:
            can_preview = True
            preview_type = 'office'
        elif mime_type and mime_type.startswith('video/'):
            can_preview = True
            preview_type = 'video'
        elif mime_type and mime_type.startswith('audio/'):
            can_preview = True
            preview_type = 'audio'

        return {
            'file_name': file_name,
            'file_size': file_stat.st_size,
            'file_size_display': self._format_size(file_stat.st_size),
            'mime_type': mime_type,
            'extension': file_ext,
            'previewable': can_preview,
            'preview_type': preview_type,
            'modified_time': file_stat.st_mtime,
        }

    def _format_size(self, size: int) -> str:
        """格式化文件大小"""
        for unit in ['B', 'KB', 'MB', 'GB']:
            if size < 1024:
                return f'{size:.1f} {unit}'
            size /= 1024
        return f'{size:.1f} TB'

    def get_text_content(self, file_path: str, max_size: int = 1024 * 1024) -> dict:
        """获取文本文件内容"""
        file_stat = os.stat(file_path)

        if file_stat.st_size > max_size:
            return {
                'content': '',
                'truncated': True,
                'message': f'文件过大（{self._format_size(file_stat.st_size)}），无法预览',
            }

        try:
            # 尝试不同编码读取
            for encoding in ['utf-8', 'gbk', 'gb2312', 'latin-1']:
                try:
                    with open(file_path, 'r', encoding=encoding) as f:
                        content = f.read()
                    return {'content': content, 'encoding': encoding, 'truncated': False}
                except UnicodeDecodeError:
                    continue

            return {'content': '', 'error': '无法解析文件编码'}
        except Exception as e:
            return {'content': '', 'error': str(e)}

    def get_excel_preview(self, file_path: str, sheet_name: str = None, max_rows: int = 100) -> dict:
        """获取Excel文件预览"""
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter

            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheets = wb.sheetnames

            if sheet_name and sheet_name in sheets:
                ws = wb[sheet_name]
            else:
                ws = wb.active
                sheet_name = ws.title

            # 读取数据
            data = []
            for row_idx, row in enumerate(ws.iter_rows(max_row=max_rows + 1)):
                if row_idx >= max_rows:
                    break
                row_data = []
                for cell in row:
                    value = cell.value
                    if value is None:
                        value = ''
                    elif hasattr(value, 'strftime'):
                        value = value.strftime('%Y-%m-%d %H:%M:%S')
                    else:
                        value = str(value)
                    row_data.append(value)
                data.append(row_data)

            # 获取列数
            max_col = max(len(row) for row in data) if data else 0
            headers = [get_column_letter(i + 1) for i in range(max_col)]

            wb.close()

            return {
                'sheets': sheets,
                'current_sheet': sheet_name,
                'headers': headers,
                'data': data,
                'total_rows': ws.max_row,
                'truncated': ws.max_row > max_rows if ws.max_row else False,
            }
        except ImportError:
            return {'error': 'openpyxl 库未安装'}
        except Exception as e:
            return {'error': str(e)}

    def get_image_thumbnail(self, file_path: str, max_size: tuple = (800, 800)) -> str:
        """获取图片缩略图路径"""
        try:
            from PIL import Image

            # 生成缓存文件名
            file_hash = hashlib.md5(f'{file_path}_{max_size}'.encode()).hexdigest()
            cache_path = self.preview_cache_dir / f'{file_hash}.jpg'

            # 检查缓存
            if cache_path.exists():
                file_mtime = os.path.getmtime(file_path)
                cache_mtime = os.path.getmtime(cache_path)
                if cache_mtime > file_mtime:
                    return str(cache_path)

            # 生成缩略图
            with Image.open(file_path) as img:
                # 转换为RGB（处理PNG透明背景）
                if img.mode in ('RGBA', 'LA', 'P'):
                    background = Image.new('RGB', img.size, (255, 255, 255))
                    if img.mode == 'P':
                        img = img.convert('RGBA')
                    background.paste(img, mask=img.split()[-1] if img.mode == 'RGBA' else None)
                    img = background
                elif img.mode != 'RGB':
                    img = img.convert('RGB')

                img.thumbnail(max_size, Image.Resampling.LANCZOS)
                img.save(cache_path, 'JPEG', quality=85, optimize=True)

            return str(cache_path)
        except ImportError:
            return file_path  # 没有PIL则返回原图路径
        except Exception:
            return file_path


class FilePreviewView(APIView):
    """文件预览接口"""

    permission_classes = [IsAuthenticated]

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.service = FilePreviewService()

    def get(self, request):
        """
        获取文件预览
        参数:
        - path: 文件路径（相对于MEDIA_ROOT或绝对路径）
        - action: info(获取信息), content(获取内容), download(下载)
        - sheet: Excel工作表名称
        """
        file_path = request.query_params.get('path', '')
        action = request.query_params.get('action', 'info')

        if not file_path:
            return Response({'error': '请提供文件路径'}, status=400)

        # 处理相对路径
        if not os.path.isabs(file_path):
            file_path = os.path.join(settings.MEDIA_ROOT, file_path)

        # 安全检查：确保文件在允许的目录内
        real_path = os.path.realpath(file_path)
        allowed_dirs = [
            os.path.realpath(settings.MEDIA_ROOT),
        ]
        if hasattr(settings, 'FILE_PREVIEW_ALLOWED_DIRS'):
            allowed_dirs.extend(settings.FILE_PREVIEW_ALLOWED_DIRS)

        is_allowed = any(real_path.startswith(d) for d in allowed_dirs)
        if not is_allowed:
            return Response({'error': '无权访问该文件'}, status=403)

        if not os.path.exists(file_path):
            return Response({'error': '文件不存在'}, status=404)

        if action == 'info':
            return Response(self.service.get_preview_info(file_path))

        elif action == 'content':
            info = self.service.get_preview_info(file_path)

            if info['preview_type'] == 'text':
                content = self.service.get_text_content(file_path)
                return Response({**info, **content})

            elif info['preview_type'] == 'office':
                if info['extension'] in ('.xls', '.xlsx'):
                    sheet_name = request.query_params.get('sheet')
                    excel_data = self.service.get_excel_preview(file_path, sheet_name)
                    return Response({**info, **excel_data})
                else:
                    # 其他Office文档返回基本信息
                    return Response({**info, 'message': '请下载后查看完整内容'})

            elif info['preview_type'] == 'image':
                # 返回缩略图URL
                thumb_path = self.service.get_image_thumbnail(file_path)
                return Response({**info, 'thumbnail_path': thumb_path.replace(settings.MEDIA_ROOT, '/media')})

            else:
                return Response(info)

        elif action == 'download':
            mime_type = self.service.get_mime_type(file_path)
            file_name = os.path.basename(file_path)

            response = FileResponse(open(file_path, 'rb'), content_type=mime_type)
            response['Content-Disposition'] = f'attachment; filename="{file_name}"'
            return response

        elif action == 'view':
            # 直接预览（用于图片、PDF等）
            mime_type = self.service.get_mime_type(file_path)
            file_name = os.path.basename(file_path)

            response = FileResponse(open(file_path, 'rb'), content_type=mime_type)
            response['Content-Disposition'] = f'inline; filename="{file_name}"'
            return response

        else:
            return Response({'error': '不支持的操作'}, status=400)


class FileUploadView(APIView):
    """文件上传接口"""

    permission_classes = [IsAuthenticated]

    def post(self, request):
        """
        上传文件
        参数:
        - file: 文件
        - category: 分类（可选）
        - description: 描述（可选）
        """
        file = request.FILES.get('file')
        if not file:
            return Response({'error': '请选择要上传的文件'}, status=400)

        category = request.data.get('category', 'general')

        # 创建存储目录
        from datetime import datetime

        date_path = datetime.now().strftime('%Y/%m/%d')
        upload_dir = Path(settings.MEDIA_ROOT) / 'uploads' / category / date_path
        upload_dir.mkdir(parents=True, exist_ok=True)

        # 生成唯一文件名
        file_ext = os.path.splitext(file.name)[1]
        file_hash = hashlib.md5(f'{file.name}_{datetime.now().isoformat()}'.encode()).hexdigest()[:12]
        safe_name = f'{file_hash}{file_ext}'

        file_path = upload_dir / safe_name

        # 保存文件
        with open(file_path, 'wb+') as destination:
            for chunk in file.chunks():
                destination.write(chunk)

        # 获取文件信息
        service = FilePreviewService()
        info = service.get_preview_info(str(file_path))

        return Response(
            {
                'success': True,
                'file_path': str(file_path).replace(settings.MEDIA_ROOT, '').lstrip('/'),
                'original_name': file.name,
                'saved_name': safe_name,
                'url': f'/media/uploads/{category}/{date_path}/{safe_name}',
                **info,
            }
        )


class FileListView(APIView):
    """文件列表接口"""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        """
        获取文件列表
        参数:
        - directory: 目录路径（相对于MEDIA_ROOT）
        - recursive: 是否递归
        """
        directory = request.query_params.get('directory', 'uploads')
        recursive = request.query_params.get('recursive', 'false').lower() == 'true'

        base_path = Path(settings.MEDIA_ROOT) / directory

        if not base_path.exists():
            return Response({'error': '目录不存在'}, status=404)

        if not base_path.is_dir():
            return Response({'error': '不是有效的目录'}, status=400)

        service = FilePreviewService()
        files = []
        directories = []

        if recursive:
            for item in base_path.rglob('*'):
                if item.is_file():
                    rel_path = item.relative_to(settings.MEDIA_ROOT)
                    info = service.get_preview_info(str(item))
                    files.append({'path': str(rel_path), 'name': item.name, **info})
        else:
            for item in base_path.iterdir():
                rel_path = item.relative_to(settings.MEDIA_ROOT)
                if item.is_file():
                    info = service.get_preview_info(str(item))
                    files.append({'path': str(rel_path), 'name': item.name, **info})
                elif item.is_dir():
                    # 统计目录下文件数量
                    file_count = sum(1 for _ in item.rglob('*') if _.is_file())
                    directories.append({'path': str(rel_path), 'name': item.name, 'file_count': file_count})

        return Response(
            {
                'directory': directory,
                'directories': directories,
                'files': files,
                'total_files': len(files),
                'total_directories': len(directories),
            }
        )
