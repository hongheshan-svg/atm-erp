"""
数据导入导出视图
"""

from django.http import HttpResponse
from rest_framework import status
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView

from .import_export_service import ExportService, ImportService, max_length, required
from .permissions import module_menu_permission

# 主数据(客户/供应商/物料)导入导出含 PII 与商业机密;BOM 导出含保密项目物料清单。
# 原先裸 IsAuthenticated 放行任意登录用户(审计 batch1 授权旁路),改为模块菜单门控。
MasterdataMenuAccess = module_menu_permission('masterdata')
ProjectsMenuAccess = module_menu_permission('projects')


class ItemImportView(APIView):
    """物料导入"""

    parser_classes = [MultiPartParser]
    permission_classes = [MasterdataMenuAccess]

    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': '请上传文件'}, status=status.HTTP_400_BAD_REQUEST)

        # 解析Excel
        column_mapping = {
            'A': 'code',
            'B': 'name',
            'C': 'spec',
            'D': 'unit',
            'E': 'category',
            'F': 'brand',
            'G': 'model',
        }

        result = ImportService.parse_excel(file, column_mapping)
        if not result['success']:
            return Response({'error': result['error']}, status=status.HTTP_400_BAD_REQUEST)

        # 验证数据
        from apps.masterdata.models import Item

        validators = {
            'code': [required, max_length(50)],
            'name': [required, max_length(200)],
        }

        validation = ImportService.validate_data(result['data'], validators)

        if validation['errors']:
            return Response(
                {
                    'success': False,
                    'valid_count': len(validation['valid']),
                    'error_count': len(validation['errors']),
                    'errors': validation['errors'][:50],  # 只返回前50个错误
                }
            )

        # 导入数据
        # 左键是解析后的列名，右键必须是 Item 的真实字段名：
        # Item 的唯一编码字段是 sku（没有 code 字段），原先映射成 'code' 会让
        # Item(**obj_data) 每行抛 TypeError 被逐行吞掉 → created 恒为 0。
        field_mapping = {
            'code': 'sku',
            'name': 'name',
            'spec': 'specification',
            'unit': 'unit',
            'brand': 'brand',
            'model': 'model',
        }

        import_result = ImportService.import_data(Item, validation['valid'], field_mapping, request.user)

        # 一行都没导进去时不能再报 success，否则前端显示导入成功但库里没有数据
        succeeded = import_result['created'] > 0 or not import_result['errors']
        return Response(
            {
                'success': succeeded,
                'created': import_result['created'],
                'errors': import_result['errors'],
            }
        )

    def get(self, request):
        """下载导入模板"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = '物料导入模板'

        headers = ['物料编码*', '物料名称*', '规格型号', '单位', '分类', '品牌', '型号']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='E0E0E0', fill_type='solid')

        # 示例数据
        example = ['M001', '螺栓M8x20', 'M8x20 不锈钢', '个', '紧固件', '国标', '304']
        for col, value in enumerate(example, 1):
            ws.cell(row=2, column=col, value=value)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="item_import_template.xlsx"'
        wb.save(response)
        return response


class ItemExportView(APIView):
    """物料导出"""

    permission_classes = [MasterdataMenuAccess]

    def get(self, request):
        from apps.masterdata.models import Item
        from apps.masterdata.serializers import ItemSerializer

        queryset = Item.objects.filter(is_deleted=False)

        # 应用筛选
        category = request.query_params.get('category')
        if category:
            queryset = queryset.filter(category_id=category)

        search = request.query_params.get('search')
        if search:
            queryset = queryset.filter(name__icontains=search) | queryset.filter(code__icontains=search)

        serializer = ItemSerializer(queryset[:5000], many=True)

        columns = [
            {'field': 'code', 'title': '物料编码', 'width': 15},
            {'field': 'name', 'title': '物料名称', 'width': 30},
            {'field': 'specification', 'title': '规格型号', 'width': 20},
            {'field': 'unit', 'title': '单位', 'width': 8},
            {'field': 'category_detail.name', 'title': '分类', 'width': 15},
            {'field': 'brand', 'title': '品牌', 'width': 15},
            {'field': 'model', 'title': '型号', 'width': 15},
            {'field': 'reference_price', 'title': '参考价', 'width': 12},
        ]

        export_format = request.query_params.get('format', 'excel')
        if export_format == 'csv':
            return ExportService.export_to_csv(serializer.data, columns, 'items')
        return ExportService.export_to_excel(serializer.data, columns, 'items', '物料列表')


class SupplierImportView(APIView):
    """供应商导入"""

    parser_classes = [MultiPartParser]
    permission_classes = [MasterdataMenuAccess]

    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': '请上传文件'}, status=status.HTTP_400_BAD_REQUEST)

        column_mapping = {
            'A': 'code',
            'B': 'name',
            'C': 'contact_person',
            'D': 'contact_phone',
            'E': 'email',
            'F': 'address',
        }

        result = ImportService.parse_excel(file, column_mapping)
        if not result['success']:
            return Response({'error': result['error']}, status=status.HTTP_400_BAD_REQUEST)

        from apps.masterdata.models import Supplier

        validators = {
            'code': [required, max_length(50)],
            'name': [required, max_length(200)],
        }

        validation = ImportService.validate_data(result['data'], validators)

        if validation['errors']:
            return Response(
                {
                    'success': False,
                    'valid_count': len(validation['valid']),
                    'error_count': len(validation['errors']),
                    'errors': validation['errors'][:50],
                }
            )

        field_mapping = {
            'code': 'code',
            'name': 'name',
            'contact_person': 'contact_person',
            'contact_phone': 'phone',
            'email': 'email',
            'address': 'address',
        }

        import_result = ImportService.import_data(Supplier, validation['valid'], field_mapping, request.user)

        return Response({'success': True, 'created': import_result['created'], 'errors': import_result['errors']})

    def get(self, request):
        """下载导入模板"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = '供应商导入模板'

        headers = ['供应商编码*', '供应商名称*', '联系人', '联系电话', '邮箱', '地址']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='E0E0E0', fill_type='solid')

        example = [
            'S001',
            '深圳XX精密机械有限公司',
            '张三',
            '13800138000',
            'zhangsan@example.com',
            '深圳市宝安区XX路XX号',
        ]
        for col, value in enumerate(example, 1):
            ws.cell(row=2, column=col, value=value)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="supplier_import_template.xlsx"'
        wb.save(response)
        return response


class SupplierExportView(APIView):
    """供应商导出"""

    permission_classes = [MasterdataMenuAccess]

    def get(self, request):
        from apps.masterdata.models import Supplier
        from apps.masterdata.serializers import SupplierSerializer

        queryset = Supplier.objects.filter(is_deleted=False)
        serializer = SupplierSerializer(queryset[:5000], many=True)

        columns = [
            {'field': 'code', 'title': '供应商编码', 'width': 15},
            {'field': 'name', 'title': '供应商名称', 'width': 30},
            {'field': 'contact_person', 'title': '联系人', 'width': 12},
            {'field': 'phone', 'title': '联系电话', 'width': 15},
            {'field': 'email', 'title': '邮箱', 'width': 25},
            {'field': 'address', 'title': '地址', 'width': 40},
        ]

        export_format = request.query_params.get('format', 'excel')
        if export_format == 'csv':
            return ExportService.export_to_csv(serializer.data, columns, 'suppliers')
        return ExportService.export_to_excel(serializer.data, columns, 'suppliers', '供应商列表')


class CustomerExportView(APIView):
    """客户导出"""

    permission_classes = [MasterdataMenuAccess]

    def get(self, request):
        from apps.masterdata.models import Customer
        from apps.masterdata.serializers import CustomerSerializer

        queryset = Customer.objects.filter(is_deleted=False)
        serializer = CustomerSerializer(queryset[:5000], many=True)

        columns = [
            {'field': 'code', 'title': '客户编码', 'width': 15},
            {'field': 'name', 'title': '客户名称', 'width': 30},
            {'field': 'contact_person', 'title': '联系人', 'width': 12},
            {'field': 'phone', 'title': '联系电话', 'width': 15},
            {'field': 'email', 'title': '邮箱', 'width': 25},
            {'field': 'address', 'title': '地址', 'width': 40},
        ]

        return ExportService.export_to_excel(serializer.data, columns, 'customers', '客户列表')


class BOMExportView(APIView):
    """BOM清单导出"""

    permission_classes = [ProjectsMenuAccess]

    def get(self, request):
        from apps.projects.models import ProjectBOM

        project_id = request.query_params.get('project')
        if not project_id:
            return Response({'error': '请指定项目'}, status=status.HTTP_400_BAD_REQUEST)

        queryset = ProjectBOM.objects.filter(project_id=project_id, is_deleted=False).select_related('item', 'project')

        data = []
        for bom in queryset:
            data.append(
                {
                    'project_name': bom.project.name if bom.project else '',
                    'item_code': bom.item.code if bom.item else '',
                    'item_name': bom.item.name if bom.item else '',
                    'item_spec': bom.item.specification if bom.item else '',
                    'quantity': bom.quantity,
                    'unit': bom.unit,
                    'remark': bom.remark,
                }
            )

        columns = [
            {'field': 'project_name', 'title': '项目名称', 'width': 25},
            {'field': 'item_code', 'title': '物料编码', 'width': 15},
            {'field': 'item_name', 'title': '物料名称', 'width': 25},
            {'field': 'item_spec', 'title': '规格型号', 'width': 20},
            {'field': 'quantity', 'title': '数量', 'width': 10},
            {'field': 'unit', 'title': '单位', 'width': 8},
            {'field': 'remark', 'title': '备注', 'width': 30},
        ]

        return ExportService.export_to_excel(data, columns, 'bom', 'BOM清单')
