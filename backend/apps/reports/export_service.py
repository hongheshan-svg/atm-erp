"""
报表导出增强服务
Report Export Enhancement Service
支持多种格式导出、批量导出、定时导出等
"""
import io
import os
import json
from datetime import datetime, date
from decimal import Decimal
from django.db import models
from django.http import HttpResponse
from django.conf import settings
from rest_framework import viewsets, serializers, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated

from apps.core.models import BaseModel
from apps.core.mixins import SoftDeleteMixin, UserTrackingMixin


class ExportTemplate(BaseModel):
    """
    导出模板配置
    """
    EXPORT_FORMATS = [
        ('EXCEL', 'Excel (.xlsx)'),
        ('CSV', 'CSV (.csv)'),
        ('PDF', 'PDF (.pdf)'),
        ('JSON', 'JSON (.json)'),
    ]
    
    code = models.CharField(max_length=50, unique=True, verbose_name='模板编码')
    name = models.CharField(max_length=200, verbose_name='模板名称')
    model_name = models.CharField(max_length=100, verbose_name='模型名称')
    
    # 导出字段配置
    fields_config = models.JSONField(
        default=list,
        verbose_name='字段配置',
        help_text='[{"field": "code", "label": "编码", "width": 15}, ...]'
    )
    
    # 筛选条件
    default_filters = models.JSONField(
        default=dict,
        blank=True,
        verbose_name='默认筛选条件'
    )
    
    # 排序
    default_ordering = models.JSONField(
        default=list,
        blank=True,
        verbose_name='默认排序'
    )
    
    # 格式
    default_format = models.CharField(
        max_length=20,
        choices=EXPORT_FORMATS,
        default='EXCEL',
        verbose_name='默认格式'
    )
    
    # 样式
    header_style = models.JSONField(
        default=dict,
        blank=True,
        verbose_name='表头样式'
    )
    cell_style = models.JSONField(
        default=dict,
        blank=True,
        verbose_name='单元格样式'
    )
    
    is_active = models.BooleanField(default=True, verbose_name='启用')
    description = models.TextField(blank=True, verbose_name='描述')
    
    class Meta:
        db_table = 'reports_export_template'
        verbose_name = '导出模板'
        verbose_name_plural = verbose_name
        ordering = ['model_name', 'code']
    
    def __str__(self):
        return f'{self.code} - {self.name}'


class ExportHistory(BaseModel):
    """
    导出历史记录
    """
    STATUS_CHOICES = [
        ('PENDING', '待处理'),
        ('PROCESSING', '处理中'),
        ('COMPLETED', '已完成'),
        ('FAILED', '失败'),
    ]
    
    template = models.ForeignKey(
        ExportTemplate,
        on_delete=models.SET_NULL,
        null=True,
        blank=True,
        related_name='export_histories',
        verbose_name='导出模板'
    )
    
    export_name = models.CharField(max_length=200, verbose_name='导出名称')
    export_format = models.CharField(max_length=20, verbose_name='导出格式')
    
    # 筛选条件
    filters_used = models.JSONField(default=dict, blank=True, verbose_name='使用的筛选条件')
    
    # 结果
    record_count = models.IntegerField(default=0, verbose_name='导出记录数')
    file_path = models.CharField(max_length=500, blank=True, verbose_name='文件路径')
    file_size = models.IntegerField(default=0, verbose_name='文件大小(字节)')
    
    status = models.CharField(
        max_length=20,
        choices=STATUS_CHOICES,
        default='PENDING',
        verbose_name='状态'
    )
    error_message = models.TextField(blank=True, verbose_name='错误信息')
    
    started_at = models.DateTimeField(null=True, blank=True, verbose_name='开始时间')
    completed_at = models.DateTimeField(null=True, blank=True, verbose_name='完成时间')
    
    class Meta:
        db_table = 'reports_export_history'
        verbose_name = '导出历史'
        verbose_name_plural = verbose_name
        ordering = ['-created_at']
    
    def __str__(self):
        return f'{self.export_name} - {self.created_at}'


class ScheduledExport(BaseModel):
    """
    定时导出任务
    """
    FREQUENCY_CHOICES = [
        ('DAILY', '每天'),
        ('WEEKLY', '每周'),
        ('MONTHLY', '每月'),
    ]
    
    template = models.ForeignKey(
        ExportTemplate,
        on_delete=models.CASCADE,
        related_name='scheduled_exports',
        verbose_name='导出模板'
    )
    
    name = models.CharField(max_length=200, verbose_name='任务名称')
    frequency = models.CharField(
        max_length=20,
        choices=FREQUENCY_CHOICES,
        default='DAILY',
        verbose_name='频率'
    )
    
    # 执行时间
    execute_time = models.TimeField(verbose_name='执行时间')
    execute_day = models.IntegerField(
        null=True,
        blank=True,
        verbose_name='执行日',
        help_text='周几(1-7)或几号(1-31)'
    )
    
    # 筛选条件
    filters = models.JSONField(default=dict, blank=True, verbose_name='筛选条件')
    
    # 收件人
    recipients = models.JSONField(
        default=list,
        verbose_name='收件人邮箱'
    )
    
    is_active = models.BooleanField(default=True, verbose_name='启用')
    last_run_at = models.DateTimeField(null=True, blank=True, verbose_name='上次运行时间')
    next_run_at = models.DateTimeField(null=True, blank=True, verbose_name='下次运行时间')
    
    class Meta:
        db_table = 'reports_scheduled_export'
        verbose_name = '定时导出'
        verbose_name_plural = verbose_name
        ordering = ['-created_at']
    
    def __str__(self):
        return self.name


# =====================
# Export Service
# =====================

class ExportService:
    """导出服务"""
    
    @staticmethod
    def export_to_excel(queryset, fields_config, filename='export'):
        """导出到Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        except ImportError:
            # 如果没有openpyxl，使用xlsxwriter
            import xlsxwriter
            return ExportService._export_excel_xlsxwriter(queryset, fields_config, filename)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Data'
        
        # 样式定义
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 写入表头
        for col, field in enumerate(fields_config, 1):
            cell = ws.cell(row=1, column=col, value=field.get('label', field.get('field')))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = field.get('width', 15)
        
        # 写入数据
        for row_idx, obj in enumerate(queryset, 2):
            for col_idx, field in enumerate(fields_config, 1):
                field_name = field.get('field')
                value = ExportService._get_field_value(obj, field_name)
                
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output, f'{filename}.xlsx', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    @staticmethod
    def _export_excel_xlsxwriter(queryset, fields_config, filename):
        """使用xlsxwriter导出Excel"""
        import xlsxwriter
        
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet('Data')
        
        # 格式
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#4472C4',
            'font_color': 'white',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1
        })
        
        cell_format = workbook.add_format({
            'border': 1,
            'valign': 'vcenter'
        })
        
        # 写入表头
        for col, field in enumerate(fields_config):
            worksheet.write(0, col, field.get('label', field.get('field')), header_format)
            worksheet.set_column(col, col, field.get('width', 15))
        
        # 写入数据
        for row_idx, obj in enumerate(queryset, 1):
            for col_idx, field in enumerate(fields_config):
                field_name = field.get('field')
                value = ExportService._get_field_value(obj, field_name)
                worksheet.write(row_idx, col_idx, value, cell_format)
        
        workbook.close()
        output.seek(0)
        
        return output, f'{filename}.xlsx', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    @staticmethod
    def export_to_csv(queryset, fields_config, filename='export'):
        """导出到CSV"""
        import csv
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # 写入表头
        headers = [field.get('label', field.get('field')) for field in fields_config]
        writer.writerow(headers)
        
        # 写入数据
        for obj in queryset:
            row = []
            for field in fields_config:
                field_name = field.get('field')
                value = ExportService._get_field_value(obj, field_name)
                row.append(value)
            writer.writerow(row)
        
        output.seek(0)
        
        # 转换为字节
        byte_output = io.BytesIO()
        byte_output.write('\ufeff'.encode('utf-8'))  # BOM for Excel
        byte_output.write(output.getvalue().encode('utf-8'))
        byte_output.seek(0)
        
        return byte_output, f'{filename}.csv', 'text/csv; charset=utf-8'
    
    @staticmethod
    def export_to_json(queryset, fields_config, filename='export'):
        """导出到JSON"""
        data = []
        
        for obj in queryset:
            row = {}
            for field in fields_config:
                field_name = field.get('field')
                value = ExportService._get_field_value(obj, field_name)
                if isinstance(value, (datetime, date)):
                    value = value.isoformat()
                elif isinstance(value, Decimal):
                    value = float(value)
                row[field_name] = value
            data.append(row)
        
        output = io.BytesIO()
        output.write(json.dumps(data, ensure_ascii=False, indent=2).encode('utf-8'))
        output.seek(0)
        
        return output, f'{filename}.json', 'application/json'
    
    @staticmethod
    def _get_field_value(obj, field_path):
        """获取字段值，支持嵌套字段如 'customer.name'"""
        if hasattr(obj, 'get'):
            # 字典类型
            value = obj.get(field_path)
        else:
            # 模型对象
            parts = field_path.split('__')
            value = obj
            for part in parts:
                if value is None:
                    break
                if hasattr(value, part):
                    value = getattr(value, part)
                    if callable(value):
                        value = value()
                elif hasattr(value, 'get'):
                    value = value.get(part)
                else:
                    value = None
        
        # 格式化
        if isinstance(value, datetime):
            value = value.strftime('%Y-%m-%d %H:%M:%S')
        elif isinstance(value, date):
            value = value.strftime('%Y-%m-%d')
        elif isinstance(value, Decimal):
            value = float(value)
        elif value is None:
            value = ''
        
        return value


# =====================
# Serializers
# =====================

class ExportTemplateSerializer(serializers.ModelSerializer):
    class Meta:
        model = ExportTemplate
        fields = '__all__'
        read_only_fields = ['created_by', 'updated_by']


class ExportHistorySerializer(serializers.ModelSerializer):
    template_name = serializers.CharField(source='template.name', read_only=True)
    status_display = serializers.CharField(source='get_status_display', read_only=True)
    
    class Meta:
        model = ExportHistory
        fields = '__all__'
        read_only_fields = ['created_by', 'updated_by']


class ScheduledExportSerializer(serializers.ModelSerializer):
    template_name = serializers.CharField(source='template.name', read_only=True)
    frequency_display = serializers.CharField(source='get_frequency_display', read_only=True)
    
    class Meta:
        model = ScheduledExport
        fields = '__all__'
        read_only_fields = ['created_by', 'updated_by', 'last_run_at', 'next_run_at']


# =====================
# ViewSets
# =====================

class ExportTemplateViewSet(SoftDeleteMixin, UserTrackingMixin, viewsets.ModelViewSet):
    """导出模板管理"""
    queryset = ExportTemplate.objects.filter(is_deleted=False)
    serializer_class = ExportTemplateSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['model_name', 'default_format', 'is_active']
    search_fields = ['code', 'name', 'description']
    
    @action(detail=False, methods=['post'])
    def init_templates(self, request):
        """初始化常用导出模板"""
        templates = [
            {
                'code': 'PROJECT_LIST',
                'name': '项目列表',
                'model_name': 'projects.Project',
                'fields_config': [
                    {'field': 'code', 'label': '项目编号', 'width': 15},
                    {'field': 'name', 'label': '项目名称', 'width': 30},
                    {'field': 'customer__name', 'label': '客户', 'width': 20},
                    {'field': 'status', 'label': '状态', 'width': 10},
                    {'field': 'start_date', 'label': '开始日期', 'width': 12},
                    {'field': 'end_date', 'label': '结束日期', 'width': 12},
                    {'field': 'progress', 'label': '进度(%)', 'width': 10},
                ]
            },
            {
                'code': 'SALES_ORDER_LIST',
                'name': '销售订单列表',
                'model_name': 'sales.SalesOrder',
                'fields_config': [
                    {'field': 'order_no', 'label': '订单号', 'width': 18},
                    {'field': 'customer__name', 'label': '客户', 'width': 25},
                    {'field': 'order_date', 'label': '订单日期', 'width': 12},
                    {'field': 'total_amount', 'label': '金额', 'width': 15},
                    {'field': 'status', 'label': '状态', 'width': 10},
                    {'field': 'salesperson__first_name', 'label': '销售员', 'width': 12},
                ]
            },
            {
                'code': 'PURCHASE_ORDER_LIST',
                'name': '采购订单列表',
                'model_name': 'purchase.PurchaseOrder',
                'fields_config': [
                    {'field': 'order_no', 'label': '订单号', 'width': 18},
                    {'field': 'supplier__name', 'label': '供应商', 'width': 25},
                    {'field': 'order_date', 'label': '订单日期', 'width': 12},
                    {'field': 'total_amount', 'label': '金额', 'width': 15},
                    {'field': 'status', 'label': '状态', 'width': 10},
                ]
            },
            {
                'code': 'INVENTORY_LIST',
                'name': '库存清单',
                'model_name': 'inventory.Stock',
                'fields_config': [
                    {'field': 'item__code', 'label': '物料编码', 'width': 15},
                    {'field': 'item__name', 'label': '物料名称', 'width': 30},
                    {'field': 'warehouse__name', 'label': '仓库', 'width': 15},
                    {'field': 'quantity', 'label': '库存数量', 'width': 12},
                    {'field': 'item__unit', 'label': '单位', 'width': 8},
                ]
            },
        ]
        
        created = 0
        for tpl in templates:
            _, c = ExportTemplate.objects.get_or_create(
                code=tpl['code'],
                defaults={
                    **tpl,
                    'created_by': request.user
                }
            )
            if c:
                created += 1
        
        return Response({'success': True, 'created': created})
    
    @action(detail=True, methods=['post'])
    def export(self, request, pk=None):
        """执行导出"""
        from django.apps import apps
        
        template = self.get_object()
        export_format = request.data.get('format', template.default_format)
        filters = request.data.get('filters', template.default_filters)
        
        # 获取模型
        try:
            app_label, model_name = template.model_name.split('.')
            model = apps.get_model(app_label, model_name)
        except Exception as e:
            return Response({'error': f'模型不存在: {e}'}, status=400)
        
        # 构建查询
        queryset = model.objects.all()
        if hasattr(model, 'is_deleted'):
            queryset = queryset.filter(is_deleted=False)
        
        if filters:
            queryset = queryset.filter(**filters)
        
        if template.default_ordering:
            queryset = queryset.order_by(*template.default_ordering)
        
        # 执行导出
        filename = f'{template.name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        
        if export_format == 'EXCEL':
            output, name, content_type = ExportService.export_to_excel(
                queryset, template.fields_config, filename
            )
        elif export_format == 'CSV':
            output, name, content_type = ExportService.export_to_csv(
                queryset, template.fields_config, filename
            )
        elif export_format == 'JSON':
            output, name, content_type = ExportService.export_to_json(
                queryset, template.fields_config, filename
            )
        else:
            return Response({'error': '不支持的导出格式'}, status=400)
        
        # 记录导出历史
        ExportHistory.objects.create(
            template=template,
            export_name=filename,
            export_format=export_format,
            filters_used=filters,
            record_count=queryset.count(),
            status='COMPLETED',
            completed_at=datetime.now(),
            created_by=request.user
        )
        
        response = HttpResponse(output.read(), content_type=content_type)
        response['Content-Disposition'] = f'attachment; filename="{name}"'
        return response


class ExportHistoryViewSet(viewsets.ReadOnlyModelViewSet):
    """导出历史"""
    queryset = ExportHistory.objects.filter(is_deleted=False)
    serializer_class = ExportHistorySerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['template', 'status', 'export_format']
    ordering_fields = ['created_at', 'record_count']
    
    def get_queryset(self):
        if self.request.user.is_superuser:
            return self.queryset
        return self.queryset.filter(created_by=self.request.user)


class ScheduledExportViewSet(SoftDeleteMixin, UserTrackingMixin, viewsets.ModelViewSet):
    """定时导出管理"""
    queryset = ScheduledExport.objects.filter(is_deleted=False)
    serializer_class = ScheduledExportSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['template', 'frequency', 'is_active']
    
    @action(detail=True, methods=['post'])
    def toggle_active(self, request, pk=None):
        """切换启用状态"""
        task = self.get_object()
        task.is_active = not task.is_active
        task.save()
        return Response(self.get_serializer(task).data)
    
    @action(detail=True, methods=['post'])
    def run_now(self, request, pk=None):
        """立即执行"""
        task = self.get_object()
        # 这里应该触发 Celery 任务
        return Response({'success': True, 'message': '任务已加入队列'})
