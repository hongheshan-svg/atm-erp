"""
Views for masterdata app.
"""
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.http import HttpResponse
import pandas as pd
from io import BytesIO
from apps.core.mixins import SoftDeleteMixin, UserTrackingMixin
from apps.core.permission_mixin import PermissionMixin
from .models import ItemCategory, Item, Customer, Supplier, Warehouse, WarehouseLocation
from .serializers import (
    ItemCategorySerializer, ItemSerializer, CustomerSerializer,
    SupplierSerializer, WarehouseSerializer, WarehouseLocationSerializer,
    WarehouseLocationTreeSerializer
)
from .item_code_generator import ItemCodeGenerator


class ItemCategoryViewSet(PermissionMixin, SoftDeleteMixin, UserTrackingMixin, viewsets.ModelViewSet):
    permission_module = 'masterdata'
    permission_resource = 'item_category'
    """
    ViewSet for ItemCategory management.
    """
    queryset = ItemCategory.objects.all()
    serializer_class = ItemCategorySerializer
    filterset_fields = ['parent', 'is_deleted']
    search_fields = ['code', 'name']
    ordering_fields = ['sort_order', 'code', 'created_at']
    
    @action(detail=False, methods=['get'])
    def tree(self, request):
        """Get category tree structure."""
        categories = self.get_queryset().filter(is_deleted=False)
        
        def build_tree(parent_id=None):
            result = []
            items = categories.filter(parent_id=parent_id)
            for item in items:
                node = ItemCategorySerializer(item).data
                node['children'] = build_tree(item.id)
                result.append(node)
            return result
        
        tree_data = build_tree()
        return Response(tree_data)


class ItemViewSet(PermissionMixin, SoftDeleteMixin, UserTrackingMixin, viewsets.ModelViewSet):
    """
    ViewSet for Item management.
    NOTE: 物料是主数据，所有用户都可以查看，不应用数据范围限制
    NOTE: 价格字段通过字段权限控制可见性
    """
    queryset = Item.objects.all()
    serializer_class = ItemSerializer
    filterset_fields = ['category', 'item_type', 'is_active', 'is_deleted']
    search_fields = ['sku', 'name', 'specification', 'barcode']
    ordering_fields = ['sku', 'created_at', 'standard_cost']

    # Permission configuration
    permission_module = 'masterdata'
    permission_resource = 'item'
    allow_authenticated_read = True
    
    def perform_destroy(self, instance):
        """
        重写删除方法，使用软删除。
        物料可能被项目BOM、采购单、入库单等引用，不能直接硬删除。
        """
        from django.utils import timezone
        instance.is_deleted = True
        instance.deleted_at = timezone.now()
        instance.save(update_fields=['is_deleted', 'deleted_at'])
    
    @action(detail=False, methods=['post'])
    def generate_code(self, request):
        """
        生成物料编码
        
        参数：
        - level1_code: 一级代码 ('1'=有图, '2'=无图)
        - level2_code: 二级代码 ('1'-'8')
        """
        level1_code = request.data.get('level1_code')
        level2_code = request.data.get('level2_code')
        
        if not level1_code or not level2_code:
            return Response(
                {'error': '请提供一级代码和二级代码'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            code = ItemCodeGenerator.generate_code(level1_code, level2_code)
            code_info = ItemCodeGenerator.parse_code(code)
            
            return Response({
                'code': code,
                'info': code_info
            })
        except ValueError as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=False, methods=['get'])
    def get_level2_choices(self, request):
        """获取二级代码选项"""
        return Response(ItemCodeGenerator.get_level2_choices())
    
    @action(detail=False, methods=['post'])
    def import_excel(self, request):
        """
        Import items from Excel file with all fields.
        
        智能导入逻辑（查缺补漏）：
        1. 如果文件中的物料编码与系统中已存在的相同，则跳过该物料（不导入也不更新）
        2. 如果没有物料编码，先根据"规格型号+版本/品牌"在数据库中查找是否已存在相同物料
           - 如果找到匹配的，使用已有的编码（复用已有物料，跳过创建）
           - 如果没找到，根据"有图/无图"和"物料类型"自动生成新编码
        
        编码规则：有图/无图(1位) + 物料类型(1位) + 年份(2位) + 三级流水号(6位)
        - 有图/无图: 1=有图, 2=无图
        - 物料类型: 1=机加, 2=钣金, 3=特殊工艺, 4=其他, 5=机械类, 6=电气类, 7=耗材辅料, 8=办公用品
        - 年份: 有图=当前年份后两位, 无图=99
        - 三级流水号: 按年份循环累加(000001-999999)
        """
        file = request.FILES.get('file')
        if not file:
            return Response(
                {'error': '请上传文件'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            df = pd.read_excel(file)
            
            # Helper function to find column by keywords
            def find_column(df, keywords):
                for col in df.columns:
                    for keyword in keywords:
                        if keyword in str(col):
                            return col
                return None
            
            # Find required columns - 物料名称是必须的
            sku_col = find_column(df, ['物料编码', 'SKU', 'sku', '编码'])
            name_col = find_column(df, ['物料名称', '名称', 'name'])
            
            if not name_col:
                return Response(
                    {'error': 'Excel文件必须包含"物料名称"列'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Find code generation columns (用于自动生成编码)
            # 有图/无图列 (一级代码)
            level1_col = find_column(df, ['有图/无图', '有图无图', '一级代码', 'level1'])
            # 物料分类列 (二级代码) - 用于编码生成，不是物料属性(item_type)
            level2_col = find_column(df, ['物料分类', '二级代码', 'level2'])
            
            # Find optional columns
            spec_col = find_column(df, ['规格型号', '规格', 'specification', '图号/型号', '图号'])
            brand_col = find_column(df, ['品牌', 'brand'])
            model_col = find_column(df, ['型号', 'model'])
            version_brand_col = find_column(df, ['版本/品牌'])
            unit_col = find_column(df, ['单位', 'unit'])
            item_type_col = find_column(df, ['物料属性', 'item_type'])  # 原材料/产成品/半成品/服务
            purchase_col = find_column(df, ['采购单价', '采购价', 'purchase'])
            sale_col = find_column(df, ['销售单价', '销售价', 'sale'])
            cost_col = find_column(df, ['标准成本', 'standard_cost'])
            tax_col = find_column(df, ['税率', 'tax'])
            mfr_col = find_column(df, ['生产厂家', '厂家', 'manufacturer'])
            origin_col = find_column(df, ['产地', 'origin'])
            safety_col = find_column(df, ['安全库存', 'safety'])
            lead_col = find_column(df, ['采购周期', 'lead_time'])
            min_col = find_column(df, ['最小库存', 'min_stock'])
            max_col = find_column(df, ['最大库存', 'max_stock'])
            barcode_col = find_column(df, ['条形码', 'barcode'])
            desc_col = find_column(df, ['描述', 'description'])
            status_col = find_column(df, ['状态', 'status', 'is_active'])
            
            created_count = 0
            matched_count = 0     # 匹配到已有物料的计数
            skip_exist_count = 0  # 系统中已存在的跳过计数
            skip_dup_count = 0    # 文件内重复的跳过计数
            error_rows = []
            
            # 有图/无图映射 (文字 -> 代码)
            level1_map = {
                '1': '1', '有图': '1', '有': '1',
                '2': '2', '无图': '2', '无': '2',
            }
            
            # 物料类型映射 (文字 -> 代码)
            level2_map = {
                '1': '1', '机加': '1',
                '2': '2', '钣金': '2',
                '3': '3', '特殊工艺': '3',
                '4': '4', '其他': '4',
                '5': '5', '机械类': '5',
                '6': '6', '电气类': '6',
                '7': '7', '耗材辅料': '7',
                '8': '8', '办公用品': '8',
            }
            
            # Track processed SKUs within this import
            processed_skus = set()
            # Track processed spec+brand combinations within this import
            processed_spec_brand = set()
            
            for idx, row in df.iterrows():
                row_num = idx + 2
                
                # 获取物料编码（可以为空）
                sku = ''
                if sku_col and pd.notna(row.get(sku_col)):
                    sku = str(row[sku_col]).strip()
                
                name = str(row[name_col]).strip() if pd.notna(row[name_col]) else ''
                
                # 物料名称是必须的
                if not name:
                    error_rows.append({'row': row_num, 'error': '物料名称为空'})
                    continue
                
                # Skip example rows
                if sku and (sku.startswith('MAT00') or '示例' in sku):
                    continue
                if '示例' in name:
                    continue
                
                # Get optional values helper functions
                def get_val(col, default=''):
                    if col and pd.notna(row.get(col)):
                        return str(row[col]).strip()
                    return default
                
                def get_num(col, default=0):
                    if col and pd.notna(row.get(col)):
                        try:
                            return float(row[col])
                        except (ValueError, TypeError):
                            return default
                    return default
                
                def get_int(col, default=0):
                    if col and pd.notna(row.get(col)):
                        try:
                            return int(float(row[col]))
                        except (ValueError, TypeError):
                            return default
                    return default
                
                try:
                    # Parse version/brand to brand/model first (需要用于匹配)
                    vb = get_val(version_brand_col)
                    parsed_brand = get_val(brand_col)
                    parsed_model = get_val(model_col)
                    if vb and not (parsed_brand or parsed_model):
                        if '/' in vb:
                            parts = [p.strip() for p in vb.split('/', 1)]
                            parsed_brand = parts[0]
                            parsed_model = parts[1] if len(parts) > 1 else ''
                        else:
                            parsed_brand = vb
                    
                    # 获取规格型号
                    specification = get_val(spec_col, '')
                    
                    # 组合版本/品牌用于匹配
                    version_brand = f"{parsed_brand}/{parsed_model}".strip('/') if parsed_brand or parsed_model else ''
                    
                    # 如果没有物料编码，需要智能匹配或自动生成
                    if not sku:
                        # 先尝试根据"规格型号+版本/品牌"在数据库中查找已有物料
                        existing_item = None
                        if specification or version_brand:
                            # 构建查询条件
                            query = Item.objects.filter(is_deleted=False)
                            if specification:
                                query = query.filter(specification=specification)
                            if version_brand:
                                # 匹配 brand/model 组合
                                if parsed_brand and parsed_model:
                                    query = query.filter(brand=parsed_brand, model=parsed_model)
                                elif parsed_brand:
                                    query = query.filter(brand=parsed_brand)
                            
                            # 如果有匹配条件，尝试查找
                            if specification or (parsed_brand or parsed_model):
                                existing_item = query.first()
                        
                        if existing_item:
                            # 找到匹配的已有物料，使用其编码
                            sku = existing_item.sku
                            matched_count += 1
                            # 跳过创建，因为物料已存在
                            skip_exist_count += 1
                            continue
                        else:
                            # 没有找到匹配的物料，需要生成新编码
                            # 获取有图/无图和物料类型
                            level1_raw = get_val(level1_col, '')
                            level2_raw = get_val(level2_col, '')
                            
                            if not level1_raw or not level2_raw:
                                error_rows.append({
                                    'row': row_num, 
                                    'error': '物料编码为空且无法匹配已有物料时，必须提供"有图/无图"和"物料类型"用于自动生成编码'
                                })
                                continue
                            
                            # 转换为标准代码
                            level1_code = level1_map.get(str(level1_raw).strip(), '')
                            level2_code = level2_map.get(str(level2_raw).strip(), '')
                            
                            if not level1_code:
                                error_rows.append({
                                    'row': row_num, 
                                    'error': f'无效的"有图/无图": {level1_raw}，应为"有图"或"无图"'
                                })
                                continue
                            
                            if not level2_code:
                                error_rows.append({
                                    'row': row_num, 
                                    'error': f'无效的"物料类型": {level2_raw}，应为: 机加/钣金/特殊工艺/其他/机械类/电气类/耗材辅料/办公用品'
                                })
                                continue
                            
                            # 使用编码生成器生成编码
                            sku = ItemCodeGenerator.generate_code(level1_code, level2_code)
                    
                    # 检查文件内重复（按编码）
                    if sku in processed_skus:
                        skip_dup_count += 1
                        continue
                    processed_skus.add(sku)
                    
                    # 检查文件内重复（按规格型号+版本/品牌）
                    spec_brand_key = f"{specification}|{version_brand}"
                    if spec_brand_key in processed_spec_brand and spec_brand_key != '|':
                        skip_dup_count += 1
                        continue
                    if spec_brand_key != '|':
                        processed_spec_brand.add(spec_brand_key)
                    
                    # 检查系统中是否已存在该物料编码 - 如果存在则跳过
                    if Item.objects.filter(sku=sku, is_deleted=False).exists():
                        skip_exist_count += 1
                        continue

                    # Parse status
                    status_val = get_val(status_col, '启用')
                    is_active = status_val not in ['禁用', '停用', 'false', 'False', '0', 'INACTIVE']

                    # Parse item_type (物料属性)
                    item_type_val = get_val(item_type_col, '原材料')
                    item_type_map = {
                        '原材料': 'MATERIAL', 'MATERIAL': 'MATERIAL',
                        '产成品': 'PRODUCT', 'PRODUCT': 'PRODUCT',
                        '半成品': 'SEMI', 'SEMI': 'SEMI',
                        '服务': 'SERVICE', 'SERVICE': 'SERVICE',
                    }
                    item_type = item_type_map.get(item_type_val, 'MATERIAL')
                    
                    # Parse unit (单位) - 支持中文和英文
                    unit_val = get_val(unit_col, '个')
                    unit_map = {
                        '个': 'PCS', 'PCS': 'PCS', 'pcs': 'PCS',
                        '千克': 'KG', 'KG': 'KG', 'kg': 'KG', '公斤': 'KG',
                        '米': 'M', 'M': 'M', 'm': 'M',
                        '平方米': 'M2', 'M2': 'M2', 'm2': 'M2', '㎡': 'M2',
                        '立方米': 'M3', 'M3': 'M3', 'm3': 'M3', '㎥': 'M3',
                        '套': 'SET', 'SET': 'SET', 'set': 'SET',
                        '箱': 'BOX', 'BOX': 'BOX', 'box': 'BOX',
                        '包': 'PACK', 'PACK': 'PACK', 'pack': 'PACK',
                        '小时': 'HOUR', 'HOUR': 'HOUR', 'hour': 'HOUR', 'H': 'HOUR', 'h': 'HOUR',
                    }
                    unit = unit_map.get(unit_val, 'PCS')

                    # 创建新物料（不更新已有物料）
                    item = Item.objects.create(
                        sku=sku,
                        name=name,
                        specification=specification,
                        brand=parsed_brand,
                        model=parsed_model,
                        unit=unit,
                        item_type=item_type,
                        purchase_price=get_num(purchase_col),
                        sale_price=get_num(sale_col),
                        standard_cost=get_num(cost_col),
                        tax_rate=get_int(tax_col, 13),
                        manufacturer=get_val(mfr_col),
                        origin_country=get_val(origin_col),
                        safety_stock=get_num(safety_col),
                        lead_time=get_int(lead_col),
                        min_stock=get_num(min_col),
                        max_stock=get_num(max_col),
                        is_active=is_active,
                        barcode=get_val(barcode_col),
                        description=get_val(desc_col),
                        created_by=request.user,
                        updated_by=request.user,
                    )
                    created_count += 1
                    
                except Exception as e:
                    error_rows.append({'row': row_num, 'error': str(e)})
            
            return Response({
                'message': f'导入完成：新增 {created_count} 条，匹配已有 {matched_count} 条，跳过已存在 {skip_exist_count} 条，跳过重复 {skip_dup_count} 条',
                'created': created_count,
                'matched_count': matched_count,
                'skip_exist_count': skip_exist_count,
                'skip_dup_count': skip_dup_count,
                'errors': error_rows
            })
        
        except Exception as e:
            return Response(
                {'error': f'导入失败: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Export items to Excel file with fields matching import template."""
        items = self.filter_queryset(self.get_queryset())
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            workbook = writer.book
            worksheet = workbook.add_worksheet('物料主数据')
            
            # Define formats - 与模板保持一致的颜色
            required_format = workbook.add_format({
                'bold': True,
                'bg_color': '#C00000',  # 红色 - 必填
                'font_color': 'white',
                'border': 1,
                'align': 'center',
                'valign': 'vcenter'
            })
            conditional_format = workbook.add_format({
                'bold': True,
                'bg_color': '#ED7D31',  # 橙色 - 条件必填
                'font_color': 'white',
                'border': 1,
                'align': 'center',
                'valign': 'vcenter'
            })
            optional_format = workbook.add_format({
                'bold': True,
                'bg_color': '#4472C4',  # 蓝色 - 选填
                'font_color': 'white',
                'border': 1,
                'align': 'center',
                'valign': 'vcenter'
            })
            data_format = workbook.add_format({
                'border': 1,
                'align': 'left',
                'valign': 'vcenter'
            })
            money_format = workbook.add_format({
                'border': 1,
                'align': 'right',
                'valign': 'vcenter',
                'num_format': '#,##0.00'
            })
            number_format = workbook.add_format({
                'border': 1,
                'align': 'right',
                'valign': 'vcenter',
                'num_format': '#,##0'
            })
            
            # Column headers matching import template exactly
            # (name, width, format_type)
            headers = [
                ('物料编码', 12, 'optional'),
                ('有图/无图', 10, 'conditional'),
                ('物料分类', 10, 'conditional'),
                ('物料名称*', 20, 'required'),
                ('规格型号', 15, 'optional'),
                ('版本/品牌', 12, 'optional'),
                ('单位', 8, 'optional'),
                ('物料属性', 10, 'optional'),
                ('采购单价', 10, 'optional'),
                ('销售单价', 10, 'optional'),
                ('标准成本', 10, 'optional'),
                ('税率(%)', 8, 'optional'),
                ('生产厂家', 15, 'optional'),
                ('产地', 10, 'optional'),
                ('安全库存', 10, 'optional'),
                ('采购周期(天)', 12, 'optional'),
                ('状态', 8, 'optional'),
            ]
            
            # Write headers with matching format
            format_map = {
                'required': required_format,
                'conditional': conditional_format,
                'optional': optional_format
            }
            for col, (header, width, format_type) in enumerate(headers):
                fmt = format_map[format_type]
                worksheet.write(0, col, header, fmt)
                worksheet.set_column(col, col, width)
            
            # Write data
            for row_idx, item in enumerate(items, 1):
                version_brand = f"{item.brand or ''}/{item.model or ''}" if item.brand or item.model else ''
                version_brand = version_brand.strip('/')
                
                # 解析物料编码获取有图/无图和物料分类
                code_info = ItemCodeGenerator.parse_code(item.sku)
                level1_display = ''
                level2_display = ''
                if code_info.get('valid'):
                    level1_display = code_info.get('level1_name', '')  # 有图/无图
                    level2_display = code_info.get('level2_name', '')  # 机加/钣金等（物料分类）
                
                # 状态显示
                status_display = '启用' if item.is_active else '禁用'
                
                worksheet.write(row_idx, 0, item.sku, data_format)
                worksheet.write(row_idx, 1, level1_display, data_format)
                worksheet.write(row_idx, 2, level2_display, data_format)
                worksheet.write(row_idx, 3, item.name, data_format)
                worksheet.write(row_idx, 4, item.specification or '', data_format)
                worksheet.write(row_idx, 5, version_brand, data_format)
                worksheet.write(row_idx, 6, item.get_unit_display(), data_format)
                worksheet.write(row_idx, 7, item.get_item_type_display(), data_format)
                worksheet.write(row_idx, 8, float(item.purchase_price), money_format)
                worksheet.write(row_idx, 9, float(item.sale_price), money_format)
                worksheet.write(row_idx, 10, float(item.standard_cost), money_format)
                worksheet.write(row_idx, 11, item.tax_rate, data_format)
                worksheet.write(row_idx, 12, item.manufacturer or '', data_format)
                worksheet.write(row_idx, 13, item.origin_country or '', data_format)
                worksheet.write(row_idx, 14, float(item.safety_stock), number_format)
                worksheet.write(row_idx, 15, item.lead_time, number_format)
                worksheet.write(row_idx, 16, status_display, data_format)
            
            worksheet.set_row(0, 25)  # 表头行高
            worksheet.freeze_panes(1, 0)
        
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=items.xlsx'
        return response
    
    @action(detail=False, methods=['get'])
    def export_template(self, request):
        """
        Export item import template with headers and example data.
        
        智能导入规则（查缺补漏）：
        1. 如果提供物料编码，且系统中已存在相同编码，则跳过该行（不导入）
        2. 如果物料编码为空，先根据"规格型号+版本/品牌"匹配已有物料
           - 如果匹配到，复用已有物料编码（跳过创建）
           - 如果没匹配到，根据"有图/无图"和"物料类型"自动生成编码
        """
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            workbook = writer.book
            worksheet = workbook.add_worksheet('物料导入模板')
            
            # Define formats
            required_format = workbook.add_format({
                'bold': True,
                'bg_color': '#C00000',
                'font_color': 'white',
                'border': 1,
                'align': 'center',
                'valign': 'vcenter',
                'text_wrap': True
            })
            conditional_format = workbook.add_format({
                'bold': True,
                'bg_color': '#ED7D31',  # 橙色表示条件必填
                'font_color': 'white',
                'border': 1,
                'align': 'center',
                'valign': 'vcenter',
                'text_wrap': True
            })
            optional_format = workbook.add_format({
                'bold': True,
                'bg_color': '#4472C4',
                'font_color': 'white',
                'border': 1,
                'align': 'center',
                'valign': 'vcenter',
                'text_wrap': True
            })
            example_format = workbook.add_format({
                'bg_color': '#FFF2CC',
                'border': 1,
                'italic': True,
                'font_color': '#666666'
            })
            
            # Column headers: (name, width, format_type) 
            # format_type: 'required', 'conditional', 'optional'
            headers = [
                ('物料编码', 12, 'optional'),           # 可选，为空时自动匹配或生成
                ('有图/无图', 10, 'conditional'),       # 物料编码为空且无法匹配时必填
                ('物料分类', 10, 'conditional'),        # 物料编码为空且无法匹配时必填（机加/钣金/机械类等）
                ('物料名称*', 20, 'required'),
                ('规格型号', 15, 'optional'),           # 用于智能匹配已有物料
                ('版本/品牌', 12, 'optional'),          # 用于智能匹配已有物料
                ('单位', 8, 'optional'),
                ('物料属性', 10, 'optional'),           # 原材料/产成品/半成品/服务
                ('采购单价', 10, 'optional'),
                ('销售单价', 10, 'optional'),
                ('标准成本', 10, 'optional'),
                ('税率(%)', 8, 'optional'),
                ('生产厂家', 15, 'optional'),
                ('产地', 10, 'optional'),
                ('安全库存', 10, 'optional'),
                ('采购周期(天)', 12, 'optional'),
                ('状态', 8, 'optional'),
            ]
            
            # Write headers
            format_map = {
                'required': required_format,
                'conditional': conditional_format,
                'optional': optional_format
            }
            for col, (header, width, format_type) in enumerate(headers):
                fmt = format_map[format_type]
                worksheet.write(0, col, header, fmt)
                worksheet.set_column(col, col, width)
            
            # Write example data - 示例1：有物料编码
            example_data1 = [
                '1126000001',     # 物料编码（有值时直接使用）
                '',               # 有图/无图（物料编码有值时可不填）
                '',               # 物料分类（物料编码有值时可不填）
                '示例物料A',
                'ABC-100-50',     # 规格型号/图号
                '品牌A/V1.0',
                'PCS',
                '原材料',         # 物料属性
                '100.00',
                '150.00',
                '100.00',
                '13',
                '示例厂家',
                '中国',
                '10',
                '7',
                '启用',
            ]
            for col, value in enumerate(example_data1):
                worksheet.write(1, col, value, example_format)
            
            # Write example data - 示例2：无物料编码，智能匹配或自动生成
            example_data2 = [
                '',               # 物料编码（为空，先匹配后生成）
                '有图',           # 有图/无图
                '机械类',         # 物料分类
                '示例物料B',
                'XYZ-200',        # 规格型号/图号（用于匹配已有物料）
                '品牌B',          # 版本/品牌（用于匹配已有物料）
                'SET',
                '原材料',         # 物料属性
                '200.00',
                '300.00',
                '180.00',
                '13',
                '',
                '',
                '5',
                '14',
                '启用',
            ]
            for col, value in enumerate(example_data2):
                worksheet.write(2, col, value, example_format)
            
            worksheet.set_row(0, 30)
            worksheet.freeze_panes(1, 0)
            
            # Help sheet
            help_sheet = workbook.add_worksheet('填写说明')
            title_format = workbook.add_format({'bold': True, 'font_size': 16, 'font_color': '#4472C4'})
            section_format = workbook.add_format({'bold': True, 'font_size': 12, 'font_color': '#C00000'})
            orange_text = workbook.add_format({'bold': True, 'font_color': '#ED7D31'})
            green_text = workbook.add_format({'bold': True, 'font_color': '#00B050'})
            
            help_content = [
                ('物料导入模板填写说明', title_format),
                ('', None),
                ('【智能导入规则 - 查缺补漏】', section_format),
                ('  1. 如果填写了物料编码，且系统中已存在相同编码，则跳过该行（不导入）', None),
                ('  2. 如果物料编码为空：', None),
                ('     a) 先根据"规格型号+版本/品牌"在数据库中查找匹配的已有物料', None),
                ('     b) 如果找到匹配的物料，则复用其编码（跳过创建，避免重复）', None),
                ('     c) 如果没找到匹配的物料，根据"有图/无图"和"物料分类"自动生成新编码', None),
                ('', None),
                ('【必填字段】红色标题', section_format),
                ('  • 物料名称*：物料的名称', None),
                ('', None),
                ('【条件必填字段】橙色标题', orange_text),
                ('  当物料编码为空且无法匹配已有物料时，以下字段必填：', None),
                ('  • 有图/无图：填"有图"或"无图"', None),
                ('  • 物料分类：机加/钣金/特殊工艺/其他/机械类/电气类/耗材辅料/办公用品（用于生成编码）', None),
                ('', None),
                ('【智能匹配字段】', green_text),
                ('  以下字段用于智能匹配已有物料（查缺补漏）：', None),
                ('  • 规格型号：物料的图号/型号，用于匹配数据库中已有物料', None),
                ('  • 版本/品牌：品牌和版本信息，用于匹配数据库中已有物料', None),
                ('', None),
                ('【编码生成规则】', section_format),
                ('  编码格式：有图/无图(1位) + 物料分类(1位) + 年份(2位) + 流水号(6位)', None),
                ('  • 有图/无图：有图=1，无图=2', None),
                ('  • 物料分类：1=机加，2=钣金，3=特殊工艺，4=其他，5=机械类，6=电气类，7=耗材辅料，8=办公用品', None),
                ('  • 年份："有图"使用当前年份后两位；"无图"固定为99', None),
                ('  • 流水号：同一年份内自动累加，范围000001-999999', None),
                ('  • 示例：1526000001 表示 有图+机械类+26年+第1个', None),
                ('', None),
                ('【选填字段】蓝色标题', section_format),
                ('  • 物料编码：可不填，系统自动匹配或生成', None),
                ('  • 物料属性：原材料/产成品/半成品/服务，默认原材料', None),
                ('  • 单位：个/套/千克/米/平方米/立方米/箱/包/小时（或PCS/SET/KG等英文），默认个', None),
                ('  • 采购单价/销售单价/标准成本：数字，默认0', None),
                ('  • 税率(%)：增值税率，默认13', None),
                ('  • 生产厂家/产地：文本', None),
                ('  • 安全库存/采购周期(天)：数字，默认0', None),
                ('  • 状态：启用/禁用，默认启用', None),
                ('', None),
                ('【注意事项】', section_format),
                ('  • 系统会自动根据规格型号和版本/品牌匹配已有物料，避免重复创建', None),
                ('  • 系统中已存在的物料编码不会被更新，会直接跳过', None),
                ('  • 导入前请删除示例行', None),
                ('  • 建议先少量导入测试，确认无误后再批量导入', None),
            ]
            
            help_sheet.set_column(0, 0, 80)
            for row, (text, fmt) in enumerate(help_content):
                if fmt:
                    help_sheet.write(row, 0, text, fmt)
                else:
                    help_sheet.write(row, 0, text)
        
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=item_import_template.xlsx'
        return response
    
    @action(detail=False, methods=['post'])
    def bulk_delete(self, request):
        """批量删除物料（软删除）"""
        from django.utils import timezone
        
        ids = request.data.get('ids', [])
        if not ids:
            return Response({'error': '请选择要删除的物料'}, status=status.HTTP_400_BAD_REQUEST)
        
        # 使用软删除，因为物料可能被其他业务数据引用
        deleted_count = Item.objects.filter(id__in=ids, is_deleted=False).update(
            is_deleted=True,
            deleted_at=timezone.now()
        )
        return Response({
            'message': f'成功删除 {deleted_count} 条物料',
            'deleted_count': deleted_count
        })
    
    @action(detail=True, methods=['get'])
    def generate_barcode(self, request, pk=None):
        """Generate barcode for item"""
        item = self.get_object()
        
        try:
            from apps.inventory.barcode_service import BarcodeService
            return BarcodeService.generate_item_barcode_response(item)
        except Exception as e:
            return Response(
                {'error': f'Barcode generation failed: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'])
    def generate_qrcode(self, request, pk=None):
        """Generate QR code for item"""
        item = self.get_object()
        
        try:
            from apps.inventory.barcode_service import BarcodeService
            return BarcodeService.generate_item_qrcode_response(item)
        except Exception as e:
            return Response(
                {'error': f'QR code generation failed: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class CustomerViewSet(PermissionMixin, SoftDeleteMixin, UserTrackingMixin, viewsets.ModelViewSet):
    permission_module = 'masterdata'
    permission_resource = 'customer'
    allow_authenticated_read = True
    """
    ViewSet for Customer management.
    NOTE: 客户是主数据，所有用户都可以查看，不应用数据范围限制
    """
    queryset = Customer.objects.all()
    serializer_class = CustomerSerializer
    filterset_fields = ['status', 'is_deleted']
    search_fields = ['code', 'name', 'contact_person', 'phone', 'tax_number']
    ordering_fields = ['code', 'created_at']
    
    def perform_destroy(self, instance):
        """客户直接使用软删除，避免触发关联查询导致的数据库结构问题"""
        from django.utils import timezone
        instance.is_deleted = True
        instance.deleted_at = timezone.now()
        instance.save(update_fields=['is_deleted', 'deleted_at'])
    
    def perform_create(self, serializer):
        """Auto-generate customer code if not provided, and create credit profile."""
        from apps.core.utils import generate_code
        from apps.masterdata.credit_management import CustomerCredit

        code = serializer.validated_data.get('code', '').strip()
        if not code:
            code = generate_code('C', rule_type='CUSTOMER')
        customer = serializer.save(code=code)

        # 自动为新客户建立信用档案，初始额度取客户的信用额度字段，保持两处口径一致
        CustomerCredit.objects.get_or_create(
            customer=customer,
            defaults={
                'credit_limit': customer.credit_limit or 0,
                'created_by': self.request.user if self.request.user.is_authenticated else None,
            },
        )

    @action(detail=False, methods=['get'])
    def download_template(self, request):
        """Download customer import template Excel file."""
        import io
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '客户导入模板'
        
        # 与列表页面显示的列对应
        headers = [
            '客户编码', '客户名称', '联系人', '电话', '税号', '开户银行', '地址', '状态'
        ]
        
        # Header style
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Example row - 与列表页面对应
        example = [
            'C001', '示例公司名称', '张三', '13800138000', 
            '91110000XXXXXXXX', '中国工商银行北京分行',
            '北京市朝阳区xxx路xxx号', '激活'
        ]
        for col, val in enumerate(example, 1):
            ws.cell(row=2, column=col, value=val)
        
        # Column widths - 与列表页面对应
        widths = [12, 25, 10, 15, 22, 20, 30, 8]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="customer_import_template.xlsx"'
        return response
    
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Export customers to Excel."""
        import io
        import openpyxl
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        from django.utils import timezone
        
        queryset = self.get_queryset().filter(is_deleted=False)
        
        status_filter = request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '客户列表'
        
        headers = [
            '客户编码', '客户名称', '简称', '联系人', '电话', '邮箱', '地址',
            '开票名称', '税号', '开户银行', '银行账号', '注册地址', '注册电话',
            '信用额度', '付款条款', '状态', '备注'
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        for row, customer in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=customer.code)
            ws.cell(row=row, column=2, value=customer.name)
            ws.cell(row=row, column=3, value=customer.short_name)
            ws.cell(row=row, column=4, value=customer.contact_person)
            ws.cell(row=row, column=5, value=customer.phone)
            ws.cell(row=row, column=6, value=customer.email)
            ws.cell(row=row, column=7, value=customer.address)
            ws.cell(row=row, column=8, value=customer.invoice_title)
            ws.cell(row=row, column=9, value=customer.tax_number)
            ws.cell(row=row, column=10, value=customer.bank_name)
            ws.cell(row=row, column=11, value=customer.bank_account)
            ws.cell(row=row, column=12, value=customer.registered_address)
            ws.cell(row=row, column=13, value=customer.registered_phone)
            ws.cell(row=row, column=14, value=float(customer.credit_limit))
            ws.cell(row=row, column=15, value=customer.payment_terms)
            ws.cell(row=row, column=16, value=customer.get_status_display())
            ws.cell(row=row, column=17, value=customer.notes)
        
        column_widths = [12, 30, 15, 10, 15, 25, 40, 30, 22, 25, 25, 40, 15, 12, 15, 8, 30]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="customers_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        return response
    
    @action(detail=False, methods=['post'])
    def import_excel(self, request):
        """Import customers from Excel."""
        import io
        import openpyxl
        from decimal import Decimal, InvalidOperation
        from django.db import transaction
        
        file = request.FILES.get('file')
        if not file:
            return Response({'error': '请上传文件'}, status=status.HTTP_400_BAD_REQUEST)
        
        if not file.name.endswith(('.xlsx', '.xls')):
            return Response({'error': '只支持Excel文件格式(.xlsx, .xls)'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file.read()))
            sheet = wb.active
            
            # Get headers from first row
            headers = [cell.value for cell in sheet[1]]
            
            # Map columns
            col_map = {
                '客户编码': 'code', '客户名称': 'name', '简称': 'short_name',
                '联系人': 'contact_person', '电话': 'phone', '邮箱': 'email', '地址': 'address',
                '开票名称': 'invoice_title', '税号': 'tax_number', '开户银行': 'bank_name',
                '银行账号': 'bank_account', '注册地址': 'registered_address', '注册电话': 'registered_phone',
                '信用额度': 'credit_limit', '付款条款': 'payment_terms', '状态': 'status', '备注': 'notes'
            }
            
            header_to_field = {}
            for idx, header in enumerate(headers):
                if header and header in col_map:
                    header_to_field[idx] = col_map[header]
            
            success_count = 0
            update_count = 0
            skip_count = 0
            errors = []
            
            # Track processed codes within this import
            processed_codes = set()
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                try:
                    data = {}
                    for col_idx, cell in enumerate(row):
                        if col_idx in header_to_field:
                            field = header_to_field[col_idx]
                            value = cell.value
                            if value is not None:
                                data[field] = str(value).strip() if not isinstance(value, (int, float)) else value

                    if not data.get('code') or not data.get('name'):
                        continue

                    # Check for in-file duplicate
                    code = str(data.get('code', '')).strip()
                    if code in processed_codes:
                        skip_count += 1
                        continue
                    processed_codes.add(code)

                    # Handle credit_limit
                    if 'credit_limit' in data:
                        try:
                            data['credit_limit'] = Decimal(str(data['credit_limit']))
                        except (InvalidOperation, ValueError):
                            data['credit_limit'] = Decimal('0')

                    # Handle status
                    status_map = {'激活': 'ACTIVE', '停用': 'INACTIVE', '潜在客户': 'POTENTIAL'}
                    if 'status' in data and data['status'] in status_map:
                        data['status'] = status_map[data['status']]
                    elif 'status' not in data:
                        data['status'] = 'ACTIVE'

                    # 每行独立事务(savepoint)：一行出错只回滚该行，不污染整批后续行
                    with transaction.atomic():
                        # 用 all_objects 查重(含软删除)：唯一约束 customer_code 不分软删，
                        # 若编码被软删客户占用，复活并更新，避免 create 撞唯一约束 IntegrityError
                        existing = Customer.all_objects.filter(code=data['code']).first()
                        if existing:
                            if existing.is_deleted:
                                existing.is_deleted = False
                                existing.deleted_at = None
                            for key, value in data.items():
                                if value:
                                    setattr(existing, key, value)
                            existing.save()
                            update_count += 1
                        else:
                            Customer.objects.create(**data)
                            success_count += 1

                except Exception as e:
                    errors.append({'row': row_idx, 'error': str(e)})
            
            return Response({
                'success_count': success_count,
                'update_count': update_count,
                'skip_count': skip_count,
                'error_count': len(errors),
                'errors': errors[:10]
            })
        
        except Exception as e:
            return Response({'error': f'导入失败: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['post'])
    def bulk_delete(self, request):
        """批量删除客户"""
        ids = request.data.get('ids', [])
        if not ids:
            return Response({'error': '请选择要删除的客户'}, status=status.HTTP_400_BAD_REQUEST)
        
        from django.utils import timezone as tz
        deleted_count = Customer.objects.filter(id__in=ids, is_deleted=False).update(
            is_deleted=True, deleted_at=tz.now()
        )
        return Response({
            'message': f'成功删除 {deleted_count} 个客户',
            'deleted_count': deleted_count
        })


class SupplierViewSet(PermissionMixin, SoftDeleteMixin, UserTrackingMixin, viewsets.ModelViewSet):
    permission_module = 'masterdata'
    permission_resource = 'supplier'
    allow_authenticated_read = True
    """
    ViewSet for Supplier management.
    NOTE: 供应商是主数据，所有用户都可以查看，不应用数据范围限制
    """
    queryset = Supplier.objects.all()
    serializer_class = SupplierSerializer
    filterset_fields = ['status', 'is_deleted']
    search_fields = ['code', 'name', 'contact_person', 'phone', 'tax_number']
    ordering_fields = ['code', 'created_at']
    
    def perform_create(self, serializer):
        """Auto-generate supplier code if not provided."""
        from apps.core.utils import generate_code
        
        code = serializer.validated_data.get('code', '').strip()
        if not code:
            code = generate_code('S', rule_type='SUPPLIER')
        serializer.save(code=code)
    
    @action(detail=False, methods=['get'])
    def download_template(self, request):
        """Download supplier import template Excel file."""
        import io
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '供应商导入模板'
        
        # 与列表页面显示的列对应
        headers = [
            '供应商编码', '供应商名称', '联系人', '电话', '税号', '开户银行', '地址', '状态'
        ]
        
        # Header style
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Example row - 与列表页面对应
        example = [
            'S001', '示例供应商名称', '李四', '13900139000',
            '91310000XXXXXXXX', '中国建设银行上海分行',
            '上海市浦东新区xxx路xxx号', '激活'
        ]
        for col, val in enumerate(example, 1):
            ws.cell(row=2, column=col, value=val)
        
        # Column widths - 与列表页面对应
        widths = [12, 25, 10, 15, 22, 20, 30, 8]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="supplier_import_template.xlsx"'
        return response
    
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Export suppliers to Excel."""
        import io
        import openpyxl
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        from django.utils import timezone
        
        queryset = self.get_queryset().filter(is_deleted=False)
        
        status_filter = request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '供应商列表'
        
        headers = [
            '供应商编码', '供应商名称', '简称', '联系人', '电话', '邮箱', '地址',
            '开票名称', '税号', '开户银行', '银行账号', '注册地址', '注册电话',
            '付款条款', '状态', '备注'
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        for row, supplier in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=supplier.code)
            ws.cell(row=row, column=2, value=supplier.name)
            ws.cell(row=row, column=3, value=supplier.short_name)
            ws.cell(row=row, column=4, value=supplier.contact_person)
            ws.cell(row=row, column=5, value=supplier.phone)
            ws.cell(row=row, column=6, value=supplier.email)
            ws.cell(row=row, column=7, value=supplier.address)
            ws.cell(row=row, column=8, value=supplier.invoice_title)
            ws.cell(row=row, column=9, value=supplier.tax_number)
            ws.cell(row=row, column=10, value=supplier.bank_name)
            ws.cell(row=row, column=11, value=supplier.bank_account)
            ws.cell(row=row, column=12, value=supplier.registered_address)
            ws.cell(row=row, column=13, value=supplier.registered_phone)
            ws.cell(row=row, column=14, value=supplier.payment_terms)
            ws.cell(row=row, column=15, value=supplier.get_status_display())
            ws.cell(row=row, column=16, value=supplier.notes)
        
        column_widths = [12, 30, 15, 10, 15, 25, 40, 30, 22, 25, 25, 40, 15, 15, 8, 30]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="suppliers_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        return response
    
    @action(detail=False, methods=['post'])
    def import_excel(self, request):
        """Import suppliers from Excel."""
        import io
        import openpyxl
        from django.db import transaction
        
        file = request.FILES.get('file')
        if not file:
            return Response({'error': '请上传文件'}, status=status.HTTP_400_BAD_REQUEST)
        
        if not file.name.endswith(('.xlsx', '.xls')):
            return Response({'error': '只支持Excel文件格式(.xlsx, .xls)'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file.read()))
            sheet = wb.active
            
            headers = [cell.value for cell in sheet[1]]
            
            col_map = {
                '供应商编码': 'code', '供应商名称': 'name', '简称': 'short_name',
                '联系人': 'contact_person', '电话': 'phone', '邮箱': 'email', '地址': 'address',
                '开票名称': 'invoice_title', '税号': 'tax_number', '开户银行': 'bank_name',
                '银行账号': 'bank_account', '注册地址': 'registered_address', '注册电话': 'registered_phone',
                '付款条款': 'payment_terms', '状态': 'status', '备注': 'notes'
            }
            
            header_to_field = {}
            for idx, header in enumerate(headers):
                if header and header in col_map:
                    header_to_field[idx] = col_map[header]
            
            success_count = 0
            update_count = 0
            skip_count = 0
            errors = []
            
            # Track processed codes within this import
            processed_codes = set()
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                try:
                    data = {}
                    for col_idx, cell in enumerate(row):
                        if col_idx in header_to_field:
                            field = header_to_field[col_idx]
                            value = cell.value
                            if value is not None:
                                data[field] = str(value).strip() if not isinstance(value, (int, float)) else value

                    if not data.get('code') or not data.get('name'):
                        continue

                    # Check for in-file duplicate
                    code = str(data.get('code', '')).strip()
                    if code in processed_codes:
                        skip_count += 1
                        continue
                    processed_codes.add(code)

                    status_map = {'激活': 'ACTIVE', '停用': 'INACTIVE', '潜在供应商': 'POTENTIAL'}
                    if 'status' in data and data['status'] in status_map:
                        data['status'] = status_map[data['status']]
                    elif 'status' not in data:
                        data['status'] = 'ACTIVE'

                    # 每行独立事务(savepoint)：一行出错只回滚该行，不污染整批后续行
                    with transaction.atomic():
                        # 用 all_objects 查重(含软删除)：唯一约束不分软删，
                        # 编码被软删供应商占用时复活更新，避免 create 撞唯一约束 IntegrityError
                        existing = Supplier.all_objects.filter(code=data['code']).first()
                        if existing:
                            if existing.is_deleted:
                                existing.is_deleted = False
                                existing.deleted_at = None
                            for key, value in data.items():
                                if value:
                                    setattr(existing, key, value)
                            existing.save()
                            update_count += 1
                        else:
                            Supplier.objects.create(**data)
                            success_count += 1

                except Exception as e:
                    errors.append({'row': row_idx, 'error': str(e)})
            
            return Response({
                'success_count': success_count,
                'update_count': update_count,
                'skip_count': skip_count,
                'error_count': len(errors),
                'errors': errors[:10]
            })
        
        except Exception as e:
            return Response({'error': f'导入失败: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['post'])
    def bulk_delete(self, request):
        """批量删除供应商"""
        ids = request.data.get('ids', [])
        if not ids:
            return Response({'error': '请选择要删除的供应商'}, status=status.HTTP_400_BAD_REQUEST)
        
        from django.utils import timezone as tz
        deleted_count = Supplier.objects.filter(id__in=ids, is_deleted=False).update(
            is_deleted=True, deleted_at=tz.now()
        )
        return Response({
            'message': f'成功删除 {deleted_count} 个供应商',
            'deleted_count': deleted_count
        })


class WarehouseViewSet(PermissionMixin, SoftDeleteMixin, UserTrackingMixin, viewsets.ModelViewSet):
    permission_module = 'masterdata'
    permission_resource = 'warehouse'
    allow_authenticated_read = True
    """
    ViewSet for Warehouse management.
    """
    queryset = Warehouse.objects.all()
    serializer_class = WarehouseSerializer
    filterset_fields = ['warehouse_type', 'is_active', 'is_deleted']
    search_fields = ['code', 'name', 'address']
    ordering_fields = ['code', 'created_at']
    
    @action(detail=True, methods=['get'])
    def locations(self, request, pk=None):
        """Get all locations for a warehouse."""
        warehouse = self.get_object()
        locations = warehouse.locations.filter(is_deleted=False).order_by('sort_order', 'code')
        serializer = WarehouseLocationSerializer(locations, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def location_tree(self, request, pk=None):
        """Get location tree for a warehouse."""
        warehouse = self.get_object()
        root_locations = warehouse.locations.filter(
            parent__isnull=True,
            is_deleted=False
        ).order_by('sort_order', 'code')
        serializer = WarehouseLocationTreeSerializer(root_locations, many=True)
        return Response(serializer.data)


class WarehouseLocationViewSet(PermissionMixin, SoftDeleteMixin, UserTrackingMixin, viewsets.ModelViewSet):
    permission_module = 'masterdata'
    permission_resource = 'warehouse_location'
    """
    ViewSet for WarehouseLocation management.
    """
    queryset = WarehouseLocation.objects.all()
    serializer_class = WarehouseLocationSerializer
    filterset_fields = ['warehouse', 'parent', 'location_type', 'is_active', 'is_deleted']
    search_fields = ['code', 'name', 'full_path']
    ordering_fields = ['warehouse', 'sort_order', 'code', 'created_at']

    def perform_destroy(self, instance):
        """软删除库位时级联软删除其全部子库位，避免产生孤儿库位。"""
        from django.utils import timezone

        now = timezone.now()
        targets = [instance, *instance.get_descendants()]
        for loc in targets:
            if not loc.is_deleted:
                loc.is_deleted = True
                loc.deleted_at = now
                loc.save(update_fields=['is_deleted', 'deleted_at'])

    @action(detail=False, methods=['get'])
    def tree(self, request):
        """Get location tree for all warehouses."""
        warehouse_id = request.query_params.get('warehouse')
        
        if warehouse_id:
            root_locations = self.get_queryset().filter(
                warehouse_id=warehouse_id,
                parent__isnull=True,
                is_deleted=False
            ).order_by('sort_order', 'code')
        else:
            root_locations = self.get_queryset().filter(
                parent__isnull=True,
                is_deleted=False
            ).order_by('warehouse', 'sort_order', 'code')
        
        serializer = WarehouseLocationTreeSerializer(root_locations, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def children(self, request, pk=None):
        """Get direct children of a location."""
        location = self.get_object()
        children = location.children.filter(is_deleted=False).order_by('sort_order', 'code')
        serializer = WarehouseLocationSerializer(children, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def path(self, request, pk=None):
        """Get the full path from root to this location."""
        location = self.get_object()
        path = [location]
        current = location
        
        while current.parent:
            current = current.parent
            path.insert(0, current)
        
        serializer = WarehouseLocationSerializer(path, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['post'])
    def bulk_create(self, request):
        """Bulk create locations for a warehouse."""
        warehouse_id = request.data.get('warehouse')
        locations_data = request.data.get('locations', [])
        
        if not warehouse_id:
            return Response(
                {'error': '请选择仓库'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            warehouse = Warehouse.objects.get(id=warehouse_id)
        except Warehouse.DoesNotExist:
            return Response(
                {'error': '仓库不存在'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        created = []
        errors = []
        
        for loc_data in locations_data:
            try:
                location = WarehouseLocation.objects.create(
                    warehouse=warehouse,
                    code=loc_data.get('code'),
                    name=loc_data.get('name'),
                    location_type=loc_data.get('location_type', 'BIN'),
                    parent_id=loc_data.get('parent'),
                    created_by=request.user
                )
                created.append(WarehouseLocationSerializer(location).data)
            except Exception as e:
                errors.append({
                    'code': loc_data.get('code'),
                    'error': str(e)
                })
        
        return Response({
            'created': len(created),
            'errors': errors,
            'locations': created
        })

