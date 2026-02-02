"""
CAD BOM统一导入模块 - Unified CAD BOM Import
支持: CREO/SolidWorks/AutoCAD/Inventor/CATIA/UG等CAD软件BOM导入
自动识别CAD软件类型，匹配或创建物料(Item)
"""
import os, re, csv, json, logging
from decimal import Decimal
from datetime import datetime
from difflib import SequenceMatcher
from typing import List, Dict, Optional, Tuple
import pandas as pd

from django.db import models, transaction
from django.conf import settings
from django.utils import timezone
from rest_framework import viewsets, serializers, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated

from apps.core.models import BaseModel
from apps.core.mixins import SoftDeleteMixin, UserTrackingMixin

logger = logging.getLogger(__name__)

# 支持的CAD软件类型
CAD_SOFTWARE_TYPES = [
    ('AUTO', '自动识别'),
    ('CREO', 'Creo/Pro-E'),
    ('SOLIDWORKS', 'SolidWorks'),
    ('AUTOCAD', 'AutoCAD'),
    ('INVENTOR', 'Inventor'),
    ('CATIA', 'CATIA'),
    ('UG', 'UG/NX'),
    ('FUSION360', 'Fusion 360'),
    ('GENERIC', '通用格式'),
]

# 各CAD软件BOM列名特征（用于自动识别）
CAD_SIGNATURE_COLUMNS = {
    'CREO': ['PTC_WM_', 'CREO_', 'PRO_E_', '.prt', '.asm', '.drw'],
    'SOLIDWORKS': ['SW-', 'SOLIDWORKS', 'SLDPRT', 'SLDASM', 'Configuration'],
    'AUTOCAD': ['ACAD_', 'AUTOCAD', 'DWG', 'Block Name', 'Layer'],
    'INVENTOR': ['IPT', 'IAM', 'Autodesk', 'iProperty'],
    'CATIA': ['CATIA', 'CATPART', 'CATPRODUCT', 'V5_', 'V6_'],
    'UG': ['NX_', 'UG_', 'UGII', 'Unigraphics'],
}

# 统一BOM列名映射（多CAD软件兼容，与系统BOM格式保持一致）
UNIFIED_BOM_COLUMN_MAPPING = {
    # 层级
    'Level': 'level', 'LVL': 'level', 'Lvl': 'level', '层级': 'level', '等级': 'level',
    'LEVEL': 'level', 'Indent': 'level', 'Tree Level': 'level',
    
    # 物料编码（映射到系统字段item_code）
    'Part Number': 'part_number', 'Part No': 'part_number', 'Part No.': 'part_number',
    'PartNumber': 'part_number', 'PART_NUMBER': 'part_number', 'Item Number': 'part_number',
    'Item No': 'part_number', 'Component': 'part_number', 'P/N': 'part_number',
    '物料编码': 'part_number', '料号': 'part_number', '零件号': 'part_number', '编码': 'part_number',
    'PTC_WM_PART_NUMBER': 'part_number', 'SW-Part Number': 'part_number',
    
    # 文件名（CAD特有）
    'Name': 'file_name', 'File Name': 'file_name', 'FileName': 'file_name',
    'Component Name': 'file_name', 'Model Name': 'file_name',
    '文件名': 'file_name', '组件名称': 'file_name', '模型名': 'file_name',
    
    # 物料名称（映射到系统字段item_name）
    'Description': 'description', 'Desc': 'description', 'DESC': 'description',
    'Part Name': 'description', 'Title': 'description',
    '名称': 'description', '描述': 'description', '零件名称': 'description', '物料名称': 'description',
    'PTC_WM_DESCRIPTION': 'description', 'SW-Description': 'description',
    
    # 数量
    'Qty': 'quantity', 'QTY': 'quantity', 'Quantity': 'quantity', 'Count': 'quantity',
    'Amount': 'quantity', 'Cnt': 'quantity',
    '数量': 'quantity', '用量': 'quantity', '计划数量': 'quantity',
    
    # 单位
    'Unit': 'unit', 'Units': 'unit', 'UOM': 'unit', 'Unit of Measure': 'unit',
    '单位': 'unit',
    
    # 材质规格（映射到系统字段material_spec）
    'Material': 'material', 'Mat': 'material', 'MAT': 'material',
    '材料': 'material', '材质': 'material', '材质规格': 'material',
    'PTC_WM_MATERIAL': 'material', 'SW-Material': 'material',
    
    # 规格型号
    'Specification': 'specification', 'Spec': 'specification', 'Model': 'specification',
    '规格型号': 'specification', '规格': 'specification', '型号': 'specification',
    
    # 物料类型
    'Type': 'component_type', 'Part Type': 'component_type', 'Component Type': 'component_type',
    '类型': 'component_type', '零件类型': 'component_type', '物料类型': 'component_type',
    
    # 有图/无图
    'Has Drawing': 'has_drawing', 'Drawing': 'has_drawing',
    '有图/无图': 'has_drawing', '图纸': 'has_drawing',
    
    # 版本/品牌
    'Revision': 'revision', 'Rev': 'revision', 'REV': 'revision', 'Version': 'revision',
    'Brand': 'revision', '版本': 'revision', '修订': 'revision', '品牌': 'revision', '版本/品牌': 'revision',
    
    # 重量
    'Weight': 'weight', 'Mass': 'weight', 'MASS': 'weight',
    '重量': 'weight', '质量': 'weight',
    
    # 供应商
    'Vendor': 'vendor', 'Supplier': 'vendor', 'VENDOR': 'vendor',
    '供应商': 'vendor', '厂家': 'vendor',
    
    # 需求日期
    'Required Date': 'required_date', 'Due Date': 'required_date', 'Need Date': 'required_date',
    '需求日期': 'required_date', '交期': 'required_date',
    
    # 申请人
    'Requester': 'requester', 'Requested By': 'requester', 'Owner': 'requester',
    '申请人': 'requester', '需求人': 'requester',
}

ITEM_PROPERTY_RULES = {
    'prefix_rules': {
        'STD': 'STANDARD', 'GB': 'STANDARD', 'DIN': 'STANDARD', 'ISO': 'STANDARD', 'JIS': 'STANDARD',
        'PUR': 'PURCHASED', 'BUY': 'PURCHASED', 'ELE': 'PURCHASED',
        'OUT': 'OUTSOURCED', 'OEM': 'OUTSOURCED', 'SUB': 'OUTSOURCED',
        'MFG': 'SELF_MADE', 'MAKE': 'SELF_MADE', 'FAB': 'SELF_MADE',
        'ASM': 'ASSEMBLY', 'ASSY': 'ASSEMBLY',
    },
    'material_rules': {
        'STANDARD': ['螺丝', '螺母', '螺栓', '垫圈', '轴承', '密封圈', 'O型圈', '弹簧', '销', '键', '挡圈'],
        'PURCHASED': ['电机', 'PLC', '传感器', '气缸', '电磁阀', '继电器', '变频器', '触摸屏', '伺服', '步进'],
    },
}

# 材质规格标准化映射（非标自动化行业常用材质）
MATERIAL_MAPPING = {
    # 碳钢
    'Q235': {'category': '碳钢', 'grade': 'Q235', 'standard': 'GB/T 700'},
    'Q235A': {'category': '碳钢', 'grade': 'Q235A', 'standard': 'GB/T 700'},
    'Q235B': {'category': '碳钢', 'grade': 'Q235B', 'standard': 'GB/T 700'},
    'Q345': {'category': '碳钢', 'grade': 'Q345', 'standard': 'GB/T 1591'},
    'Q345B': {'category': '碳钢', 'grade': 'Q345B', 'standard': 'GB/T 1591'},
    '45': {'category': '碳钢', 'grade': '45号钢', 'standard': 'GB/T 699'},
    '45#': {'category': '碳钢', 'grade': '45号钢', 'standard': 'GB/T 699'},
    '45钢': {'category': '碳钢', 'grade': '45号钢', 'standard': 'GB/T 699'},
    '40CR': {'category': '合金钢', 'grade': '40Cr', 'standard': 'GB/T 3077'},
    '42CRMO': {'category': '合金钢', 'grade': '42CrMo', 'standard': 'GB/T 3077'},
    'GCR15': {'category': '轴承钢', 'grade': 'GCr15', 'standard': 'GB/T 18254'},
    
    # 不锈钢
    'SUS304': {'category': '不锈钢', 'grade': '304', 'standard': 'JIS G4303'},
    '304': {'category': '不锈钢', 'grade': '304', 'standard': 'GB/T 1220'},
    '0CR18NI9': {'category': '不锈钢', 'grade': '304', 'standard': 'GB/T 1220'},
    'SUS316': {'category': '不锈钢', 'grade': '316', 'standard': 'JIS G4303'},
    '316': {'category': '不锈钢', 'grade': '316', 'standard': 'GB/T 1220'},
    '316L': {'category': '不锈钢', 'grade': '316L', 'standard': 'GB/T 1220'},
    'SUS316L': {'category': '不锈钢', 'grade': '316L', 'standard': 'JIS G4303'},
    '201': {'category': '不锈钢', 'grade': '201', 'standard': 'GB/T 1220'},
    '303': {'category': '不锈钢', 'grade': '303', 'standard': 'GB/T 1220'},
    
    # 铝合金
    '6061': {'category': '铝合金', 'grade': '6061-T6', 'standard': 'GB/T 3190'},
    '6061-T6': {'category': '铝合金', 'grade': '6061-T6', 'standard': 'GB/T 3190'},
    '6063': {'category': '铝合金', 'grade': '6063', 'standard': 'GB/T 3190'},
    '7075': {'category': '铝合金', 'grade': '7075-T6', 'standard': 'GB/T 3190'},
    '7075-T6': {'category': '铝合金', 'grade': '7075-T6', 'standard': 'GB/T 3190'},
    '2024': {'category': '铝合金', 'grade': '2024', 'standard': 'GB/T 3190'},
    '5052': {'category': '铝合金', 'grade': '5052', 'standard': 'GB/T 3190'},
    'AL': {'category': '铝合金', 'grade': '铝', 'standard': ''},
    'ALUMINUM': {'category': '铝合金', 'grade': '铝', 'standard': ''},
    
    # 铜合金
    'H62': {'category': '黄铜', 'grade': 'H62', 'standard': 'GB/T 5231'},
    'H59': {'category': '黄铜', 'grade': 'H59', 'standard': 'GB/T 5231'},
    'T2': {'category': '紫铜', 'grade': 'T2', 'standard': 'GB/T 5231'},
    'C3604': {'category': '黄铜', 'grade': 'C3604', 'standard': 'JIS H3250'},
    'BRASS': {'category': '黄铜', 'grade': '黄铜', 'standard': ''},
    'COPPER': {'category': '紫铜', 'grade': '紫铜', 'standard': ''},
    
    # 工程塑料
    'POM': {'category': '工程塑料', 'grade': 'POM', 'standard': '', 'alias': '赛钢'},
    '赛钢': {'category': '工程塑料', 'grade': 'POM', 'standard': ''},
    'PEEK': {'category': '工程塑料', 'grade': 'PEEK', 'standard': ''},
    'PTFE': {'category': '工程塑料', 'grade': 'PTFE', 'standard': '', 'alias': '铁氟龙'},
    '铁氟龙': {'category': '工程塑料', 'grade': 'PTFE', 'standard': ''},
    'TEFLON': {'category': '工程塑料', 'grade': 'PTFE', 'standard': ''},
    'PA66': {'category': '工程塑料', 'grade': 'PA66', 'standard': '', 'alias': '尼龙66'},
    'NYLON': {'category': '工程塑料', 'grade': 'PA', 'standard': ''},
    '尼龙': {'category': '工程塑料', 'grade': 'PA', 'standard': ''},
    'MC901': {'category': '工程塑料', 'grade': 'MC901', 'standard': '', 'alias': '蓝色尼龙'},
    'UHMWPE': {'category': '工程塑料', 'grade': 'UHMWPE', 'standard': '', 'alias': '超高分子量聚乙烯'},
    'PC': {'category': '工程塑料', 'grade': 'PC', 'standard': '', 'alias': '聚碳酸酯'},
    'ACRYLIC': {'category': '工程塑料', 'grade': 'PMMA', 'standard': '', 'alias': '亚克力'},
    '亚克力': {'category': '工程塑料', 'grade': 'PMMA', 'standard': ''},
    'ABS': {'category': '工程塑料', 'grade': 'ABS', 'standard': ''},
    'PVC': {'category': '工程塑料', 'grade': 'PVC', 'standard': ''},
    
    # 橡胶
    'NBR': {'category': '橡胶', 'grade': 'NBR', 'standard': '', 'alias': '丁腈橡胶'},
    '丁腈': {'category': '橡胶', 'grade': 'NBR', 'standard': ''},
    'EPDM': {'category': '橡胶', 'grade': 'EPDM', 'standard': '', 'alias': '三元乙丙'},
    'FKM': {'category': '橡胶', 'grade': 'FKM', 'standard': '', 'alias': '氟橡胶'},
    'VITON': {'category': '橡胶', 'grade': 'FKM', 'standard': ''},
    '硅胶': {'category': '橡胶', 'grade': 'VMQ', 'standard': ''},
    'SILICONE': {'category': '橡胶', 'grade': 'VMQ', 'standard': ''},
    'PU': {'category': '橡胶', 'grade': 'PU', 'standard': '', 'alias': '聚氨酯'},
}

# 表面处理标准化
SURFACE_TREATMENT_MAPPING = {
    '镀锌': {'code': 'ZN', 'name': '镀锌', 'standard': 'GB/T 9799'},
    '镀锌发蓝': {'code': 'ZN-BL', 'name': '镀锌发蓝', 'standard': ''},
    '镀镍': {'code': 'NI', 'name': '镀镍', 'standard': 'GB/T 9798'},
    '镀铬': {'code': 'CR', 'name': '镀铬', 'standard': 'GB/T 11379'},
    '硬铬': {'code': 'HCR', 'name': '硬铬', 'standard': 'GB/T 11379'},
    '发黑': {'code': 'BK', 'name': '发黑', 'standard': 'GB/T 15519'},
    '阳极氧化': {'code': 'ANOD', 'name': '阳极氧化', 'standard': 'GB/T 8013'},
    '硬质阳极氧化': {'code': 'HARD-ANOD', 'name': '硬质阳极氧化', 'standard': 'GB/T 8013'},
    '喷涂': {'code': 'PAINT', 'name': '喷涂', 'standard': ''},
    '喷塑': {'code': 'POWDER', 'name': '喷塑', 'standard': ''},
    '达克罗': {'code': 'DACRO', 'name': '达克罗', 'standard': ''},
    '钝化': {'code': 'PASS', 'name': '钝化', 'standard': ''},
    '抛光': {'code': 'POLISH', 'name': '抛光', 'standard': ''},
    '拉丝': {'code': 'BRUSH', 'name': '拉丝', 'standard': ''},
    '喷砂': {'code': 'SANDBLAST', 'name': '喷砂', 'standard': ''},
    '电泳': {'code': 'ED', 'name': '电泳', 'standard': ''},
    '镀锡': {'code': 'SN', 'name': '镀锡', 'standard': ''},
}


class MaterialExtractionService:
    """材质规格提取与标准化服务"""
    
    @classmethod
    def extract_material(cls, material_str: str) -> Dict:
        """
        从材质字符串提取并标准化材质信息
        
        Args:
            material_str: 原始材质字符串，如 "SUS304 Φ20" 或 "Q235B 10*50*100"
            
        Returns:
            {
                'category': '不锈钢',
                'grade': '304',
                'standard': 'JIS G4303',
                'dimension': 'Φ20',
                'original': 'SUS304 Φ20',
                'normalized': '304不锈钢 Φ20'
            }
        """
        if not material_str:
            return {'category': '', 'grade': '', 'standard': '', 'dimension': '', 'original': '', 'normalized': ''}
        
        original = material_str.strip()
        upper = original.upper()
        
        # 尝试匹配标准材质
        matched_material = None
        matched_key = ''
        
        for key, info in MATERIAL_MAPPING.items():
            if key.upper() in upper:
                # 优先匹配更长的关键词
                if not matched_key or len(key) > len(matched_key):
                    matched_material = info
                    matched_key = key
        
        if matched_material:
            # 提取尺寸信息（去除材质关键词后的部分）
            dimension = cls._extract_dimension(original, matched_key)
            
            return {
                'category': matched_material['category'],
                'grade': matched_material['grade'],
                'standard': matched_material.get('standard', ''),
                'dimension': dimension,
                'original': original,
                'normalized': f"{matched_material['grade']}{matched_material['category']} {dimension}".strip()
            }
        
        # 未匹配到标准材质，尝试提取尺寸
        dimension = cls._extract_dimension(original, '')
        
        return {
            'category': '',
            'grade': '',
            'standard': '',
            'dimension': dimension,
            'original': original,
            'normalized': original
        }
    
    @classmethod
    def _extract_dimension(cls, material_str: str, material_key: str) -> str:
        """提取尺寸规格信息"""
        # 移除材质关键词
        remaining = material_str
        if material_key:
            remaining = re.sub(re.escape(material_key), '', material_str, flags=re.IGNORECASE).strip()
        
        # 常见尺寸格式匹配
        dimension_patterns = [
            r'[Φφ]\s*\d+(?:\.\d+)?(?:\s*[xX×*]\s*\d+(?:\.\d+)?)*',  # Φ20, Φ20*100
            r'\d+(?:\.\d+)?\s*[xX×*]\s*\d+(?:\.\d+)?(?:\s*[xX×*]\s*\d+(?:\.\d+)?)*',  # 10*50*100
            r'\d+(?:\.\d+)?\s*[mM]{2}',  # 10mm
            r'DN\s*\d+',  # DN50
            r'M\d+(?:\.\d+)?(?:\s*[xX×*]\s*\d+)?',  # M10, M10*50
        ]
        
        for pattern in dimension_patterns:
            match = re.search(pattern, remaining)
            if match:
                return match.group().strip()
        
        # 返回去除材质后的剩余部分
        return remaining.strip()
    
    @classmethod
    def extract_surface_treatment(cls, treatment_str: str) -> Dict:
        """
        提取并标准化表面处理信息
        
        Returns:
            {'code': 'ZN', 'name': '镀锌', 'standard': 'GB/T 9799', 'original': '...'}
        """
        if not treatment_str:
            return {'code': '', 'name': '', 'standard': '', 'original': ''}
        
        original = treatment_str.strip()
        
        for key, info in SURFACE_TREATMENT_MAPPING.items():
            if key in original:
                return {
                    'code': info['code'],
                    'name': info['name'],
                    'standard': info.get('standard', ''),
                    'original': original
                }
        
        return {
            'code': '',
            'name': original,
            'standard': '',
            'original': original
        }
    
    @classmethod
    def batch_extract_materials(cls, items: List[Dict]) -> List[Dict]:
        """批量提取材质信息并更新到items"""
        for item in items:
            material_str = item.get('material', '')
            if material_str:
                extracted = cls.extract_material(material_str)
                item['material_category'] = extracted['category']
                item['material_grade'] = extracted['grade']
                item['material_standard'] = extracted['standard']
                item['material_dimension'] = extracted['dimension']
                item['material_normalized'] = extracted['normalized']
        return items


class CreoBOMImportSession(BaseModel):
    """CAD BOM统一导入会话"""
    STATUS_CHOICES = [('UPLOADING', '上传中'), ('PARSING', '解析中'), ('MATCHING', '匹配中'),
        ('REVIEWING', '待确认'), ('IMPORTING', '导入中'), ('COMPLETED', '已完成'),
        ('FAILED', '失败'), ('CANCELLED', '已取消')]
    FORMAT_CHOICES = [('CSV', 'CSV'), ('XML', 'XML'), ('XLSX', 'Excel'), ('TXT', '文本')]
    
    name = models.CharField(max_length=200, verbose_name='导入名称')
    file_format = models.CharField(max_length=20, choices=FORMAT_CHOICES, default='CSV')
    cad_software = models.CharField(max_length=20, choices=CAD_SOFTWARE_TYPES, default='AUTO', verbose_name='CAD软件')
    detected_software = models.CharField(max_length=20, blank=True, verbose_name='检测到的CAD软件')
    status = models.CharField(max_length=20, choices=STATUS_CHOICES, default='UPLOADING')
    source_file = models.FileField(upload_to='cad_bom/', null=True, blank=True)
    source_file_name = models.CharField(max_length=255, blank=True)
    project_id = models.IntegerField(null=True, blank=True)
    column_mapping = models.JSONField(default=dict, blank=True)
    import_options = models.JSONField(default=dict, blank=True)
    total_rows = models.IntegerField(default=0)
    matched_count = models.IntegerField(default=0)
    new_item_count = models.IntegerField(default=0)
    imported_count = models.IntegerField(default=0)
    error_count = models.IntegerField(default=0)
    started_at = models.DateTimeField(null=True, blank=True)
    completed_at = models.DateTimeField(null=True, blank=True)
    error_log = models.TextField(blank=True)
    
    class Meta:
        db_table = 'plm_creo_bom_session'
        verbose_name = 'CAD BOM导入'


class CreoBOMImportItem(BaseModel):
    """CREO BOM导入项"""
    STATUS_CHOICES = [('PENDING', '待处理'), ('MATCHED', '已匹配'), ('NEW', '新物料'),
        ('CREATED', '已创建'), ('IMPORTED', '已导入'), ('SKIPPED', '跳过'), ('ERROR', '错误')]
    
    session = models.ForeignKey(CreoBOMImportSession, on_delete=models.CASCADE, related_name='items')
    row_number = models.IntegerField(default=0)
    level = models.IntegerField(default=0)
    raw_data = models.JSONField(default=dict)
    part_number = models.CharField(max_length=100, blank=True)
    part_name = models.CharField(max_length=200, blank=True)
    file_name = models.CharField(max_length=255, blank=True)
    quantity = models.DecimalField(max_digits=15, decimal_places=4, default=1)
    unit = models.CharField(max_length=20, default='PCS')
    material = models.CharField(max_length=100, blank=True)
    suggested_item_property = models.CharField(max_length=20, blank=True)
    status = models.CharField(max_length=20, choices=STATUS_CHOICES, default='PENDING')
    match_type = models.CharField(max_length=20, blank=True)
    matched_item_id = models.IntegerField(null=True, blank=True)
    match_score = models.FloatField(default=0)
    match_candidates = models.JSONField(default=list, blank=True)
    created_item_id = models.IntegerField(null=True, blank=True)
    imported_bom_id = models.IntegerField(null=True, blank=True)
    error_message = models.TextField(blank=True)
    
    class Meta:
        db_table = 'plm_creo_bom_item'
        ordering = ['session', 'row_number']


class CreoBOMParserService:
    """CAD BOM统一解析服务"""
    
    @classmethod
    def parse_file(cls, file_path, file_format, column_mapping=None, options=None) -> Tuple[List[Dict], str]:
        """解析BOM文件，返回(items, detected_software)"""
        column_mapping = column_mapping or UNIFIED_BOM_COLUMN_MAPPING
        options = options or {}
        
        if file_format == 'CSV':
            items = cls._parse_csv(file_path, column_mapping, options)
        elif file_format == 'XML':
            items = cls._parse_xml(file_path, column_mapping)
        elif file_format == 'XLSX':
            items = cls._parse_xlsx(file_path, column_mapping, options)
        else:
            items = cls._parse_csv(file_path, column_mapping, options)
        
        # 自动识别CAD软件类型
        detected = cls._detect_cad_software(items)
        return items, detected
    
    @classmethod
    def _detect_cad_software(cls, items: List[Dict]) -> str:
        """根据BOM内容自动识别CAD软件"""
        if not items:
            return 'GENERIC'
        
        # 收集所有原始数据的键和值
        all_text = []
        for item in items[:10]:  # 只检查前10行
            raw = item.get('raw_data', {})
            for k, v in raw.items():
                all_text.append(str(k).upper())
                all_text.append(str(v).upper() if v else '')
            # 检查文件名特征
            fn = item.get('file_name', '').upper()
            all_text.append(fn)
        
        all_text_str = ' '.join(all_text)
        
        # 检查各CAD软件的特征
        scores = {}
        for cad, signatures in CAD_SIGNATURE_COLUMNS.items():
            score = sum(1 for sig in signatures if sig.upper() in all_text_str)
            if score > 0:
                scores[cad] = score
        
        if scores:
            return max(scores, key=scores.get)
        
        # 根据文件扩展名判断
        for item in items:
            fn = item.get('file_name', '').lower()
            if '.prt' in fn or '.asm' in fn or '.drw' in fn:
                return 'CREO'
            elif '.sldprt' in fn or '.sldasm' in fn:
                return 'SOLIDWORKS'
            elif '.ipt' in fn or '.iam' in fn:
                return 'INVENTOR'
            elif '.catpart' in fn or '.catproduct' in fn:
                return 'CATIA'
        
        return 'GENERIC'
    
    @classmethod
    def _parse_csv(cls, file_path, column_mapping, options):
        items = []
        for enc in ['utf-8', 'gbk', 'gb2312', 'utf-8-sig']:
            try:
                with open(file_path, 'r', encoding=enc) as f:
                    for row_num, row in enumerate(csv.DictReader(f), 1):
                        item = cls._map_row(row, column_mapping, row_num)
                        if item: items.append(item)
                break
            except UnicodeDecodeError: continue
        return items
    
    @classmethod
    def _parse_xml(cls, file_path, column_mapping):
        items = []
        try:
            import xml.etree.ElementTree as ET
            root = ET.parse(file_path).getroot()
            for tag in ['Component', 'Part', 'Item', 'Row']:
                for row_num, comp in enumerate(root.findall(f'.//{tag}'), 1):
                    raw = {**comp.attrib}
                    for c in comp:
                        if c.text: raw[c.tag] = c.text.strip()
                    item = cls._map_row(raw, column_mapping, row_num)
                    if item: items.append(item)
                if items: break
        except Exception as e: logger.error(f"XML解析错误: {e}")
        return items
    
    @classmethod
    def _parse_xlsx(cls, file_path, column_mapping, options):
        items = []
        try:
            import openpyxl
            ws = openpyxl.load_workbook(file_path, data_only=True).active
            headers = [c.value or f'col_{c.column}' for c in ws[1]]
            for row_num, row in enumerate(ws.iter_rows(min_row=2), 1):
                row_data = {headers[i]: c.value for i, c in enumerate(row) if i < len(headers)}
                item = cls._map_row(row_data, column_mapping, row_num)
                if item: items.append(item)
        except Exception as e: logger.error(f"Excel解析错误: {e}")
        return items
    
    @classmethod
    def _map_row(cls, row_data, column_mapping, row_num):
        item = {'row_number': row_num, 'raw_data': row_data, 'level': 0, 'part_number': '',
                'part_name': '', 'file_name': '', 'quantity': 1, 'unit': 'PCS', 'material': ''}
        for col, val in row_data.items():
            if not col or not val: continue
            col_str, val_str = str(col).strip(), str(val).strip()
            for k, field in column_mapping.items():
                if col_str.lower() == k.lower():
                    if field == 'level':
                        try: item['level'] = int(val_str)
                        except: pass
                    elif field == 'quantity':
                        try: item['quantity'] = float(val_str)
                        except: pass
                    elif field in item: item[field] = val_str
                    break
        if not item['part_number'] and item['file_name']:
            item['part_number'] = re.sub(r'\.\d+$', '', re.sub(r'\.(prt|asm|drw)$', '', item['file_name'], flags=re.I)).upper()
        if not item['part_name']: item['part_name'] = item['part_number'] or item['file_name']
        return item if item['part_number'] or item['part_name'] else None


class ItemMatchingService:
    """物料匹配"""
    
    @classmethod
    def match_items(cls, items, options=None):
        from apps.masterdata.models import Item
        threshold = (options or {}).get('fuzzy_threshold', 0.7)
        all_items = list(Item.objects.filter(is_deleted=False).values('id', 'sku', 'name', 'specification'))
        sku_idx = {i['sku'].upper(): i for i in all_items}
        
        for item in items:
            pn = (item.get('part_number') or '').upper()
            name = (item.get('part_name') or '').lower()
            
            if pn in sku_idx:
                m = sku_idx[pn]
                item.update({'matched_item_id': m['id'], 'match_score': 1.0, 'match_type': 'EXACT', 'status': 'MATCHED'})
                continue
            
            best, best_score, cands = None, 0, []
            for db in all_items:
                score = (SequenceMatcher(None, pn, db['sku'].upper()).ratio() * 0.5 +
                         SequenceMatcher(None, name, db['name'].lower()).ratio() * 0.5)
                if score >= threshold:
                    cands.append({'id': db['id'], 'sku': db['sku'], 'name': db['name'], 'score': score})
                    if score > best_score: best, best_score = db, score
            
            item['match_candidates'] = sorted(cands, key=lambda x: -x['score'])[:5]
            if best:
                item.update({'matched_item_id': best['id'], 'match_score': best_score, 'match_type': 'FUZZY', 'status': 'MATCHED'})
            else:
                prop = 'PURCHASED'
                for pfx, p in ITEM_PROPERTY_RULES['prefix_rules'].items():
                    if pn.startswith(pfx): prop = p; break
                item.update({'status': 'NEW', 'match_type': 'NEW', 'suggested_item_property': prop})
        return items


class ItemCreationService:
    """物料创建"""
    
    @classmethod
    def create_from_bom(cls, items, options=None):
        from apps.masterdata.models import Item
        options = options or {}
        prefix = options.get('sku_prefix', 'CREO-')
        cat_id = options.get('default_category_id')
        created, errors = 0, []
        
        for item in items:
            if item.get('status') != 'NEW': continue
            try:
                with transaction.atomic():
                    sku = item.get('part_number') or f"{prefix}{datetime.now().strftime('%Y%m%d%H%M%S')}"
                    if Item.objects.filter(sku=sku, is_deleted=False).exists():
                        item.update({'error_message': f'SKU已存在: {sku}', 'status': 'ERROR'})
                        errors.append(f'SKU已存在: {sku}')
                        continue
                    new = Item.objects.create(
                        sku=sku, name=item.get('part_name', sku),
                        specification=item.get('material', ''),
                        item_property=item.get('suggested_item_property', 'PURCHASED'),
                        unit=item.get('unit', 'PCS'), category_id=cat_id, is_active=True,
                    )
                    item.update({'created_item_id': new.id, 'matched_item_id': new.id, 'status': 'CREATED'})
                    created += 1
            except Exception as e:
                item.update({'error_message': str(e), 'status': 'ERROR'})
                errors.append(str(e))
        return {'created': created, 'errors': errors}


class CreoBOMImportService:
    """CAD BOM导入服务"""
    
    @classmethod
    def parse_and_match(cls, session, file_path):
        session.status, session.started_at = 'PARSING', timezone.now()
        session.save()
        try:
            items, detected_software = CreoBOMParserService.parse_file(
                file_path, session.file_format, 
                session.column_mapping or UNIFIED_BOM_COLUMN_MAPPING, 
                session.import_options
            )
            session.total_rows = len(items)
            session.detected_software = detected_software
            session.status = 'MATCHING'
            session.save()
            
            matched_items = ItemMatchingService.match_items(items, session.import_options)
            mc, nc = 0, 0
            for d in matched_items:
                CreoBOMImportItem.objects.create(
                    session=session, row_number=d.get('row_number', 0), level=d.get('level', 0),
                    raw_data=d.get('raw_data', {}), part_number=d.get('part_number', ''),
                    part_name=d.get('part_name', ''), file_name=d.get('file_name', ''),
                    quantity=Decimal(str(d.get('quantity', 1))), unit=d.get('unit', 'PCS'),
                    material=d.get('material', ''), status=d.get('status', 'PENDING'),
                    match_type=d.get('match_type', ''), matched_item_id=d.get('matched_item_id'),
                    match_score=d.get('match_score', 0), match_candidates=d.get('match_candidates', []),
                    suggested_item_property=d.get('suggested_item_property', ''),
                )
                if d.get('status') == 'MATCHED': mc += 1
                elif d.get('status') == 'NEW': nc += 1
            session.matched_count, session.new_item_count = mc, nc
            session.status = 'REVIEWING'
            session.save()
            return {'success': True, 'total': len(items), 'matched': mc, 'new': nc}
        except Exception as e:
            session.status, session.error_log = 'FAILED', str(e)
            session.save()
            return {'success': False, 'error': str(e)}
    
    @classmethod
    def create_new_items(cls, session, options=None):
        new_items = session.items.filter(status='NEW')
        items_data = list(new_items.values())
        result = ItemCreationService.create_from_bom(items_data, options)
        for d in items_data:
            if d.get('created_item_id'):
                new_items.filter(id=d['id']).update(
                    created_item_id=d['created_item_id'], matched_item_id=d['created_item_id'], status='CREATED')
        return result
    
    @classmethod
    def import_to_bom(cls, session):
        from apps.projects.models import Project, ProjectBOM
        if not session.project_id: return {'success': False, 'error': '未指定项目'}
        try: project = Project.objects.get(id=session.project_id)
        except: return {'success': False, 'error': '项目不存在'}
        
        session.status = 'IMPORTING'
        session.save()
        imp, skip, err = 0, 0, 0
        
        for item in session.items.filter(status__in=['MATCHED', 'CREATED']):
            item_id = item.matched_item_id or item.created_item_id
            if not item_id:
                item.status, item.error_message = 'SKIPPED', '无匹配物料'
                item.save(); skip += 1; continue
            try:
                with transaction.atomic():
                    ex = ProjectBOM.objects.filter(project=project, item_id=item_id, is_deleted=False).first()
                    if ex: ex.planned_qty += item.quantity; ex.save(); bom = ex
                    else: bom = ProjectBOM.objects.create(project=project, item_id=item_id, 
                            planned_qty=item.quantity, level=item.level, notes=f'CREO导入: {item.file_name}')
                    item.imported_bom_id, item.status = bom.id, 'IMPORTED'
                    item.save(); imp += 1
            except Exception as e:
                item.status, item.error_message = 'ERROR', str(e)
                item.save(); err += 1
        
        session.imported_count, session.error_count = imp, err
        session.status = 'COMPLETED' if err == 0 else 'FAILED'
        session.completed_at = timezone.now()
        session.save()
        return {'success': True, 'imported': imp, 'skipped': skip, 'errors': err}
    
    @classmethod
    def import_hierarchical_bom(cls, session, options=None):
        """
        导入层级BOM结构（支持装配体-子装配-零件父子关系）
        根据level字段建立parent关系
        """
        from apps.projects.models import Project, ProjectBOM
        options = options or {}
        
        if not session.project_id:
            return {'success': False, 'error': '未指定项目'}
        
        try:
            project = Project.objects.get(id=session.project_id)
        except Project.DoesNotExist:
            return {'success': False, 'error': '项目不存在'}
        
        session.status = 'IMPORTING'
        session.save()
        
        imported, skipped, errors = 0, 0, 0
        level_stack = {}  # 用于追踪各层级的最后一个BOM项，以建立父子关系
        
        # 按行号排序处理，确保父级先于子级创建
        items = session.items.filter(status__in=['MATCHED', 'CREATED']).order_by('row_number')
        
        for item in items:
            item_id = item.matched_item_id or item.created_item_id
            if not item_id:
                item.status = 'SKIPPED'
                item.error_message = '无匹配物料'
                item.save()
                skipped += 1
                continue
            
            try:
                with transaction.atomic():
                    level = item.level or 0
                    
                    # 确定父级BOM项
                    parent_bom = None
                    if level > 0:
                        # 查找上一层级的BOM项作为父级
                        for check_level in range(level - 1, -1, -1):
                            if check_level in level_stack:
                                parent_bom = level_stack[check_level]
                                break
                    
                    # 检查是否已存在
                    existing = ProjectBOM.objects.filter(
                        project=project,
                        item_id=item_id,
                        parent=parent_bom,
                        is_deleted=False
                    ).first()
                    
                    if existing:
                        existing.planned_qty += item.quantity
                        existing.save()
                        bom = existing
                    else:
                        # 解析非标件类型
                        custom_part_type = cls._identify_custom_part_type(item)
                        is_custom = bool(custom_part_type)
                        
                        bom = ProjectBOM.objects.create(
                            project=project,
                            item_id=item_id,
                            planned_qty=item.quantity,
                            level=level,
                            parent=parent_bom,
                            # 非标件信息
                            is_custom_part=is_custom,
                            custom_part_type=custom_part_type,
                            cad_file_name=item.file_name,
                            cad_software=session.detected_software or session.cad_software,
                            # 材质信息
                            material_spec=item.material,
                            # 装配体代号
                            assembly_code=item.part_number if level == 0 else '',
                            notes=f'CAD导入: {item.file_name}'
                        )
                    
                    # 更新层级栈
                    level_stack[level] = bom
                    # 清除更深层级的引用
                    for k in list(level_stack.keys()):
                        if k > level:
                            del level_stack[k]
                    
                    item.imported_bom_id = bom.id
                    item.status = 'IMPORTED'
                    item.save()
                    imported += 1
                    
            except Exception as e:
                item.status = 'ERROR'
                item.error_message = str(e)
                item.save()
                errors += 1
                logger.exception(f"导入BOM项失败: {item.part_number}")
        
        session.imported_count = imported
        session.error_count = errors
        session.status = 'COMPLETED' if errors == 0 else 'FAILED'
        session.completed_at = timezone.now()
        session.save()
        
        return {
            'success': True,
            'imported': imported,
            'skipped': skipped,
            'errors': errors,
            'hierarchy_levels': len(level_stack)
        }
    
    @classmethod
    def _identify_custom_part_type(cls, item):
        """识别非标件类型"""
        # 自制件关键词匹配
        CUSTOM_PART_PATTERNS = {
            'MACHINED': ['机加', 'CNC', '车', '铣', '磨', '钻', '镗', 'MACHINED'],
            'SHEET_METAL': ['钣金', '折弯', '冲压', 'SHEET', '激光切割', '数冲'],
            'WELDING': ['焊接', '焊件', 'WELD', '焊合件'],
            'ASSEMBLY': ['组件', '装配', 'ASSY', 'ASM', '总成'],
            'CASTING': ['铸造', '铸件', 'CAST'],
            'FORGING': ['锻造', '锻件', 'FORG'],
            'PLASTIC': ['塑料', '注塑', 'PLASTIC', 'MOLD'],
        }
        
        # 检查文件名和零件名
        check_text = f"{item.file_name} {item.part_name} {item.part_number}".upper()
        
        for part_type, keywords in CUSTOM_PART_PATTERNS.items():
            for kw in keywords:
                if kw.upper() in check_text:
                    return part_type
        
        # 检查文件扩展名判断是否为CAD自制件
        file_ext = item.file_name.lower().split('.')[-1] if item.file_name else ''
        if file_ext in ['prt', 'sldprt', 'ipt']:
            # CAD零件文件，可能是自制件
            # 进一步根据材质判断
            material = item.material.upper() if item.material else ''
            if any(m in material for m in ['Q235', 'Q345', '45#', 'SUS', '铝', 'ALUMINUM', 'AL']):
                return 'MACHINED'
        
        return ''
    
    @classmethod
    def get_bom_tree(cls, project_id):
        """获取项目BOM树形结构"""
        from apps.projects.models import ProjectBOM
        
        def build_tree(parent_id=None, level=0):
            items = ProjectBOM.objects.filter(
                project_id=project_id,
                parent_id=parent_id,
                is_deleted=False
            ).select_related('item').order_by('level', 'sort_order')
            
            result = []
            for item in items:
                node = {
                    'id': item.id,
                    'item_id': item.item_id,
                    'item_code': item.item.sku if item.item else '',
                    'item_name': item.item.name if item.item else '',
                    'level': item.level,
                    'planned_qty': float(item.planned_qty),
                    'is_custom_part': item.is_custom_part,
                    'custom_part_type': item.custom_part_type,
                    'has_drawing': item.has_drawing,
                    'drawing_no': item.drawing_no,
                    'material_spec': item.material_spec,
                    'cad_file_name': item.cad_file_name,
                    'children': build_tree(item.id, level + 1)
                }
                result.append(node)
            return result
        
        return build_tree()


# Serializers
class CreoBOMImportItemSerializer(serializers.ModelSerializer):
    status_display = serializers.CharField(source='get_status_display', read_only=True)
    class Meta:
        model = CreoBOMImportItem
        fields = '__all__'

class CreoBOMImportSessionSerializer(serializers.ModelSerializer):
    status_display = serializers.CharField(source='get_status_display', read_only=True)
    items = CreoBOMImportItemSerializer(many=True, read_only=True)
    class Meta:
        model = CreoBOMImportSession
        fields = '__all__'

class CreoBOMUploadSerializer(serializers.Serializer):
    file = serializers.FileField()
    name = serializers.CharField(max_length=200, required=False)
    file_format = serializers.ChoiceField(choices=['CSV', 'XML', 'XLSX', 'TXT'], default='CSV')
    cad_software = serializers.ChoiceField(choices=['AUTO', 'CREO', 'SOLIDWORKS', 'AUTOCAD', 'INVENTOR', 'CATIA', 'UG', 'FUSION360', 'GENERIC'], default='AUTO')
    project_id = serializers.IntegerField(required=False, allow_null=True)
    options = serializers.JSONField(required=False, default=dict)


# ViewSet
class CreoBOMImportViewSet(SoftDeleteMixin, UserTrackingMixin, viewsets.ModelViewSet):
    queryset = CreoBOMImportSession.objects.filter(is_deleted=False)
    serializer_class = CreoBOMImportSessionSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['file_format', 'status', 'project_id']
    
    @action(detail=False, methods=['post'])
    def upload(self, request):
        """上传CAD BOM文件（支持CREO/SolidWorks/AutoCAD等自动识别）"""
        ser = CreoBOMUploadSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        f = ser.validated_data['file']
        session = CreoBOMImportSession.objects.create(
            name=ser.validated_data.get('name') or f.name,
            file_format=ser.validated_data['file_format'],
            cad_software=ser.validated_data.get('cad_software', 'AUTO'),
            project_id=ser.validated_data.get('project_id'),
            import_options=ser.validated_data.get('options', {}),
            source_file=f, source_file_name=f.name, created_by=request.user
        )
        result = CreoBOMImportService.parse_and_match(session, session.source_file.path)
        session.refresh_from_db()
        return Response({'session': CreoBOMImportSessionSerializer(session).data, 'result': result})
    
    @action(detail=True, methods=['post'])
    def create_items(self, request, pk=None):
        session = self.get_object()
        result = CreoBOMImportService.create_new_items(session, request.data.get('options', {}))
        session.refresh_from_db()
        return Response({'session': CreoBOMImportSessionSerializer(session).data, 'result': result})
    
    @action(detail=True, methods=['post'])
    def import_bom(self, request, pk=None):
        session = self.get_object()
        if request.data.get('project_id'):
            session.project_id = request.data['project_id']
            session.save()
        result = CreoBOMImportService.import_to_bom(session)
        session.refresh_from_db()
        return Response({'session': CreoBOMImportSessionSerializer(session).data, 'result': result})
    
    @action(detail=True, methods=['post'])
    def import_hierarchy(self, request, pk=None):
        """导入层级BOM（支持装配体-子装配-零件父子关系）"""
        session = self.get_object()
        if request.data.get('project_id'):
            session.project_id = request.data['project_id']
            session.save()
        result = CreoBOMImportService.import_hierarchical_bom(session, request.data.get('options', {}))
        session.refresh_from_db()
        return Response({'session': CreoBOMImportSessionSerializer(session).data, 'result': result})
    
    @action(detail=False, methods=['get'])
    def bom_tree(self, request):
        """获取项目BOM树形结构"""
        project_id = request.query_params.get('project_id')
        if not project_id:
            return Response({'error': '请指定project_id'}, status=400)
        tree = CreoBOMImportService.get_bom_tree(int(project_id))
        return Response({'tree': tree})
    
    @action(detail=True, methods=['post'])
    def manual_match(self, request, pk=None):
        session = self.get_object()
        try:
            item = session.items.get(id=request.data.get('item_id'))
            item.matched_item_id = request.data.get('matched_material_id')
            item.match_type, item.match_score, item.status = 'MANUAL', 1.0, 'MATCHED'
            item.save()
            return Response(CreoBOMImportItemSerializer(item).data)
        except CreoBOMImportItem.DoesNotExist:
            return Response({'error': '不存在'}, status=404)
    
    @action(detail=False, methods=['get'])
    def supported_formats(self, request):
        """获取支持的文件格式和CAD软件"""
        from .bom_format import BOM_FORMAT_CONFIG, get_excel_headers, PROJECT_BOM_IMPORT_TEMPLATE
        
        return Response({
            'formats': [
                {'code': 'CSV', 'name': 'CSV', 'extensions': ['.csv']},
                {'code': 'XML', 'name': 'XML', 'extensions': ['.xml']}, 
                {'code': 'XLSX', 'name': 'Excel', 'extensions': ['.xlsx', '.xls']},
                {'code': 'TXT', 'name': '文本', 'extensions': ['.txt', '.bom']}
            ],
            'cad_software': [
                {'code': 'AUTO', 'name': '自动识别', 'description': '根据BOM内容自动识别CAD软件'},
                {'code': 'CREO', 'name': 'Creo/Pro-E', 'extensions': ['.prt', '.asm', '.drw']},
                {'code': 'SOLIDWORKS', 'name': 'SolidWorks', 'extensions': ['.sldprt', '.sldasm']},
                {'code': 'AUTOCAD', 'name': 'AutoCAD', 'extensions': ['.dwg', '.dxf']},
                {'code': 'INVENTOR', 'name': 'Inventor', 'extensions': ['.ipt', '.iam']},
                {'code': 'CATIA', 'name': 'CATIA', 'extensions': ['.catpart', '.catproduct']},
                {'code': 'UG', 'name': 'UG/NX', 'extensions': ['.prt']},
                {'code': 'FUSION360', 'name': 'Fusion 360', 'extensions': ['.f3d']},
                {'code': 'GENERIC', 'name': '通用格式', 'description': '标准BOM格式'},
            ],
            'column_mapping': UNIFIED_BOM_COLUMN_MAPPING,
            'bom_fields': [f['label'] for f in PROJECT_BOM_IMPORT_TEMPLATE],
            'bom_format': BOM_FORMAT_CONFIG,
        })
    
    @action(detail=False, methods=['get'])
    def download_template(self, request):
        """下载CAD BOM导入模板（与系统BOM格式一致）"""
        from io import BytesIO
        from django.http import HttpResponse
        from .bom_format import PROJECT_BOM_IMPORT_TEMPLATE, get_excel_headers
        import pandas as pd
        
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            workbook = writer.book
            worksheet = workbook.add_worksheet('CAD_BOM导入模板')
            
            # 格式定义
            title_fmt = workbook.add_format({'bold': True, 'font_size': 14, 'font_color': '#4472C4'})
            header_fmt = workbook.add_format({'bold': True, 'bg_color': '#4472C4', 'font_color': 'white', 
                                              'border': 1, 'align': 'center', 'valign': 'vcenter'})
            required_fmt = workbook.add_format({'bold': True, 'bg_color': '#C00000', 'font_color': 'white',
                                                'border': 1, 'align': 'center', 'valign': 'vcenter'})
            example_fmt = workbook.add_format({'bg_color': '#FFF2CC', 'border': 1, 'italic': True})
            readonly_fmt = workbook.add_format({'bg_color': '#E0E0E0', 'border': 1, 'italic': True})
            
            # 标题
            worksheet.merge_range(0, 0, 0, 10, 'CAD BOM导入模板 - 非标自动化行业标准格式', title_fmt)
            worksheet.write(1, 0, '说明：红色列为必填项，灰色列为系统自动填充，黄色为示例数据')
            
            # 表头
            headers = get_excel_headers(PROJECT_BOM_IMPORT_TEMPLATE)
            for col, (label, width, htype) in enumerate(headers):
                fmt = required_fmt if htype == 'required' else header_fmt
                worksheet.write(3, col, label, fmt)
                worksheet.set_column(col, col, width)
            
            # 示例数据
            examples = [
                [1, 'MAT-001', '有图', '机加', '(自动)', '(自动)', 'V1.0', '(自动)', 10, '2026-02-01', '张三'],
                [2, 'MAT-002', '无图', '电气类', '(自动)', '(自动)', '', '(自动)', 5, '2026-02-15', '李四'],
            ]
            for row_idx, row_data in enumerate(examples):
                for col_idx, val in enumerate(row_data):
                    fmt = readonly_fmt if str(val) == '(自动)' else example_fmt
                    worksheet.write(4 + row_idx, col_idx, val, fmt)
            
            # 冻结标题行
            worksheet.freeze_panes(4, 0)
            worksheet.set_row(3, 25)
        
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=CAD_BOM_Import_Template.xlsx'
        return response